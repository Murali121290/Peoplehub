from utils.compat import Blueprint, request, jsonify, get_client_ip
from utils.jwt_helper import jwt_required, get_jwt_identity
from utils.employee_cache import get_all_employees_cached
from models.database import db
from models.attendance import Attendance
from datetime import datetime
from models.employee import Employee
from models.user import User
from datetime import date
from sqlalchemy import extract, or_, func
from sqlalchemy import extract
from sqlalchemy.orm.attributes import flag_modified
from datetime import timedelta
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from utils.compat import send_file
from zoneinfo import ZoneInfo
from models.leave import LeaveRequest, LeaveLedger
from models.payment_details import PaymentDetails
from io import BytesIO
import os
import pymysql
from middleware.auth import auth_required
from utils.jwt_helper import get_jwt_identity



from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)
from openpyxl.utils import get_column_letter


attendance_bp = Blueprint(
    "attendance",
    __name__
)


def get_ist_now():
    return datetime.now(ZoneInfo("Asia/Kolkata")).replace(tzinfo=None)


def get_ist_today():
    return datetime.now(ZoneInfo("Asia/Kolkata")).date()


@attendance_bp.route("/debug-attendance", methods=["GET"])
def debug_attendance():
    try:
        from models.attendance import Attendance
        from models.employee import Employee
        import datetime
        emp = Employee.query.filter(Employee.first_name.like("%Sangeetha%")).first()
        if emp:
            start_date = datetime.date(2026, 7, 25)
            end_date = datetime.date(2026, 8, 24)
            records = Attendance.query.filter(
                Attendance.user_id == emp.user_id,
                Attendance.attendance_date >= start_date,
                Attendance.attendance_date <= end_date
            ).all()
            details = []
            for r in records:
                details.append({
                    "date": str(r.attendance_date),
                    "type": str(type(r.attendance_date)),
                    "status": r.status
                })
            return jsonify({
                "success": True,
                "employee": f"{emp.first_name} {emp.last_name}",
                "user_id": emp.user_id,
                "employee_id": emp.employee_id,
                "records": details
            })
        return jsonify({"success": False, "error": "Employee not found"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


def is_office_network(ip_address: str) -> bool:
    import os
    if not ip_address:
        return False
    prefixes = os.environ.get("OFFICE_IP_PREFIXES", "10.,172.18.,127.0.0.1,::1").split(",")
    return any(ip_address.startswith(prefix.strip()) for prefix in prefixes if prefix.strip())


@attendance_bp.route("/checkin", methods=["POST"])
@jwt_required()
def check_in():

    try:

        data = request.json

        user_id = data.get("user_id")

        if not user_id:
            return jsonify({
                "success": False,
                "message": "User ID is required"
            }), 400

        if str(get_jwt_identity()) != str(user_id):
            return jsonify({
                "success": False,
                "message": "Unauthorized"
            }), 403

        employee = Employee.query.filter_by(
            user_id=user_id
        ).first()

        if not employee:
            return jsonify({
                "success": False,
                "message": "Employee not found"
            }), 404

        payload_ip = (data.get("client_ip") or "").strip()
        client_ip = payload_ip if payload_ip else get_client_ip()

        work_mode = (employee.work_mode or "Office").strip().lower()
        if work_mode == "office":
            if not is_office_network(client_ip):
                return jsonify({
                    "success": False,
                    "message": "Check-in is only allowed from the office network for employees in Office mode."
                }), 400


        # =====================================
        # SHIFT VALIDATION
        # =====================================
        from models.shift_request import ShiftRequest
        from sqlalchemy import or_
        today_date = get_ist_today()

        # Check if there is an approved shift request for today
        emp_ids = [employee.employee_id] if employee.employee_id else []

        approved_request = ShiftRequest.query.filter(
            ShiftRequest.employee_id.in_(emp_ids),
            ShiftRequest.status == "Approved",
            ShiftRequest.from_date <= today_date,
            ShiftRequest.to_date >= today_date
        ).first()

        if approved_request:
            shift_name = (approved_request.requested_shift or "").strip().lower()

        else:
            shift_name = (
                employee.shift_timing or ""
            ).strip().lower()

        # Use IST time for shift validation (not server UTC)
        current_time = datetime.now(ZoneInfo("Asia/Kolkata")).time()

        # First Shift (07:00 AM - 04:00 PM)
        if shift_name == "first shift":


            allowed_time = datetime.strptime(
                "07:00",
                "%H:%M"
            ).time()

            if current_time < allowed_time:

                return jsonify({
                    "success": False,
                    "message":
                    "First Shift check-in allowed only after 07:00 AM"
                }), 400

        # General Shift (09:00 AM - 06:00 PM)
        elif shift_name == "general shift":

            allowed_time = datetime.strptime(
                "09:00",
                "%H:%M"
            ).time()

            # if current_time < allowed_time:

            #     return jsonify({
            #         "success": False,
            #         "message":
            #         "General Shift check-in allowed only after 09:00 AM"
            #     }), 400

        # Second Shift (12:00 PM - 09:00 PM)
        elif shift_name == "second shift":

            allowed_time = datetime.strptime(
                "12:00",
                "%H:%M"
            ).time()

            if current_time < allowed_time:

                return jsonify({
                    "success": False,
                    "message":
                    "Second Shift check-in allowed only after 12:00 PM"
                }), 400

        # Night Shift (10:00 PM - 06:00 AM)
        elif shift_name == "night shift":

            allowed_time = datetime.strptime(
                "22:00",
                "%H:%M"
            ).time()

            if current_time < allowed_time:

                return jsonify({
                    "success": False,
                    "message":
                    "You have a night shift"
                }), 400

        # =====================================
        # APPROVED LEAVE CHECK — block check-in on approved leave days
        # =====================================
        from models.leave import LeaveRequest
        from sqlalchemy import or_ as sql_or

        approved_leave_today = LeaveRequest.query.filter(
            sql_or(
                LeaveRequest.employee_id == employee.employee_id,
                LeaveRequest.employee_id == str(employee.id)
            ),
            LeaveRequest.status == "Approved",
            LeaveRequest.request_type == "Leave",
            LeaveRequest.from_date <= today_date,
            LeaveRequest.to_date >= today_date
        ).first()

        if approved_leave_today and (approved_leave_today.total_days is None or approved_leave_today.total_days > 0.5):
            leave_type = approved_leave_today.leave_type or "Leave"
            return jsonify({
                "success": False,
                "message": f"You are on approved {leave_type} today. Check-in is not allowed on leave days.",
                "on_leave": True
            }), 403

        # =====================================
        # CHECK ALREADY CHECKED IN
        # =====================================

        today = get_ist_today()

        attendance = Attendance.query.filter_by(
            user_id=user_id,
            attendance_date=today
        ).first()

        if attendance and attendance.check_in:
            if not attendance.check_out:
                return jsonify({
                    "success": False,
                    "message": "You are already checked in."
                }), 400

            return jsonify({
                "success": False,
                "message": "You have already checked out for today. You cannot check in again."
            }), 400
        elif attendance:
            # =====================================
            # UPDATE EXISTING PLACEHOLDER ROW
            # =====================================
            attendance.check_in = get_ist_now()
            attendance.check_in_ip = client_ip
            attendance.status = "Check In"
        else:
            # =====================================
            # CREATE ATTENDANCE
            # =====================================
            attendance = Attendance(
                user_id=user_id,
                attendance_date=today,
                check_in=get_ist_now(),
                check_in_ip=client_ip,
                status="Check In"
            )

        db.session.add(attendance)

        # Update and resolve related notifications
        try:
            from models.notification import Notification
            from extensions import socketio

            # 1. Look up manager's employee details to locate their socket room
            manager_name = employee.reporting_manager.strip().lower() if employee.reporting_manager else ""
            manager_emp = None
            for e in [e for e in get_all_employees_cached() if (e.status or "").lower() != "inactive"]:
                full_name = f"{e.first_name} {e.last_name}".strip().lower()
                is_match = (full_name == manager_name) or (len(manager_name.split()) == 1 and full_name.split()[0] == manager_name) or (len(full_name.split()) == 1 and manager_name.split()[0] == full_name)
                if is_match:
                    manager_emp = e
                    break

            # 3. Mark the reminder notification for employee as completed/resolved
            reminder_notifs = Notification.query.filter(
                Notification.title == "🔔 Check-In Reminder",
                Notification.related_id == employee.id,
                Notification.resolved == False
            ).all()

            for r in reminder_notifs:
                r.resolved = True
                r.status = "Completed"
                r.resolved_at = datetime.utcnow()
                socketio.emit(
                    "manager_notification_resolved",
                    {"notification_id": r.id, "status": "Completed"},
                    to=str(employee.id)
                )


        except Exception as delete_err:
            print("Failed to process check-in notifications update:", str(delete_err))

        db.session.commit()

        # Emit attendance_update socket event for real-time dashboard updates
        try:
            from extensions import socketio
            from datetime import timedelta
            
            # Calculate virtual check_in time for frontend timer to ignore gaps
            if attendance.check_in:
                virtual_check_in = attendance.check_in + timedelta(minutes=(attendance.total_gap_minutes or 0))
                check_in_str = virtual_check_in.strftime("%I:%M %p")
            else:
                check_in_str = None
                
            payload = {
                "id": employee.id,
                "user_id": employee.user_id,
                "attendance_status": "Present",
                "check_in": check_in_str,
                "check_out": None,
                "working_hours": 0.0,
                "lunch_minutes": attendance.lunch_minutes or 0,
                "tea_minutes": attendance.tea_minutes or 0,
                "shift": employee.shift_timing or "General Shift",
                "manager_status": attendance.manager_status or "Pending",
                "checked_in": True,
                "lunch_break": False,
                "tea_break": False,
                "is_paused": False,
                "paused_start": None,
                "paused_minutes": 0
            }
            socketio.emit("attendance_update", payload)
        except Exception as socket_err:
            print("Failed to emit check-in socket:", str(socket_err))

        return jsonify({
            "success": True,
            "message":
            "Checked In Successfully"
        })

    except Exception as e:

        db.session.rollback()

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@attendance_bp.route("/checkout", methods=["POST"])
@jwt_required()
def check_out():

    try:

        data = request.json

        user_id = data.get("user_id")

        if not user_id:
            return jsonify({
                "success": False,
                "error": "User ID is required"
            }), 400

        if str(get_jwt_identity()) != str(user_id):
            return jsonify({
                "success": False,
                "error": "Unauthorized"
            }), 403

        print("CHECKOUT USER ID:", user_id)

        attendance = Attendance.query.filter(
            Attendance.user_id == user_id,
            Attendance.check_in.isnot(None),
            Attendance.check_out.is_(None)
        ).order_by(
            Attendance.id.desc()
        ).first()

        print("ATTENDANCE FOUND:", attendance)

        if not attendance:
            return jsonify({
                "success": False,
                "error": "No active check-in found"
            }), 404

        if not attendance.check_in:
            return jsonify({
                "success": False,
                "error": "Check-in time missing"
            }), 400

        attendance.check_out = get_ist_now()

        employee = Employee.query.filter_by(user_id=user_id).first()
        payload_ip = (data.get("client_ip") or "").strip()
        checkout_ip = payload_ip if payload_ip else get_client_ip()

        if employee:
            work_mode = (employee.work_mode or "Office").strip().lower()
            if work_mode == "office":
                if not is_office_network(checkout_ip):
                    return jsonify({
                        "success": False,
                        "error": "Check-out is only allowed from the office network for employees in Office mode."
                    }), 400

        attendance.check_out_ip = checkout_ip

        # IP Mismatch check (Option B) - Only enforced for Office mode employees (ignore WFH & Hybrid)
        is_office_mode = False
        if employee:
            work_mode = (employee.work_mode or "Office").strip().lower()
            if work_mode == "office":
                is_office_mode = True

        check_in_ip = attendance.check_in_ip
        if is_office_mode and check_in_ip and check_in_ip != checkout_ip:
            attendance.manager_status = "Need Clarification"
            history = list(attendance.clarification_history or [])
            msg_entry = {
                "sender_id": "system",
                "sender_name": "System",
                "sender_role": "system",
                "comment": f"IP address mismatch detected (Checked in from {check_in_ip}, Checked out from {checkout_ip}).",
                "timestamp": get_ist_now().isoformat()
            }
            history.append(msg_entry)
            attendance.clarification_history = history
            flag_modified(attendance, "clarification_history")

        if attendance.is_paused:
            attendance.is_paused = False
            paused_end = get_ist_now()
            if attendance.paused_start:
                added_mins = int((paused_end - attendance.paused_start).total_seconds() / 60)
                attendance.paused_minutes = (attendance.paused_minutes or 0) + added_mins
                attendance.paused_start = None

        total_seconds = (
            attendance.check_out -
            attendance.check_in
        ).total_seconds()

        break_minutes = (
            attendance.total_break_minutes or 0
        )
        
        gap_minutes = (
            attendance.total_gap_minutes or 0
        )

        paused_minutes = (
            attendance.paused_minutes or 0
        )

        total_seconds -= break_minutes * 60
        total_seconds -= gap_minutes * 60
        total_seconds -= paused_minutes * 60

        hours_decimal = total_seconds / 3600
        attendance.total_hours = int(hours_decimal * 100) / 100

        calculate_attendance_status(attendance)

        db.session.commit()

        # Emit attendance_update socket event for real-time dashboard updates
        try:
            from extensions import socketio
            if employee:
                payload = {
                    "id": employee.id,
                    "user_id": employee.user_id,
                    "attendance_status": "Checked Out",
                    "check_in": attendance.check_in.strftime("%I:%M %p") if attendance.check_in else None,
                    "check_out": attendance.check_out.strftime("%I:%M %p") if attendance.check_out else None,
                    "working_hours": attendance.total_hours or 0.0,
                    "lunch_minutes": attendance.lunch_minutes or 0,
                    "tea_minutes": attendance.tea_minutes or 0,
                    "shift": employee.shift_timing or "General Shift",
                    "manager_status": attendance.manager_status or "Pending",
                    "checked_in": False,
                    "lunch_break": False,
                    "tea_break": False,
                    "is_paused": False,
                    "paused_start": None,
                    "paused_minutes": attendance.paused_minutes or 0
                }
                socketio.emit("attendance_update", payload)
        except Exception as socket_err:
            print("Failed to emit checkout socket:", str(socket_err))

        is_mismatch = is_office_mode and bool(check_in_ip and check_in_ip != checkout_ip)
        message = "Checked Out Successfully (IP mismatch detected, clarification required)" if is_mismatch else "Checked Out Successfully"
        return jsonify({
            "success": True,
            "message": message,
            "check_in": attendance.check_in.strftime("%Y-%m-%d %H:%M:%S"),
            "check_out": attendance.check_out.strftime("%Y-%m-%d %H:%M:%S"),
            "total_hours": attendance.total_hours,
            "ip_mismatch": bool(is_mismatch)
        }), 200

    except Exception as e:

        db.session.rollback()

        print("CHECKOUT ERROR:", str(e))

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

def calculate_attendance_status(attendance):
    """
    Calculate and update the attendance status (Present, Half Day, Absent)
    based on the gross duration from check-in to check-out (including breaks).
    - < 4 hours: Absent
    - < 7 hours: Half Day
    - >= 7 hours: Present
    """
    # 1. Determine effective check-in and check-out (Web or Card)
    eff_in = attendance.check_in or attendance.card_check_in
    eff_out = attendance.check_out or attendance.card_check_out

    if eff_in and eff_out:
        # If checked out, calculate gross hours from in to out
        gross_seconds = (eff_out - eff_in).total_seconds()
        gross_hours = max(gross_seconds, 0) / 3600
    elif eff_in:
        # If currently checked in, calculate elapsed gross hours till now (if today)
        now = get_ist_now()
        if attendance.attendance_date == now.date():
            gross_seconds = (now - eff_in).total_seconds()
            gross_hours = max(gross_seconds, 0) / 3600
        else:
            # Past date without check-out
            gross_hours = 0.0
    else:
        gross_hours = 0.0

    # 1.5 Add Approved Permission Hours
    if attendance.user_id and attendance.attendance_date:
        try:
            from models.employee import Employee
            from models.leave_request import LeaveRequest
            from sqlalchemy import or_ as sql_or
            emp = Employee.query.filter_by(user_id=attendance.user_id).first()
            if emp:
                permission = LeaveRequest.query.filter(
                    LeaveRequest.request_type == "Permission",
                    LeaveRequest.status == "Approved",
                    LeaveRequest.permission_date == attendance.attendance_date,
                    sql_or(
                        LeaveRequest.employee_id == str(emp.id),
                        LeaveRequest.employee_id == emp.employee_id
                    )
                ).first()
                if permission and permission.from_time and permission.to_time:
                    f_time = permission.from_time
                    t_time = permission.to_time
                    f_sec = f_time.hour * 3600 + f_time.minute * 60 + f_time.second
                    t_sec = t_time.hour * 3600 + t_time.minute * 60 + t_time.second
                    permission_hours = max(t_sec - f_sec, 0) / 3600.0
                    gross_hours += permission_hours
        except Exception as e:
            print("Error calculating permission hours in attendance status:", e)

    # 2. Determine status
    if eff_in and not eff_out:
        # Checked in but not yet checked out
        now = get_ist_now()
        if attendance.attendance_date == now.date():
            # Still live session today — mark as Check In
            attendance.status = "Check In"
            return
        # Past date without checkout — treat as Absent
        attendance.status = "Absent"
        return

    if gross_hours < 4.0:
        attendance.status = "Absent"
    else:
        is_weekend = attendance.attendance_date.weekday() >= 5
        req_hours = 7.0 if is_weekend else 8.0
        
        if gross_hours < req_hours:
            attendance.status = "Half Day"
        else:
            attendance.status = "Present"


def sync_biometric_to_web_entry(attendance):
    """
    Sync biometric card punch times to web punch times if they are missing,
    recalculate total hours & status, and return True if any changes were made.
    """
    updated = False
    if attendance.card_check_in and not attendance.check_in:
        attendance.check_in = attendance.card_check_in
        updated = True
    if attendance.card_check_out and not attendance.check_out:
        attendance.check_out = attendance.card_check_out
        updated = True

    if updated:
        if attendance.check_in and attendance.check_out:
            total_seconds = (attendance.check_out - attendance.check_in).total_seconds()
            break_minutes = attendance.total_break_minutes or 0
            gap_minutes = attendance.total_gap_minutes or 0
            paused_minutes = attendance.paused_minutes or 0
            total_seconds -= (break_minutes + gap_minutes + paused_minutes) * 60
            hours_decimal = max(total_seconds, 0) / 3600
            attendance.total_hours = int(hours_decimal * 100) / 100

        # Recalculate status
        calculate_attendance_status(attendance)

    return updated


@attendance_bp.route("/sync-logs", methods=["POST"])
def sync_card_logs():
    try:
        data = request.json
        logs = data.get("logs", [])
        
        if not logs:
            return jsonify({
                "success": True,
                "message": "No logs provided"
            }), 200

        print(f"Syncing {len(logs)} card logs...")
        processed_count = 0
        batch_attendance = {}

        for log in logs:
            emp_code = str(log.get("employee_code", "")).strip()
            log_time_str = log.get("log_time")
            direction = str(log.get("direction", "")).strip().lower()

            if not emp_code or not log_time_str:
                continue

            # Parse log time
            try:
                # Support various formats, standard is %Y-%m-%d %H:%M:%S
                if "T" in log_time_str:
                    log_dt = datetime.fromisoformat(log_time_str.replace("Z", ""))
                else:
                    log_dt = datetime.strptime(log_time_str, "%Y-%m-%d %H:%M:%S")
            except Exception as pe:
                print(f"Failed to parse log time '{log_time_str}': {pe}")
                continue

            # Find employee by employee_id (Postgres)
            employee = Employee.query.filter_by(employee_id=emp_code).first()
            if not employee:
                print(f"Sync skip: Employee code '{emp_code}' not found in Peoplehub database")
                continue

            att_date = log_dt.date()
            cache_key = (employee.user_id, att_date)

            # Find or create Attendance for that employee and date
            if cache_key in batch_attendance:
                attendance = batch_attendance[cache_key]
            else:
                attendance = Attendance.query.filter_by(
                    user_id=employee.user_id,
                    attendance_date=att_date
                ).first()

                if not attendance:
                    attendance = Attendance(
                        user_id=employee.user_id,
                        attendance_date=att_date,
                        status="Present",
                        shift_timing=employee.shift_timing or "General Shift"
                    )
                    db.session.add(attendance)
                batch_attendance[cache_key] = attendance

            # Update Card Punch times
            if direction == "in":
                if not attendance.card_check_in:
                    attendance.card_check_in = log_dt
                else:
                    # Keep earliest punch for check-in
                    attendance.card_check_in = min(attendance.card_check_in, log_dt)
            elif direction == "out":
                if not attendance.card_check_out:
                    attendance.card_check_out = log_dt
                else:
                    # Keep latest punch for check-out
                    attendance.card_check_out = max(attendance.card_check_out, log_dt)
            else:
                # If direction is not specified, auto-determine based on order of punches
                if not attendance.card_check_in:
                    attendance.card_check_in = log_dt
                else:
                    # Compare and set as check-in or check-out
                    if log_dt < attendance.card_check_in:
                        attendance.card_check_out = attendance.card_check_in
                        attendance.card_check_in = log_dt
                    else:
                        if not attendance.card_check_out:
                            attendance.card_check_out = log_dt
                        else:
                            attendance.card_check_out = max(attendance.card_check_out, log_dt)

            # Calculate card working hours
            if attendance.card_check_in and attendance.card_check_out:
                total_seconds = (attendance.card_check_out - attendance.card_check_in).total_seconds()
                hours_decimal = max(total_seconds, 0) / 3600
                attendance.card_working_hours = int(hours_decimal * 100) / 100
            else:
                attendance.card_working_hours = 0.0

            # Sync to web check_in / check_out columns if they are NULL
            sync_biometric_to_web_entry(attendance)

            # Recalculate status
            calculate_attendance_status(attendance)

            processed_count += 1

            # Emit dashboard update
            try:
                from extensions import socketio
                
                web_in_str = attendance.check_in.strftime("%I:%M %p") if attendance.check_in else None
                web_out_str = attendance.check_out.strftime("%I:%M %p") if attendance.check_out else None
                card_in_str = attendance.card_check_in.strftime("%I:%M %p") if attendance.card_check_in else None
                card_out_str = attendance.card_check_out.strftime("%I:%M %p") if attendance.card_check_out else None

                payload = {
                    "id": employee.id,
                    "user_id": employee.user_id,
                    "attendance_status": attendance.status,
                    "check_in": web_in_str,
                    "check_out": web_out_str,
                    "working_hours": attendance.total_hours or 0.0,
                    "card_check_in": card_in_str,
                    "card_check_out": card_out_str,
                    "card_working_hours": attendance.card_working_hours or 0.0,
                    "shift": employee.shift_timing or "General Shift",
                    "manager_status": attendance.manager_status or "Pending",
                    "checked_in": (attendance.check_in is not None and attendance.check_out is None),
                    "card_checked_in": (attendance.card_check_in is not None and attendance.card_check_out is None),
                }
                socketio.emit("attendance_update", payload)
            except Exception as se:
                print("Socket emit error during card sync:", se)

        db.session.commit()
        return jsonify({
            "success": True,
            "message": f"Successfully processed {processed_count} logs"
        }), 200

    except Exception as e:
        db.session.rollback()
        print("SYNC ERROR:", str(e))
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@attendance_bp.route("/status/<int:user_id>")
def attendance_status(user_id):

    attendance = Attendance.query.filter_by(
        user_id=user_id,
        attendance_date=get_ist_today()
    ).order_by(
        Attendance.id.desc()
    ).first()

    if not attendance:
        return jsonify({
            "checked_in": False,
            "checked_out": False
        })

    # User is checked in if they have check_in/card_check_in and haven't checked out on those channels
    web_active = (attendance.check_in is not None) and (attendance.check_out is None)
    card_active = (attendance.card_check_in is not None) and (attendance.card_check_out is None)
    is_checked_in = web_active or card_active

    has_checked_out = (attendance.check_in is not None and attendance.check_out is not None) or \
                      (attendance.card_check_in is not None and attendance.card_check_out is not None)

    effective_check_in = attendance.check_in or attendance.card_check_in

    return jsonify({
        "checked_in": is_checked_in,
        "checked_out": has_checked_out,
        "check_in": effective_check_in.isoformat() if effective_check_in else None,
        "check_out": (attendance.check_out or attendance.card_check_out).isoformat() if (attendance.check_out or attendance.card_check_out) else None,
        "lunch_break": attendance.lunch_break,
        "tea_break": attendance.tea_break,
        "lunch_start": attendance.lunch_start.isoformat() if attendance.lunch_start else None,
        "tea_start": attendance.tea_start.isoformat() if attendance.tea_start else None,
        "lunch_minutes": attendance.lunch_minutes or 0,
        "tea_minutes": attendance.tea_minutes or 0,
        "total_break_minutes": attendance.total_break_minutes or 0,
        "is_paused": attendance.is_paused or False,
        "paused_start": attendance.paused_start.isoformat() if attendance.paused_start else None,
        "paused_minutes": attendance.paused_minutes or 0,
        "working_hours": attendance.total_hours or 0.0
    })




@attendance_bp.route(
    "/lunch-break",
    methods=["POST", "PUT"]
)
@jwt_required()
def lunch_break():

    try:

        data = request.json

        if str(get_jwt_identity()) != str(data.get("user_id")):
            return jsonify({
                "success": False,
                "error": "Unauthorized"
            }), 403

        employee = Employee.query.filter_by(user_id=data.get("user_id")).first()
        if employee:
            work_mode = (employee.work_mode or "").strip().lower()
            if work_mode == "office":
                payload_ip = (data.get("client_ip") or "").strip()
                client_ip = payload_ip if payload_ip else get_client_ip()
                if not is_office_network(client_ip):
                    return jsonify({
                        "success": False,
                        "error": "Lunch break actions are only allowed from the office network for employees in Office mode."
                    }), 400

        attendance = Attendance.query.filter_by(
            user_id=data["user_id"],
            attendance_date=get_ist_today()
        ).order_by(
            Attendance.id.desc()
        ).first()

        if not attendance:
            return jsonify({
                "success": False,
                "error": "Attendance not found"
            }), 404

        # Verify they are currently checked in (web active or card active)
        web_active = (attendance.check_in is not None) and (attendance.check_out is None)
        card_active = (attendance.card_check_in is not None) and (attendance.card_check_out is None)
        if not (web_active or card_active):
            return jsonify({
                "success": False,
                "error": "User is not checked in or has already checked out."
            }), 400

        action = data.get("action")

        if action == "start":
            # Lunch is allowed only once per day — check existing lunch_minutes (no new column needed)
            if (attendance.lunch_minutes or 0) > 0 or attendance.lunch_break:
                return jsonify({
                    "success": False,
                    "error": "Lunch break is allowed only once per day."
                }), 400

            attendance.lunch_break = True
            attendance.lunch_start = get_ist_now()

        elif action == "stop":
            attendance.lunch_break = False
            attendance.lunch_end = get_ist_now()

            if attendance.lunch_start and attendance.lunch_end:
                added_mins = int((attendance.lunch_end - attendance.lunch_start).total_seconds() / 60)
                attendance.lunch_minutes = (attendance.lunch_minutes or 0) + added_mins
                attendance.lunch_start = None
                attendance.lunch_end = None

        attendance.total_break_minutes = (
            (attendance.lunch_minutes or 0) +
            (attendance.tea_minutes or 0)
        )

        db.session.commit()

        # Emit attendance_update socket event for real-time dashboard updates
        try:
            from extensions import socketio
            if employee:
                payload = {
                    "id": employee.id,
                    "user_id": employee.user_id,
                    "attendance_status": "Present" if not attendance.check_out else "Checked Out",
                    "check_in": attendance.check_in.strftime("%I:%M %p") if attendance.check_in else None,
                    "check_out": attendance.check_out.strftime("%I:%M %p") if attendance.check_out else None,
                    "working_hours": attendance.total_hours or 0.0,
                    "lunch_minutes": attendance.lunch_minutes or 0,
                    "tea_minutes": attendance.tea_minutes or 0,
                    "shift": employee.shift_timing or "General Shift",
                    "manager_status": attendance.manager_status or "Pending",
                    "checked_in": True if not attendance.check_out else False,
                    "lunch_break": attendance.lunch_break or False,
                    "tea_break": attendance.tea_break or False,
                    "lunch_start": attendance.lunch_start.isoformat() if attendance.lunch_start else None,
                    "lunch_end": attendance.lunch_end.isoformat() if attendance.lunch_end else None,
                    "is_paused": attendance.is_paused or False,
                    "paused_start": attendance.paused_start.isoformat() if attendance.paused_start else None,
                    "paused_minutes": attendance.paused_minutes or 0
                }
                socketio.emit("attendance_update", payload)
        except Exception as socket_err:
            print("Failed to emit lunch socket:", str(socket_err))

        return jsonify({
            "success": True
        })

    except Exception as e:

        print("LUNCH BREAK ERROR:", str(e))

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@attendance_bp.route(
    "/tea-break",
    methods=["POST", "PUT"]
)
@jwt_required()
def tea_break():

    try:

        data = request.json

        if str(get_jwt_identity()) != str(data.get("user_id")):
            return jsonify({
                "success": False,
                "error": "Unauthorized"
            }), 403

        employee = Employee.query.filter_by(user_id=data.get("user_id")).first()
        if employee:
            work_mode = (employee.work_mode or "").strip().lower()
            if work_mode == "office":
                payload_ip = (data.get("client_ip") or "").strip()
                client_ip = payload_ip if payload_ip else get_client_ip()
                if not is_office_network(client_ip):
                    return jsonify({
                        "success": False,
                        "error": "Tea break actions are only allowed from the office network for employees in Office mode."
                    }), 400

        attendance = Attendance.query.filter_by(
            user_id=data["user_id"],
            attendance_date=get_ist_today()
        ).order_by(
            Attendance.id.desc()
        ).first()

        if not attendance:
            return jsonify({
                "success": False,
                "error": "Attendance not found"
            }), 404

        # Verify they are currently checked in (web active or card active)
        web_active = (attendance.check_in is not None) and (attendance.check_out is None)
        card_active = (attendance.card_check_in is not None) and (attendance.card_check_out is None)
        if not (web_active or card_active):
            return jsonify({
                "success": False,
                "error": "User is not checked in or has already checked out."
            }), 400

        action = data.get("action")

        if action == "start":
            # Tea is allowed twice per day — read count via raw SQL (column exists in DB, not in ORM model)
            from sqlalchemy import text as sql_text
            tea_count_row = db.session.execute(
                sql_text("SELECT tea_count FROM attendance WHERE id = :id"),
                {"id": attendance.id}
            ).fetchone()
            tea_count = (tea_count_row[0] or 0) if tea_count_row else 0

            if tea_count >= 2:
                return jsonify({
                    "success": False,
                    "error": "Tea break is allowed only twice per day."
                }), 400

            attendance.tea_break = True
            attendance.tea_start = get_ist_now()
            # Increment tea_count via raw SQL
            db.session.execute(
                sql_text("UPDATE attendance SET tea_count = :count WHERE id = :id"),
                {"count": tea_count + 1, "id": attendance.id}
            )

        elif action == "stop":
            attendance.tea_break = False
            attendance.tea_end = get_ist_now()

            if attendance.tea_start and attendance.tea_end:
                added_mins = int((attendance.tea_end - attendance.tea_start).total_seconds() / 60)
                attendance.tea_minutes = (attendance.tea_minutes or 0) + added_mins
                attendance.tea_start = None
                attendance.tea_end = None

        attendance.total_break_minutes = (
            (attendance.lunch_minutes or 0) +
            (attendance.tea_minutes or 0)
        )

        db.session.commit()

        # Emit attendance_update socket event for real-time dashboard updates
        try:
            from extensions import socketio
            if employee:
                payload = {
                    "id": employee.id,
                    "user_id": employee.user_id,
                    "attendance_status": "Present" if not attendance.check_out else "Checked Out",
                    "check_in": attendance.check_in.strftime("%I:%M %p") if attendance.check_in else None,
                    "check_out": attendance.check_out.strftime("%I:%M %p") if attendance.check_out else None,
                    "working_hours": attendance.total_hours or 0.0,
                    "lunch_minutes": attendance.lunch_minutes or 0,
                    "tea_minutes": attendance.tea_minutes or 0,
                    "shift": employee.shift_timing or "General Shift",
                    "manager_status": attendance.manager_status or "Pending",
                    "checked_in": True if not attendance.check_out else False,
                    "lunch_break": attendance.lunch_break or False,
                    "tea_break": attendance.tea_break or False,
                    "tea_start": attendance.tea_start.isoformat() if attendance.tea_start else None,
                    "tea_end": attendance.tea_end.isoformat() if attendance.tea_end else None,
                    "is_paused": attendance.is_paused or False,
                    "paused_start": attendance.paused_start.isoformat() if attendance.paused_start else None,
                    "paused_minutes": attendance.paused_minutes or 0
                }
                socketio.emit("attendance_update", payload)
        except Exception as socket_err:
            print("Failed to emit tea socket:", str(socket_err))

        return jsonify({
            "success": True
        })

    except Exception as e:

        print("TEA BREAK ERROR:", str(e))

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@attendance_bp.route(
    "/pause",
    methods=["POST", "PUT"]
)
@jwt_required()
def pause_attendance():

    try:

        data = request.json

        if str(get_jwt_identity()) != str(data.get("user_id")):
            return jsonify({
                "success": False,
                "error": "Unauthorized"
            }), 403

        employee = Employee.query.filter_by(user_id=data.get("user_id")).first()
        if employee:
            work_mode = (employee.work_mode or "").strip().lower()
            if work_mode == "office":
                payload_ip = (data.get("client_ip") or "").strip()
                client_ip = payload_ip if payload_ip else get_client_ip()
                if not is_office_network(client_ip):
                    return jsonify({
                        "success": False,
                        "error": "Timesheet pause/resume is only allowed from the office network for employees in Office mode."
                    }), 400

        attendance = Attendance.query.filter_by(
            user_id=data["user_id"],
            attendance_date=get_ist_today()
        ).order_by(
            Attendance.id.desc()
        ).first()

        if not attendance:
            return jsonify({
                "success": False,
                "error": "Attendance not found"
            }), 404

        # Verify they are currently checked in (web active or card active)
        web_active = (attendance.check_in is not None) and (attendance.check_out is None)
        card_active = (attendance.card_check_in is not None) and (attendance.card_check_out is None)
        if not (web_active or card_active):
            return jsonify({
                "success": False,
                "error": "User is not checked in or has already checked out."
            }), 400

        action = data.get("action")

        if action == "start":
            if attendance.lunch_break or attendance.tea_break:
                return jsonify({
                    "success": False,
                    "error": "Cannot pause while on active break (lunch/tea)."
                }), 400
            if attendance.is_paused:
                return jsonify({
                    "success": False,
                    "error": "Timesheet is already paused."
                }), 400

            attendance.is_paused = True
            attendance.paused_start = get_ist_now()

        elif action == "stop":
            if not attendance.is_paused:
                return jsonify({
                    "success": False,
                    "error": "Timesheet is not currently paused."
                }), 400

            attendance.is_paused = False
            paused_end = get_ist_now()

            if attendance.paused_start:
                added_mins = int((paused_end - attendance.paused_start).total_seconds() / 60)
                attendance.paused_minutes = (attendance.paused_minutes or 0) + added_mins
                attendance.paused_start = None

        db.session.commit()

        # Emit attendance_update socket event for real-time dashboard updates
        try:
            from extensions import socketio
            if employee:
                payload = {
                    "id": employee.id,
                    "user_id": employee.user_id,
                    "attendance_status": "Present" if not attendance.check_out else "Checked Out",
                    "check_in": attendance.check_in.strftime("%I:%M %p") if attendance.check_in else None,
                    "check_out": attendance.check_out.strftime("%I:%M %p") if attendance.check_out else None,
                    "working_hours": attendance.total_hours or 0.0,
                    "lunch_minutes": attendance.lunch_minutes or 0,
                    "tea_minutes": attendance.tea_minutes or 0,
                    "paused_minutes": attendance.paused_minutes or 0,
                    "shift": employee.shift_timing or "General Shift",
                    "manager_status": attendance.manager_status or "Pending",
                    "checked_in": True if not attendance.check_out else False,
                    "lunch_break": attendance.lunch_break or False,
                    "tea_break": attendance.tea_break or False,
                    "is_paused": attendance.is_paused or False,
                    "paused_start": attendance.paused_start.isoformat() if attendance.paused_start else None
                }
                socketio.emit("attendance_update", payload)
        except Exception as socket_err:
            print("Failed to emit pause socket:", str(socket_err))

        return jsonify({
            "success": True,
            "is_paused": attendance.is_paused,
            "paused_minutes": attendance.paused_minutes
        })

    except Exception as e:

        print("PAUSE ATTENDANCE ERROR:", str(e))

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

def is_date_week_off(d):
    """
    S4Carlisle week-off rules:
    - Sundays (weekday == 6) are week-offs
    - 2nd and 4th Saturdays of the month are week-offs
    - 1st, 3rd, and 5th Saturdays are working days
    """
    if d.weekday() == 6:
        return True
    if d.weekday() == 5:
        # Calculate which Saturday of the month it is
        sat_count = 0
        for day_num in range(1, d.day + 1):
            from datetime import date
            if date(d.year, d.month, day_num).weekday() == 5:
                sat_count += 1
        if sat_count in (2, 4):
            return True
    return False

@attendance_bp.route("/check-holiday-or-weekoff", methods=["GET"])
@jwt_required()
def check_holiday_or_weekoff():
    try:
        from models.employee import Employee
        user_id = get_jwt_identity()
        employee = Employee.query.filter_by(user_id=int(user_id)).first()
        if not employee:
            return jsonify({"success": False, "message": "Employee not found"}), 404
            
        today = get_ist_today()
        
        # Check holiday override
        from models.holiday import Holiday, HolidayOverride
        override = HolidayOverride.query.filter_by(date=today).first()
        is_holiday = False
        is_week_off = False
        holiday_name = None
        
        if override:
            if override.override_type == "Holiday":
                is_holiday = True
                holiday_name = override.name or "Holiday"
            elif override.override_type == "Weekly Off":
                is_week_off = True
                holiday_name = override.name or "Weekly Off"
        else:
            holiday = Holiday.query.filter_by(date=today).first()
            if holiday:
                is_holiday = True
                holiday_name = holiday.name
            elif is_date_week_off(today):
                is_week_off = True
                
        # Check if already has a pending or approved one day wages request for today
        from models.shift_request import ShiftRequest
        existing_wages = ShiftRequest.query.filter(
            ShiftRequest.employee_id.in_([employee.id, employee.employee_id]),
            ShiftRequest.request_type == "One Day Wages",
            ShiftRequest.status.in_(["Pending", "Approved"]),
            ShiftRequest.from_date <= today,
            ShiftRequest.to_date >= today
        ).first()
        
        return jsonify({
            "success": True,
            "is_holiday_or_weekoff": is_holiday or is_week_off,
            "reason": holiday_name or ("Weekend Weekoff" if is_week_off else None),
            "already_requested": existing_wages is not None
        }), 200
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@attendance_bp.route("/history/<int:user_id>")
def attendance_history(user_id):

    employee = Employee.query.filter_by(user_id=user_id).first()
    if not employee:
        return jsonify([])

    today = get_ist_today()
    joining = employee.joining_date
    result = []

    # Read start_date from query parameters (format: YYYY-MM-DD)
    start_date_param = request.args.get("start_date")
    start_date_val = None
    if start_date_param:
        try:
            start_date_val = datetime.strptime(start_date_param, "%Y-%m-%d").date()
        except ValueError:
            start_date_val = None

    if not start_date_val:
        # Default to current cycle start date
        if today.day >= 25:
            start_date_val = date(today.year, today.month, 25)
        else:
            if today.month == 1:
                start_date_val = date(today.year - 1, 12, 25)
            else:
                start_date_val = date(today.year, today.month - 1, 25)

    # Cycle runs from start_date_val to 24th of next month
    if start_date_val.month == 12:
        end_date_limit = date(start_date_val.year + 1, 1, 24)
    else:
        end_date_limit = date(start_date_val.year, start_date_val.month + 1, 24)

    # Include today's date in the history data
    end_date = min(today, end_date_limit)

    # Fetch holidays and overrides within range
    from models.holiday import Holiday, HolidayOverride
    holidays = Holiday.query.filter(Holiday.date >= start_date_val, Holiday.date <= end_date).all()
    overrides = HolidayOverride.query.filter(HolidayOverride.date >= start_date_val, HolidayOverride.date <= end_date).all()
    
    holiday_dict = {
        (h.date.date() if hasattr(h.date, "date") else h.date): h.name
        for h in holidays
    }
    override_dict = {
        (o.date.date() if hasattr(o.date, "date") else o.date): (o.override_type, o.name)
        for o in overrides
    }

    if end_date >= start_date_val:
        num_days = (end_date - start_date_val).days + 1
        for i in range(num_days):
            current_date = end_date - timedelta(days=i)

            # Do not go before joining date if set
            if joining and current_date < joining:
                break

            record = Attendance.query.filter_by(
                user_id=user_id,
                attendance_date=current_date
            ).first()

            # Check for One Day Wages request on this date
            from models.shift_request import ShiftRequest
            wages_req = ShiftRequest.query.filter(
                ShiftRequest.employee_id.in_([employee.id, employee.employee_id]),
                ShiftRequest.request_type == "One Day Wages",
                ShiftRequest.from_date <= current_date,
                ShiftRequest.to_date >= current_date
            ).first()
            wages_status = wages_req.status if wages_req else None
            is_weekend_or_holiday = is_date_week_off(current_date) or (current_date in holiday_dict)
            is_one_day_wages = (wages_req is not None and wages_req.status == "Approved")

            # Determine the effective shift for this date:
            # Priority: approved ShiftRequest covering date > record.shift_timing > employee.shift_timing
            approved_shift_req = ShiftRequest.query.filter(
                ShiftRequest.employee_id.in_([employee.id, employee.employee_id]),
                ShiftRequest.request_type == "Shift",
                ShiftRequest.status == "Approved",
                ShiftRequest.from_date <= current_date,
                ShiftRequest.to_date >= current_date
            ).order_by(ShiftRequest.created_at.desc()).first()
            effective_shift = (
                (approved_shift_req.requested_shift if approved_shift_req and approved_shift_req.requested_shift else None)
                or (record.shift_timing if record else None)
                or employee.shift_timing
                or "General Shift"
            )

            if record:
                is_today = (record.attendance_date == today)
                
                if record.check_out:
                    working_hours = record.total_hours or 0.0
                    check_out_str = record.check_out.strftime("%I:%M %p")
                elif is_today and record.check_in:
                    # Active check-in today: compute live working hours
                    now = get_ist_now()
                    elapsed_seconds = (now - record.check_in).total_seconds()
                    break_seconds = (record.total_break_minutes or 0) * 60
                    hours_decimal = max(elapsed_seconds - break_seconds, 0) / 3600
                    working_hours = int(hours_decimal * 100) / 100
                    check_out_str = "-"
                else:
                    # Past date with missing checkout
                    working_hours = record.total_hours or 0.0
                    check_out_str = "-"

                # Calculate gross total hours (including breaks)
                eff_in = record.check_in or record.card_check_in
                eff_out = record.check_out or record.card_check_out
                if eff_in and eff_out:
                    gross_sec = (eff_out - eff_in).total_seconds()
                    gross_hours = max(gross_sec, 0) / 3600
                elif eff_in and is_today:
                    now_time = get_ist_now()
                    gross_sec = (now_time - eff_in).total_seconds()
                    gross_hours = max(gross_sec, 0) / 3600
                else:
                    gross_hours = 0.0
                gross_hours = int(gross_hours * 100) / 100

                # Derive display status: strictly use database status directly
                display_status = record.status
                if not display_status:
                    if gross_hours < 4.0:
                        display_status = "Absent"
                    elif gross_hours < 7.0:
                        display_status = "Half Day"
                    else:
                        display_status = "Present"

                # If employee has checked in but not checked out today, keep 'Check In'
                if record.check_in and not record.check_out and not record.card_check_out and is_today:
                    display_status = "Check In"

                # Override: if the day is a weekend or holiday and no check-in exists,
                # show the correct label instead of a stale "Absent" stored in DB.
                if not record.check_in:
                    override = override_dict.get(current_date)
                    if override:
                        if override[0] == "Holiday":
                            display_status = "Holiday"
                        # If override is "Working Day", keep whatever was computed
                    elif current_date in holiday_dict:
                        display_status = "Holiday"
                    elif is_date_week_off(current_date):
                        display_status = "Week Off"

                # Check for approved Permission on this date
                from models.leave import LeaveRequest as LR
                perm_req = LR.query.filter(
                    or_(
                        LR.employee_id == str(employee.id),
                        LR.employee_id == employee.employee_id
                    ),
                    LR.request_type == "Permission",
                    LR.status == "Approved",
                    LR.permission_date == current_date
                ).first()

                has_permission = False
                permission_label = ""
                actual_working_hours = working_hours
                if perm_req and perm_req.from_time and perm_req.to_time:
                    has_permission = True
                    ft = perm_req.from_time
                    tt = perm_req.to_time
                    perm_seconds = (tt.hour * 3600 + tt.minute * 60) - (ft.hour * 3600 + ft.minute * 60)
                    perm_hours = max(perm_seconds, 0) / 3600
                    virtual_working_hours = actual_working_hours + perm_hours

                    def _fmt(t):
                        h = t.hour; ampm = "AM" if h < 12 else "PM"; h12 = h % 12 or 12
                        return f"{h12:02d}:{t.minute:02d} {ampm}"

                    permission_label = f"{_fmt(ft)} – {_fmt(tt)}"



                result.append({
                    "id": record.id,
                    "date": record.attendance_date.strftime("%Y-%m-%d"),
                    "attendance_date": record.attendance_date.strftime("%Y-%m-%d"),
                    "attendance_date_formatted": record.attendance_date.strftime("%d %b %Y"),
                    "shift": effective_shift,
                    "checkIn": record.check_in.strftime("%I:%M %p") if record.check_in else "-",
                    "check_in": record.check_in.strftime("%I:%M %p") if record.check_in else "-",
                    "checkOut": check_out_str,
                    "check_out": check_out_str,
                    "workingHours": working_hours,
                    "working_hours": working_hours,
                    "total_hours": gross_hours,
                    "totalHours": gross_hours,
                    "cardCheckIn": record.card_check_in.strftime("%I:%M %p") if record.card_check_in else "-",
                    "card_check_in": record.card_check_in.strftime("%I:%M %p") if record.card_check_in else "-",
                    "cardCheckOut": record.card_check_out.strftime("%I:%M %p") if record.card_check_out else "-",
                    "card_check_out": record.card_check_out.strftime("%I:%M %p") if record.card_check_out else "-",
                    "cardWorkingHours": record.card_working_hours or 0.0,
                    "card_working_hours": record.card_working_hours or 0.0,
                    "lunchMinutes": record.lunch_minutes,
                    "lunch_minutes": record.lunch_minutes,
                    "teaMinutes": record.tea_minutes,
                    "tea_minutes": record.tea_minutes,
                    "totalBreak": record.total_break_minutes,
                    "total_break_minutes": record.total_break_minutes,
                    "status": display_status,
                    "manager_status": record.manager_status or "Pending",
                    "reporting_manager": employee.reporting_manager or "",
                    "check_in_ip": record.check_in_ip,
                    "check_out_ip": record.check_out_ip,
                    "clarification_history": record.clarification_history or [],
                    "last_manager_comment": (record.clarification_history[-1]["comment"] if record.clarification_history and record.clarification_history[-1].get("sender_role") == "manager" else "") or "",
                    "is_one_day_wages": is_one_day_wages,
                    "wages_status": wages_status,
                    "has_permission": has_permission,
                    "permission_label": permission_label,
                    "is_regularization": bool(record.is_regularization),
                    "regularization_reason": record.regularization_reason or "",
                    "regularization_check_in": record.regularization_check_in.strftime("%I:%M %p") if record.regularization_check_in else "-",
                    "regularization_check_out": record.regularization_check_out.strftime("%I:%M %p") if record.regularization_check_out else "-",
                    "regularization_total_hours": record.regularization_total_hours or 0.0,
                })

            else:
                # Check normal calendar rules for virtual status
                status = "Absent"
                
                # 1. Check HolidayOverride first
                override = override_dict.get(current_date)
                if override:
                    if override[0] == "Holiday":
                        status = "Holiday"
                    else:
                        # override is "Working Day" -> check approved leaves or mark Absent
                        from models.leave import LeaveRequest
                        leave = LeaveRequest.query.filter(
                            or_(
                                LeaveRequest.employee_id == str(employee.id),
                                LeaveRequest.employee_id == employee.employee_id
                            ),
                            LeaveRequest.status == "Approved",
                            LeaveRequest.from_date <= current_date,
                            LeaveRequest.to_date >= current_date
                        ).first()
                        
                        is_cancelled = False
                        if leave and leave.cancelled_dates:
                            is_cancelled = current_date.strftime("%Y-%m-%d") in leave.cancelled_dates
                            
                        status = ("Half Day" if (leave.total_days is not None and leave.total_days <= 0.5) else "Leave") if (leave and not is_cancelled) else "Absent"
                else:
                    # 2. Check Holiday table
                    if current_date in holiday_dict:
                        status = "Holiday"
                    # 3. Check Saturday/Sunday week-off logic
                    elif is_date_week_off(current_date):
                        status = "Week Off"
                    # 4. Check approved leaves
                    else:
                        from models.leave import LeaveRequest
                        leave = LeaveRequest.query.filter(
                            or_(
                                LeaveRequest.employee_id == str(employee.id),
                                LeaveRequest.employee_id == employee.employee_id
                            ),
                            LeaveRequest.status == "Approved",
                            LeaveRequest.from_date <= current_date,
                            LeaveRequest.to_date >= current_date
                        ).first()
                        
                        is_cancelled = False
                        if leave and leave.cancelled_dates:
                            is_cancelled = current_date.strftime("%Y-%m-%d") in leave.cancelled_dates
                            
                        status = ("Half Day" if (leave.total_days is not None and leave.total_days <= 0.5) else "Leave") if (leave and not is_cancelled) else "Absent"

                result.append({
                    "id": f"virtual-{current_date.strftime('%Y-%m-%d')}",
                    "date": current_date.strftime("%Y-%m-%d"),
                    "attendance_date": current_date.strftime("%Y-%m-%d"),
                    "attendance_date_formatted": current_date.strftime("%d %b %Y"),
                    "shift": employee.shift_timing or "General Shift",
                    "checkIn": "-",
                    "check_in": "-",
                    "checkOut": "-",
                    "check_out": "-",
                    "workingHours": 0.0,
                    "working_hours": 0.0,
                    "total_hours": 0.0,
                    "cardCheckIn": "-",
                    "card_check_in": "-",
                    "cardCheckOut": "-",
                    "card_check_out": "-",
                    "cardWorkingHours": 0.0,
                    "card_working_hours": 0.0,
                    "lunchMinutes": 0,
                    "lunch_minutes": 0,
                    "teaMinutes": 0,
                    "tea_minutes": 0,
                    "totalBreak": 0,
                    "total_break_minutes": 0,
                    "status": status,
                    "manager_status": "Pending",
                    "reporting_manager": employee.reporting_manager or "",
                    "clarification_history": [],
                    "last_manager_comment": "",
                    "is_one_day_wages": is_one_day_wages,
                    "wages_status": wages_status,
                    "regularization_check_in": "-",
                    "regularization_check_out": "-",
                    "regularization_total_hours": 0.0
                })

    return jsonify(result)


@attendance_bp.route("/", methods=["GET"])
def get_attendance():

    # Allow HR/Managers to query a specific date via ?date=YYYY-MM-DD
    date_param = request.args.get("date")
    if date_param:
        try:
            today = datetime.strptime(date_param, "%Y-%m-%d").date()
        except ValueError:
            today = get_ist_today()
    else:
        today = get_ist_today()

    employees = [e for e in get_all_employees_cached() if (e.status or "").lower() != "inactive"]

    attendance_list = []

    for employee in employees:
        # Skip if date is prior to date of joining (D.O.J)
        if employee.joining_date and today < employee.joining_date:
            continue

        attendance = Attendance.query.filter_by(
            user_id=employee.user_id,
            attendance_date=today
        ).first()

        if attendance:
            status = attendance.status
            if not status:
                # Fallback to gross hours threshold check
                eff_in = attendance.check_in or attendance.card_check_in
                eff_out = attendance.check_out or attendance.card_check_out
                if eff_in and eff_out:
                    gross_sec = (eff_out - eff_in).total_seconds()
                    gross_hours = max(gross_sec, 0) / 3600
                elif eff_in and today == get_ist_today():
                    now = get_ist_now()
                    gross_sec = (now - eff_in).total_seconds()
                    gross_hours = max(gross_sec, 0) / 3600
                else:
                    gross_hours = 0.0
                
                if gross_hours < 4.0:
                    status = "Absent"
                elif gross_hours < 7.0:
                    status = "Half Day"
                else:
                    status = "Present"

            # Override: if checked in but not yet checked out today, show 'Check In'
            if (attendance.check_in and not attendance.check_out
                    and not attendance.card_check_out
                    and today == get_ist_today()):
                status = "Check In"


            web_hrs = attendance.total_hours or 0.0
            card_hrs = attendance.card_working_hours or 0.0

            max_hrs = max(web_hrs, card_hrs)

            check_in = (
                attendance.check_in.strftime("%H:%M:%S")
                if attendance.check_in
                else "-"
            )

            check_out = (
                attendance.check_out.strftime("%H:%M:%S")
                if attendance.check_out
                else "-"
            )

            card_check_in = (
                attendance.card_check_in.strftime("%H:%M:%S")
                if attendance.card_check_in
                else "-"
            )

            card_check_out = (
                attendance.card_check_out.strftime("%H:%M:%S")
                if attendance.card_check_out
                else "-"
            )

            total_hours = attendance.total_hours or 0.0

        else:
            status = "Absent"
            check_in = "-"
            check_out = "-"
            card_check_in = "-"
            card_check_out = "-"
            total_hours = 0

        if status == "Absent":
            from models.leave import LeaveRequest
            leave = LeaveRequest.query.filter(
                or_(
                    LeaveRequest.employee_id == str(employee.id),
                    LeaveRequest.employee_id == employee.employee_id
                ),
                LeaveRequest.status == "Approved",
                LeaveRequest.from_date <= today,
                LeaveRequest.to_date >= today
            ).first()
            if leave:
                status = "Half Day" if (leave.total_days is not None and leave.total_days <= 0.5) else "Leave"

        # Check for an approved Permission on this date and credit its hours
        from models.leave import LeaveRequest as LR
        from datetime import time as dtime
        permission_req = LR.query.filter(
            or_(
                LR.employee_id == str(employee.id),
                LR.employee_id == employee.employee_id
            ),
            LR.request_type == "Permission",
            LR.status == "Approved",
            LR.permission_date == today
        ).first()

        has_permission = False
        permission_label = ""
        actual_hours = total_hours or 0.0
        if permission_req and permission_req.from_time and permission_req.to_time:
            has_permission = True
            # Calculate permission duration in hours
            ft = permission_req.from_time
            tt = permission_req.to_time
            perm_seconds = (tt.hour * 3600 + tt.minute * 60) - (ft.hour * 3600 + ft.minute * 60)
            perm_hours = max(perm_seconds, 0) / 3600
            virtual_total_hours = actual_hours + perm_hours

            def fmt_time(t):
                h = t.hour
                ampm = "AM" if h < 12 else "PM"
                h12 = h % 12 or 12
                return f"{h12:02d}:{t.minute:02d} {ampm}"

            permission_label = f"{fmt_time(ft)} – {fmt_time(tt)}"

            # Re-evaluate status now that hours include permission credit virtually
            if status in ("Absent", "Half Day") and has_permission:
                if virtual_total_hours < 4.0:
                    status = "Absent"
                elif virtual_total_hours < 8.0:
                    status = "Half Day"
                else:
                    status = "Present"


        from models.shift_request import ShiftRequest
        emp_ids = [employee.employee_id] if employee.employee_id else []

        wfh_today = ShiftRequest.query.filter(
            ShiftRequest.employee_id.in_(emp_ids),
            ShiftRequest.status == "Approved",
            ShiftRequest.request_type == "WFH",
            ShiftRequest.from_date <= today,
            ShiftRequest.to_date >= today
        ).first()

        shift_change_today = ShiftRequest.query.filter(
            ShiftRequest.employee_id.in_(emp_ids),
            ShiftRequest.status == "Approved",
            ShiftRequest.request_type == "Shift",
            ShiftRequest.from_date <= today,
            ShiftRequest.to_date >= today
        ).first()

        attendance_list.append({
            "user_id": employee.user_id,
            "employee_code": employee.employee_id,
            "employee_name": f"{employee.first_name} {employee.last_name}",
            "department": employee.department,
            "designation": employee.designation,
            "check_in": check_in,
            "check_out": check_out,
            "card_check_in": card_check_in,
            "card_check_out": card_check_out,
            "lunch_minutes": attendance.lunch_minutes if attendance else 0,
            "tea_minutes": attendance.tea_minutes if attendance else 0,
            "total_hours": total_hours,
            "attendance_date": str(today),
            "status": status,
            "is_wfh": bool(wfh_today),
            "is_shift_changed": bool(shift_change_today),
            "check_in_ip": attendance.check_in_ip if attendance else None,
            "check_out_ip": attendance.check_out_ip if attendance else None,
            "has_permission": has_permission,
            "permission_label": permission_label,
            "regularization_check_in": (
                attendance.regularization_check_in.strftime("%H:%M:%S")
                if (attendance and attendance.regularization_check_in)
                else "-"
            ),
            "regularization_check_out": (
                attendance.regularization_check_out.strftime("%H:%M:%S")
                if (attendance and attendance.regularization_check_out)
                else "-"
            ),
            "regularization_total_hours": (
                attendance.regularization_total_hours
                if attendance
                else 0.0
            ),
            "shift_timing": (
                shift_change_today.requested_shift
                if shift_change_today
                else (
                    attendance.shift_timing
                    if attendance and attendance.shift_timing
                    else employee.shift_timing or "General Shift"
                )
            )
        })


    return jsonify(attendance_list)

@attendance_bp.route("/generate-daily-attendance")
def generate_daily_attendance():

    today = date.today()

    users = User.query.filter_by(
        is_active=True
    ).all()

    count = 0

    for user in users:
        # Check if the employee joining date is set and in the future
        emp = Employee.query.filter_by(user_id=user.id).first()
        if emp and emp.joining_date and today < emp.joining_date:
            continue

        existing = Attendance.query.filter_by(
            user_id=user.id,
            attendance_date=today
        ).first()

        if not existing:

            attendance = Attendance(
                user_id=user.id,
                attendance_date=today,
                status="Absent"
            )

            db.session.add(attendance)
            count += 1

    db.session.commit()

    return jsonify({
        "success": True,
        "records_created": count
    })



def _get_period_attendance_records(days_count, include_card_fields=False):
    """Shared by /weekly and /monthly. Batches attendance/leave/shift-request
    lookups for the whole date range up front (instead of one query per
    employee per day) and looks them up from in-memory dicts inside the loop.
    """
    from models.leave import LeaveRequest
    from models.shift_request import ShiftRequest

    result = []

    employees = [e for e in get_all_employees_cached() if (e.status or "").lower() != "inactive"]

    end_date = date.today()
    start_date = end_date - timedelta(days=days_count - 1)

    user_ids = [e.user_id for e in employees]
    leave_emp_ids = list({str(e.id) for e in employees} | {e.employee_id for e in employees if e.employee_id})
    shift_emp_ids = [e.employee_id for e in employees if e.employee_id]

    # BATCH FETCH: attendance across the whole date range
    attendances = Attendance.query.filter(
        Attendance.user_id.in_(user_ids),
        Attendance.attendance_date >= start_date,
        Attendance.attendance_date <= end_date
    ).all()
    attendance_by_key = {(a.user_id, a.attendance_date): a for a in attendances}

    # BATCH FETCH: approved leave requests overlapping the range
    leaves_by_employee = {}
    if leave_emp_ids:
        leaves = LeaveRequest.query.filter(
            LeaveRequest.employee_id.in_(leave_emp_ids),
            LeaveRequest.status == "Approved",
            LeaveRequest.from_date <= end_date,
            LeaveRequest.to_date >= start_date
        ).all()
        for lr in leaves:
            leaves_by_employee.setdefault(str(lr.employee_id), []).append(lr)

    # BATCH FETCH: approved WFH/Shift requests overlapping the range
    shift_requests_by_employee = {}
    if shift_emp_ids:
        shift_requests = ShiftRequest.query.filter(
            ShiftRequest.employee_id.in_(shift_emp_ids),
            ShiftRequest.status == "Approved",
            ShiftRequest.request_type.in_(["WFH", "Shift"]),
            ShiftRequest.from_date <= end_date,
            ShiftRequest.to_date >= start_date
        ).all()
        for sr in shift_requests:
            shift_requests_by_employee.setdefault(str(sr.employee_id), []).append(sr)

    for i in range(days_count):

        current_date = end_date - timedelta(days=i)

        for employee in employees:
            # Skip if date is prior to date of joining (D.O.J)
            if employee.joining_date and current_date < employee.joining_date:
                continue

            attendance = attendance_by_key.get((employee.user_id, current_date))

            if attendance:
                status = attendance.status
                if not status:
                    # Fallback to gross hours threshold check
                    eff_in = attendance.check_in or attendance.card_check_in
                    eff_out = attendance.check_out or attendance.card_check_out
                    if eff_in and eff_out:
                        gross_sec = (eff_out - eff_in).total_seconds()
                        gross_hours = max(gross_sec, 0) / 3600
                    elif eff_in and current_date == date.today():
                        now = get_ist_now()
                        gross_sec = (now - eff_in).total_seconds()
                        gross_hours = max(gross_sec, 0) / 3600
                    else:
                        gross_hours = 0.0
                    
                    if gross_hours < 4.0:
                        status = "Absent"
                    elif gross_hours < 7.0:
                        status = "Half Day"
                    else:
                        status = "Present"

                card_check_in = (
                    attendance.card_check_in.strftime("%I:%M %p")
                    if attendance.card_check_in
                    else "-"
                )
                card_check_out = (
                    attendance.card_check_out.strftime("%I:%M %p")
                    if attendance.card_check_out
                    else "-"
                )
            else:
                status = "Absent"
                card_check_in = "-"
                card_check_out = "-"

            if status == "Absent":
                leave = None
                for k in [k for k in (str(employee.id), employee.employee_id) if k]:
                    leave = next(
                        (lr for lr in leaves_by_employee.get(k, []) if lr.from_date <= current_date <= lr.to_date),
                        None
                    )
                    if leave:
                        break
                if leave:
                    status = "Half Day" if (leave.total_days is not None and leave.total_days <= 0.5) else "Leave"

            emp_shift_requests = shift_requests_by_employee.get(employee.employee_id, []) if employee.employee_id else []

            wfh_today = next(
                (sr for sr in emp_shift_requests if sr.request_type == "WFH" and sr.from_date <= current_date <= sr.to_date),
                None
            )
            shift_change_today = next(
                (sr for sr in emp_shift_requests if sr.request_type == "Shift" and sr.from_date <= current_date <= sr.to_date),
                None
            )

            record = {

                "employee_code":
                    employee.employee_id,

                "employee_name":
                    f"{employee.first_name} {employee.last_name}",

                "team":
                    employee.department
                    if employee.department
                    else "-",

                "date":
                    current_date.strftime("%d-%m-%Y"),

                "check_in":
                    attendance.check_in.strftime("%I:%M %p")
                    if attendance and attendance.check_in
                    else "-",

                "check_out":
                    attendance.check_out.strftime("%I:%M %p")
                    if attendance and attendance.check_out
                    else "-",

                "total_hours":
                    attendance.total_hours
                    if attendance
                    else "-",

                "status": status,
                "is_wfh": bool(wfh_today),
                "is_shift_changed": bool(shift_change_today),

                "shift_timing":
                    shift_change_today.requested_shift
                    if shift_change_today
                    else (
                        attendance.shift_timing
                        if attendance and attendance.shift_timing
                        else (
                            employee.shift_timing
                            or "General Shift"
                        )
                    )
            }

            if include_card_fields:
                record["card_check_in"] = card_check_in
                record["card_check_out"] = card_check_out

            result.append(record)

    return result


@attendance_bp.route(
    "/weekly",
    methods=["GET"]
)
def get_weekly_attendance():

    try:
        result = _get_period_attendance_records(7, include_card_fields=True)
        return jsonify(result)

    except Exception as e:

        return jsonify({
            "error": str(e)
        }), 500


@attendance_bp.route(
    "/monthly",
    methods=["GET"]
)
def get_monthly_attendance():

    try:
        result = _get_period_attendance_records(30, include_card_fields=False)
        return jsonify(result)

    except Exception as e:

        return jsonify({
            "error": str(e)
        }), 500
    
@attendance_bp.route(
    "/available-months",
    methods=["GET"]
)
@auth_required
def get_available_months():
    try:
        # Get all distinct attendance dates in the database
        dates = db.session.query(Attendance.attendance_date).distinct().all()
        
        cycles = set()
        for d_tuple in dates:
            d = d_tuple[0]
            if not d:
                continue
            # Determine cycle month and year: day >= 25 is next month
            if d.day >= 25:
                if d.month == 12:
                    cycles.add((d.year + 1, 1))
                else:
                    cycles.add((d.year, d.month + 1))
            else:
                cycles.add((d.year, d.month))
                
        # Sort descending
        sorted_cycles = sorted(list(cycles), key=lambda x: (x[0], x[1]), reverse=True)
        
        results = []
        for y, m in sorted_cycles:
            dt = date(y, m, 1)
            results.append({
                "year": y,
                "month": m,
                "label": dt.strftime("%B %Y")
            })
            
        if not results:
            today = date.today()
            results.append({
                "year": today.year,
                "month": today.month,
                "label": today.strftime("%B %Y")
            })
            
        return jsonify(results)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@attendance_bp.route(
    "/export-monthly",
    methods=["GET"]
)
@auth_required
def export_monthly_attendance():

    try:

        wb = Workbook()

        ws = wb.active

        ws.title = "Attendance Report"

        month_param = request.args.get("month")
        year_param = request.args.get("year")
        
        if month_param and year_param:
            try:
                m = int(month_param)
                y = int(year_param)
                if m == 1:
                    start_date = date(y - 1, 12, 25)
                else:
                    start_date = date(y, m - 1, 25)
                end_date = date(y, m, 24)
            except ValueError:
                today = date.today()
                if today.month == 1:
                    start_date = date(today.year - 1, 12, 25)
                else:
                    start_date = date(today.year, today.month - 1, 25)
                end_date = date(today.year, today.month, 24)
        else:
            today = date.today()
            if today.month == 1:
                start_date = date(today.year - 1, 12, 25)
            else:
                start_date = date(today.year, today.month - 1, 25)
            end_date = date(today.year, today.month, 24)

        today_ist = get_ist_today()
        yesterday_ist = today_ist - timedelta(days=1)
        effective_end_date = min(end_date, yesterday_ist)

        def get_ordinal_suffix(day):
            if 11 <= day <= 13:
                return "th"
            return {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")

        # =====================================
        # STYLES
        # =====================================

        sky_blue_fill = PatternFill(
            fill_type="solid",
            fgColor="1F7A8C"  # Brand Primary Blue/Teal
        )

        yellow_fill = PatternFill(
            fill_type="solid",
            fgColor="FBBF24"  # Brand Yellow
        )

        leave_deduction_fill = PatternFill(
            fill_type="solid",
            fgColor="FDE68A"  # Brand Soft Yellow for leave highlight
        )

        white_font = Font(
            bold=True,
            color="FFFFFF",
            size=11
        )

        bold_font = Font(
            bold=True,
            size=11
        )

        thin_border = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1")
        )

        # =====================================
        # TITLE
        # =====================================

        start_day_suff = f"{start_date.day}{get_ordinal_suffix(start_date.day)}"
        end_day_suff = f"{effective_end_date.day}{get_ordinal_suffix(effective_end_date.day)}"

        # Merged A1:S1 for S4C Period Title
        ws.merge_cells("A1:S1")
        ws["A1"] = f"S4C - Attendance for the period from {start_day_suff} {start_date.strftime('%B')} {start_date.year} to {end_day_suff} {effective_end_date.strftime('%B')} {effective_end_date.year}"
        ws["A1"].fill = sky_blue_fill
        ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40

        # Merged A2:S2 for Summary Month Subtitle
        ws.merge_cells("A2:S2")
        ws["A2"] = f"Attendance Summary {effective_end_date.strftime('%B %Y')}"
        ws["A2"].fill = sky_blue_fill
        ws["A2"].font = Font(bold=True, size=11, color="FFFFFF")
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 24

        # Merged A3:S3 for Attendance Cycle Date Range
        ws.merge_cells("A3:S3")
        ws["A3"] = f"Attendance Cycle : {start_date.strftime('%d-%b-%Y')} to {effective_end_date.strftime('%d-%b-%Y')}"
        ws["A3"].fill = yellow_fill
        ws["A3"].font = bold_font
        ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[3].height = 24

        # Column headers (Row 5)
        headers = [
            "S.No",
            "Emp Code",
            "Emp Name",
            "D.O.J",
            "Designation",
            "Department",
            "No of Days in Cycle",
            "Total No of Days Worked",
            "No of Leaves Taken",
            "Previous PL",
            "After PL",
            "Previous CL/SL",
            "After CL/SL",
            "Paid Leave",
            "LOP",
            "Remarks",
            "Total ODW Days",
            "ODW Dates",
            "Late Deductions"
        ]

        ws.row_dimensions[5].height = 87

        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col_num)
            cell.value = header
            cell.fill = sky_blue_fill
            cell.font = white_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # =====================================
        # EMPLOYEE DATA
        # =====================================
        from routes.employees import is_manager_match

        manager_id = request.args.get("manager_id")
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id)) if current_user_id else None
        is_hr_or_admin = current_user and current_user.access_level.lower() in ["admin", "hr"]

        def is_employee_valid_for_report(e):
            fname = (e.first_name or "").lower()
            lname = (e.last_name or "").lower()
            full_name = f"{fname} {lname}".strip()
            
            if "test" in fname or "test" in lname or full_name == "hr admin":
                return False

            if (e.status or "").lower() == "inactive":
                return False
            if e.is_active is False and e.last_working_date:
                lwd = e.last_working_date.date() if isinstance(e.last_working_date, datetime) else e.last_working_date
                if lwd < start_date:
                    return False
            return True

        if manager_id:
            manager_emp = Employee.query.filter_by(user_id=int(manager_id)).first()
            if manager_emp:
                manager_full_name = f"{manager_emp.first_name} {manager_emp.last_name}".strip()
                employees = [
                    e for e in get_all_employees_cached()
                    if is_employee_valid_for_report(e)
                    and is_manager_match(e.reporting_manager, manager_full_name)
                ]
            else:
                employees = []
        elif not is_hr_or_admin:
            if current_user:
                caller_emp = Employee.query.filter_by(user_id=current_user.id).first()
                if caller_emp:
                    manager_full_name = f"{caller_emp.first_name} {caller_emp.last_name}".strip()
                    employees = [
                        e for e in get_all_employees_cached()
                        if is_employee_valid_for_report(e)
                        and is_manager_match(e.reporting_manager, manager_full_name)
                    ]
                else:
                    employees = []
            else:
                employees = []
        else:
            employees = [e for e in get_all_employees_cached() if is_employee_valid_for_report(e)]

        # Filter by team/department if requested
        team_param = request.args.get("team")
        if team_param and team_param != "All":
            employees = [
                e for e in employees
                if (e.department or "").strip().lower() == team_param.strip().lower()
            ]

        # Sort employees by employee code in ascending order
        def get_emp_code_val(e):
            code = getattr(e, "employee_id", None) or getattr(e, "user_id", "")
            try:
                return int(code)
            except (ValueError, TypeError):
                return 999999
        employees = sorted(employees, key=get_emp_code_val)

        # Fetch holidays and overrides within range
        from models.holiday import Holiday, HolidayOverride
        holidays = Holiday.query.filter(Holiday.date >= start_date, Holiday.date <= end_date).all()
        overrides = HolidayOverride.query.filter(HolidayOverride.date >= start_date, HolidayOverride.date <= end_date).all()
        
        holiday_dict = {h.date: h.name for h in holidays}
        override_dict = {o.date: (o.override_type, o.name) for o in overrides}

        row = 6

        def get_shift_start_time(shift_name):
            s = (shift_name or "").lower().strip()
            if "first" in s:
                return datetime.strptime("07:00", "%H:%M").time()
            elif "second" in s:
                return datetime.strptime("12:00", "%H:%M").time()
            elif "night" in s:
                return datetime.strptime("22:00", "%H:%M").time()
            else:
                return datetime.strptime("09:00", "%H:%M").time()

        for index, employee in enumerate(employees, start=1):

            attendance_records = Attendance.query.filter(
                Attendance.user_id == employee.user_id,
                Attendance.attendance_date >= start_date,
                Attendance.attendance_date <= effective_end_date
            ).all()

            attendance_by_date = {a.attendance_date: a for a in attendance_records}

            # Map approved leaves covering the dates
            emp_leaves = LeaveRequest.query.filter(
                LeaveRequest.employee_id.in_([employee.employee_id, str(employee.id)]),
                LeaveRequest.status == "Approved",
                LeaveRequest.request_type == "Leave",
                LeaveRequest.from_date <= effective_end_date,
                LeaveRequest.to_date >= start_date
            ).all()

            # Retrieve employee's live leave balances from DB
            from models.leave import EmployeeLeaveBalance
            cl_sl_bal_recs = EmployeeLeaveBalance.query.filter(
                EmployeeLeaveBalance.employee_id == employee.id,
                func.lower(EmployeeLeaveBalance.leave_type).in_(["cl/sl", "cl / sl", "sl/cl", "sl / cl"])
            ).all()
            pl_bal_recs = EmployeeLeaveBalance.query.filter(
                EmployeeLeaveBalance.employee_id == employee.id,
                func.lower(EmployeeLeaveBalance.leave_type).in_(["pl", "privilege leave"])
            ).all()
            
            current_cl_sl_bal = sum(r.available or 0.0 for r in cl_sl_bal_recs)
            current_pl_bal = sum(r.available or 0.0 for r in pl_bal_recs)

            # Find all approved leaves starting from start_date (to reconstruct previous balance)
            future_leaves = LeaveRequest.query.filter(
                LeaveRequest.employee_id.in_([employee.employee_id, str(employee.id)]),
                LeaveRequest.status == "Approved",
                LeaveRequest.request_type == "Leave",
                LeaveRequest.from_date >= start_date
            ).all()

            future_cl_sl_taken = 0.0
            future_pl_taken = 0.0
            for l in future_leaves:
                l_type = (l.leave_type or "").lower()
                if "loss of pay" in l_type or "lop" in l_type or "unpaid" in l_type:
                    continue
                is_cl_sl = "casual" in l_type or "sick" in l_type or "cl/sl" in l_type or "cl / sl" in l_type
                is_pl = "privilege" in l_type or "pl" in l_type
                if is_cl_sl:
                    future_cl_sl_taken += l.total_days or 0.0
                elif is_pl:
                    future_pl_taken += l.total_days or 0.0

            previous_cl_sl = current_cl_sl_bal + future_cl_sl_taken
            previous_pl = current_pl_bal + future_pl_taken

            total_working_days = 0.0
            total_weekoffs = 0.0
            total_holidays = 0.0
            total_odw_days = 0.0

            cl_sl_taken = 0.0
            pl_taken = 0.0
            lop_leave_taken = 0.0
            unauthorized_absences = 0.0
            late_count = 0

            leave_dates = []
            absent_dates = []

            from models.shift_request import ShiftRequest
            emp_wages = ShiftRequest.query.filter(
                ShiftRequest.employee_id.in_([employee.employee_id, str(employee.id)]),
                ShiftRequest.request_type == "One Day Wages",
                ShiftRequest.status == "Approved",
                ShiftRequest.from_date <= effective_end_date,
                ShiftRequest.to_date >= start_date
            ).all()

            emp_effective_end = effective_end_date
            if employee.is_active is False and employee.last_working_date:
                lwd = employee.last_working_date.date() if isinstance(employee.last_working_date, datetime) else employee.last_working_date
                if lwd < effective_end_date:
                    emp_effective_end = lwd

            if start_date > emp_effective_end:
                num_days_to_calculate = 0
            else:
                num_days_to_calculate = (emp_effective_end - start_date).days + 1

            odw_days_list = []

            for i in range(num_days_to_calculate):
                d = start_date + timedelta(days=i)
                
                # Skip days prior to the employee's date of joining (D.O.J)
                if employee.joining_date and d < employee.joining_date:
                    continue
                
                # Check holiday/weekoff status
                is_holiday = False
                is_week_off = False
                override = override_dict.get(d)
                if override:
                    if override[0] == "Holiday":
                        is_holiday = True
                else:
                    if d in holiday_dict:
                        is_holiday = True
                    elif is_date_week_off(d):
                        is_week_off = True

                # Check if there is an approved leave request on this date
                day_leaves = [l for l in emp_leaves if l.from_date <= d and l.to_date >= d]
                leave_val = 0.0
                is_cl_sl_leave = False
                is_pl_leave = False
                is_lop_leave = False

                if day_leaves:
                    first_leave = day_leaves[0]
                    leave_type_lower = (first_leave.leave_type or "").lower()
                    is_cl_sl_leave = "casual" in leave_type_lower or "sick" in leave_type_lower or "cl/sl" in leave_type_lower or "cl / sl" in leave_type_lower
                    is_pl_leave = "privilege" in leave_type_lower or "pl" in leave_type_lower
                    is_lop_leave = "loss of pay" in leave_type_lower or "lop" in leave_type_lower or "unpaid" in leave_type_lower
                    
                    if first_leave.total_days is not None and first_leave.total_days <= 0.5:
                        leave_val = 0.5
                    else:
                        leave_val = 1.0

                # Check if there is an approved one day wages request covering this date
                day_wages = [w for w in emp_wages if w.from_date <= d and w.to_date >= d]
                is_odw = len(day_wages) > 0
                if is_odw:
                    total_odw_days += 1.0
                    odw_days_list.append(d)

                att = attendance_by_date.get(d)
                
                # Late check-in detection
                if att:
                    eff_check_in = att.check_in or att.card_check_in
                    if eff_check_in:
                        # Determine shift timing for this day
                        approved_shift_req = ShiftRequest.query.filter(
                            ShiftRequest.employee_id.in_([employee.employee_id, str(employee.id)]),
                            ShiftRequest.request_type == "Shift",
                            ShiftRequest.status == "Approved",
                            ShiftRequest.from_date <= d,
                            ShiftRequest.to_date >= d
                        ).order_by(ShiftRequest.created_at.desc()).first()
                        
                        effective_shift = (
                            (approved_shift_req.requested_shift if approved_shift_req and approved_shift_req.requested_shift else None)
                            or att.shift_timing
                            or employee.shift_timing
                            or "General Shift"
                        )
                        
                        shift_start = get_shift_start_time(effective_shift)
                        start_dt = datetime.combine(eff_check_in.date(), shift_start)
                        grace_dt = start_dt + timedelta(minutes=15)
                        if eff_check_in > grace_dt:
                            late_count += 1

                if is_odw:
                    if is_holiday:
                        total_holidays += 1.0
                    elif is_week_off:
                        total_weekoffs += 1.0
                elif is_holiday:
                    if att and att.status == "Present":
                        total_working_days += 1.0
                    else:
                        total_holidays += 1.0
                elif is_week_off:
                    if att and att.status == "Present":
                        total_working_days += 1.0
                    else:
                        total_weekoffs += 1.0
                else:
                    # Normal working day
                    effective_status = att.status if att else None
                    if leave_val > 0.0:
                        # Prioritize approved leave
                        needed_leave = min(leave_val, 1.0)
                        is_half_leave = (needed_leave <= 0.5)
                        if is_lop_leave:
                            lop_leave_taken += needed_leave
                            if (d, is_half_leave) not in absent_dates:
                                absent_dates.append((d, is_half_leave))
                        elif is_cl_sl_leave:
                            cl_sl_taken += needed_leave
                            if (d, is_half_leave) not in leave_dates:
                                leave_dates.append((d, is_half_leave))
                        elif is_pl_leave:
                            pl_taken += needed_leave
                            if (d, is_half_leave) not in leave_dates:
                                leave_dates.append((d, is_half_leave))
                        
                        # Process remaining day portion with attendance
                        remaining_day = 1.0 - needed_leave
                        if remaining_day > 0.0:
                            if att and effective_status in ["Present", "Half Day"]:
                                total_working_days += remaining_day
                            else:
                                unauthorized_absences += remaining_day
                                if (d, True) not in absent_dates:
                                    absent_dates.append((d, True))
                    else:
                        # No approved leave
                        if att and effective_status == "Present":
                            total_working_days += 1.0
                        elif att and effective_status == "Half Day":
                            total_working_days += 0.5
                            unauthorized_absences += 0.5
                            if (d, True) not in absent_dates:
                                absent_dates.append((d, True))
                        else:
                            unauthorized_absences += 1.0
                            if (d, False) not in absent_dates:
                                absent_dates.append((d, False))

            # Apply leave balance caps and split into paid vs unpaid LOP
            # CL/SL:
            if cl_sl_taken > previous_cl_sl:
                paid_cl_sl = previous_cl_sl
                lop_cl_sl = cl_sl_taken - previous_cl_sl
            else:
                paid_cl_sl = cl_sl_taken
                lop_cl_sl = 0.0
            
            # PL:
            if pl_taken > previous_pl:
                paid_pl = previous_pl
                lop_pl = pl_taken - previous_pl
            else:
                paid_pl = pl_taken
                lop_pl = 0.0

            after_cl_sl = max(previous_cl_sl - paid_cl_sl, 0.0)
            after_pl = max(previous_pl - paid_pl, 0.0)
            
            total_leaves_taken = cl_sl_taken + pl_taken + lop_leave_taken
            total_paid_leaves = paid_cl_sl + paid_pl
            total_lop_days = unauthorized_absences + lop_cl_sl + lop_pl + lop_leave_taken

            total_days_worked = total_working_days + total_weekoffs + total_holidays + total_paid_leaves
            effective_start = max(start_date, employee.joining_date) if employee.joining_date else start_date
            if effective_start > emp_effective_end:
                total_days_cycle = 0
            else:
                total_days_cycle = (emp_effective_end - effective_start).days + 1

            # Format leave and absent remarks dynamically
            from collections import defaultdict
            def format_days_for_remarks(day_tuples):
                if not day_tuples:
                    return ""
                
                # Group by month
                by_month = defaultdict(list)
                for dt, is_half in day_tuples:
                    month_name = dt.strftime("%b")
                    suffix = " (Half)" if is_half else ""
                    day_suff = f"{dt.day}{get_ordinal_suffix(dt.day)}{suffix}"
                    by_month[month_name].append(day_suff)
                    
                month_parts = []
                for month_name, days in by_month.items():
                    if len(days) == 1:
                        month_parts.append(f"{days[0]} {month_name}")
                    elif len(days) == 2:
                        month_parts.append(f"{days[0]} and {days[1]} {month_name}")
                    else:
                        month_parts.append(f"{', '.join(days[:-1])} and {days[-1]} {month_name}")
                
                return " and ".join(month_parts)

            remark_parts = []
            if leave_dates:
                sorted_leaves = sorted(leave_dates, key=lambda x: x[0])
                remark_parts.append(f"Leave on {format_days_for_remarks(sorted_leaves)}")
            if absent_dates:
                sorted_absences = sorted(absent_dates, key=lambda x: x[0])
                remark_parts.append(f"Absent on {format_days_for_remarks(sorted_absences)}")
            
            leave_remarks = "; ".join(remark_parts)

            # Format ODW dates list
            odw_formatted_list = []
            for dt in odw_days_list:
                day_suff = f"{dt.day}{get_ordinal_suffix(dt.day)}"
                month_name = dt.strftime("%b")
                odw_formatted_list.append(f"{day_suff} {month_name}")
            
            if total_odw_days > 0:
                odw_count_str = str(int(total_odw_days)) if total_odw_days == int(total_odw_days) else str(total_odw_days)
                oneday_wages_val = f"{odw_count_str} ({', '.join(odw_formatted_list)})"
            else:
                oneday_wages_val = "0"

            # Set late deductions to 0 by default
            late_ded_str = "0"

            # Populate cells
            ws.cell(row=row, column=1).value = index
            ws.cell(row=row, column=2).value = employee.employee_id or employee.user_id
            ws.cell(row=row, column=3).value = f"{employee.first_name} {employee.last_name}"
            ws.cell(row=row, column=4).value = employee.joining_date.strftime("%d-%m-%Y") if employee.joining_date else ""
            ws.cell(row=row, column=5).value = employee.designation or "-"
            ws.cell(row=row, column=6).value = employee.department or "-"
            ws.cell(row=row, column=7).value = total_days_cycle
            ws.cell(row=row, column=8).value = total_days_worked
            ws.cell(row=row, column=9).value = total_leaves_taken
            ws.cell(row=row, column=10).value = previous_pl
            ws.cell(row=row, column=11).value = after_pl
            ws.cell(row=row, column=12).value = previous_cl_sl
            ws.cell(row=row, column=13).value = after_cl_sl
            ws.cell(row=row, column=14).value = total_paid_leaves
            ws.cell(row=row, column=15).value = total_lop_days
            ws.cell(row=row, column=16).value = leave_remarks
            ws.cell(row=row, column=17).value = int(total_odw_days) if total_odw_days == int(total_odw_days) else total_odw_days
            ws.cell(row=row, column=18).value = ", ".join(odw_formatted_list) if odw_formatted_list else "-"
            ws.cell(row=row, column=19).value = late_ded_str

            # Highlight leave fields if leaves were deducted (subtracted)
            if after_pl < previous_pl:
                ws.cell(row=row, column=10).fill = leave_deduction_fill
                ws.cell(row=row, column=11).fill = leave_deduction_fill
            if after_cl_sl < previous_cl_sl:
                ws.cell(row=row, column=12).fill = leave_deduction_fill
                ws.cell(row=row, column=13).fill = leave_deduction_fill

            # Border and alignment (centered)
            for col in range(1, 20):
                c = ws.cell(row=row, column=col)
                c.border = thin_border
                val = c.value
                if val is not None and val != "":
                    if isinstance(val, (int, float, date, datetime)) or (isinstance(val, str) and (val.isdigit() or val.startswith("EMP"))):
                        c.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        if col in [5, 6, 16, 18]:  # Designation (E), Department (F), Remarks (P), ODW Dates (R)
                            c.alignment = Alignment(vertical="center", wrap_text=True)
                        else:
                            c.alignment = Alignment(vertical="center")
                else:
                    c.alignment = Alignment(horizontal="center", vertical="center")

            row += 1

        # =====================================
        # AUTO WIDTH
        # =====================================
        for column_cells in ws.columns:
            length = max(
                len(str(cell.value))
                if cell.value
                else 0
                for cell in column_cells
            )
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max(length + 5, 12)

        # Override specific columns with minimized compact widths
        ws.column_dimensions["A"].width = 6    # S.No
        ws.column_dimensions["B"].width = 10   # Emp Code
        ws.column_dimensions["C"].width = 20   # Emp Name
        ws.column_dimensions["E"].width = 20   # Designation
        ws.column_dimensions["F"].width = 15   # Department
        ws.column_dimensions["P"].width = 25   # Remarks
        ws.column_dimensions["R"].width = 23   # ODW Dates

        # Numeric columns set to width 7
        numeric_cols = ["G", "H", "I", "J", "K", "L", "M", "N", "O", "Q", "S"]
        for col_let in numeric_cols:
            ws.column_dimensions[col_let].width = 7

        # FREEZE PANES
        # =====================================
        ws.freeze_panes = "A6"

        # Hide empty columns T, U, V
        for col_let in ["T", "U", "V"]:
            ws.column_dimensions[col_let].hidden = True

        # Page Setup: Fit to 1 page wide
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        # =====================================
        # FILTER
        # =====================================
        # ws.auto_filter.ref = f"A5:R{row-1}"
        ws.auto_filter.ref = f"A5:S{row-1}"

        # =====================================
        # SAVE FILE
        # =====================================
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="Attendance_Report.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500
    

@attendance_bp.route(
    "/credit-monthly-leaves",
    methods=["POST"]
)
def credit_monthly_leaves():

    try:

        current_month = datetime.now().strftime("%B")
        current_year = datetime.now().year

        employees = [e for e in Employee.query.all() if (e.status or "").lower() != "inactive"]

        for employee in employees:

            existing = LeaveLedger.query.filter_by(
                employee_id=employee.employee_id,
                month=current_month,
                year=current_year
            ).first()

            if existing:
                continue

            opening_cl = employee.casual_leave or 0
            opening_sl = employee.sick_leave or 0
            opening_pl = employee.privilege_leave or 0

            credit_cl = 6
            credit_sl = 6
            credit_pl = 15

            closing_cl = opening_cl + credit_cl
            closing_sl = opening_sl + credit_sl
            closing_pl = opening_pl + credit_pl

            employee.casual_leave = closing_cl
            employee.sick_leave = closing_sl
            employee.privilege_leave = closing_pl

            ledger = LeaveLedger(

                employee_id=employee.employee_id,

                month=current_month,
                year=current_year,

                opening_cl=opening_cl,
                opening_sl=opening_sl,
                opening_pl=opening_pl,

                credit_cl=credit_cl,
                credit_sl=credit_sl,
                credit_pl=credit_pl,

                closing_cl=closing_cl,
                closing_sl=closing_sl,
                closing_pl=closing_pl
            )

            db.session.add(ledger)

        db.session.commit()

        return jsonify({
            "success": True,
            "message": "Leave credited successfully"
        })

    except Exception as e:

        db.session.rollback()

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500
    
@attendance_bp.route(
    "/export-paysheet",
    methods=["GET"]
)
def export_paysheet():

    try:
        today = date.today()
        # Accept month/year filters from query params (defaults to current month/year)
        month = request.args.get("month", type=int, default=today.month)
        year  = request.args.get("year",  type=int, default=today.year)

        # Determine the payroll period for the requested month
        if month == 1:
            start_date = date(year - 1, 12, 25)
        else:
            start_date = date(year, month - 1, 25)
        end_date = date(year, month, 24)

        wb = Workbook()
        ws = wb.active
        ws.title = "Paysheet"

        header_fill = PatternFill(fill_type="solid", fgColor="D9A066")
        header_font = Font(bold=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        import calendar
        month_label = f"{calendar.month_name[month]} {year}"

        ws.merge_cells("A1:BE1")
        ws["A1"] = f"PAYSHEET REPORT - {month_label}"
        ws["A1"].font = Font(bold=True, size=16)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = [
            "S.No",
            "EMP NO",
            "Gender",
            "PF No",
            "UAN No",
            "ESI No",
            "Employee Name",
            "Department",
            "Designation",
            "Mail ID",
            "DOJ",
            "No Of Days In Month",
            "Days Payable",
            "Basic",
            "HRA",
            "LTA",
            "Other Allowance",
            "Gross Salary",
            "Earned Basic",
            "Earned HRA",
            "Earned LTA",
            "Earned Other Allowance",
            "Earned Actual Gross",
            "Attendance Bonus",
            "ODW",
            "Total",
            "Internet Charges",
            "Gross Earned Salary",
            "Earned PF Wages",
            "PF Ded Employee",
            "PF Ded Employer",
            "VPF",
            "PF & VPF Ded Employee",
            "ESI Ded Employee",
            "ESI Ded Employer",
            "Salary Advance",
            "TDS",
            "LWF",
            "PT",
            "Other Deduction",
            "Total Deduction",
            "Net Transfer",
            "Account No",
            "IFSC Code",
            "Branch Code",
            "PF Wage",
            "PF",
            "EPS Wage",
            "8.33 %",
            "3.67 %",
            "0.50 %",
            "0.50 % Employer",
            "0.01 %",
            "Bonus",
            "Actual Month CTC",
            "Earned Month CTC",
            "Remarks"
        ]

        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

        # Fetch all PaymentDetails records for this payroll period
        pay_records = PaymentDetails.query.filter_by(
            payroll_period_start=start_date,
            payroll_period_end=end_date
        ).all()

        # Build a lookup: alphanumeric employee_id -> PaymentDetails record
        pay_lookup = {r.employee_id: r for r in pay_records}

        # Get all active employees
        employees = [e for e in get_all_employees_cached() if (e.status or "").lower() != "inactive"]

        # Sort employees by employee code in ascending order (handling prefixes like EMP)
        def get_emp_code_val(e):
            code = getattr(e, "employee_id", None) or ""
            import re
            nums = re.findall(r'\d+', str(code))
            if nums:
                try:
                    return int(nums[0])
                except ValueError:
                    pass
            try:
                return int(code)
            except (ValueError, TypeError):
                return 999999
        employees = sorted(employees, key=get_emp_code_val)

        row = 4
        for index, employee in enumerate(employees, start=1):
            rec = pay_lookup.get(employee.employee_id)

            # Use saved payroll record if exists, else fall back to employee defaults
            if rec:
                no_of_days       = rec.no_of_days
                days_payable     = rec.days_payable
                basic            = rec.basic
                hra              = rec.hra
                lta              = rec.lta
                other_allowance  = rec.other_allowance
                gross_salary     = rec.gross_salary
                earned_basic     = rec.earned_basic
                earned_hra       = rec.earned_hra
                earned_lta       = rec.earned_lta
                earned_other     = rec.earned_other_allowance
                earned_actual_gross = rec.earned_actual_gross
                attendance_bonus = rec.attendance_bonus
                odw              = rec.odw
                total            = rec.total
                internet_charges = rec.internet_charges
                gross_earned     = rec.gross_earned_salary
                earned_pf_wages  = rec.earned_pf_wages
                pf_ded_employee  = rec.pf_ded_employee
                pf_ded_employer  = rec.pf  # pf field = employer PF
                vpf              = rec.vpf
                pf_vpf_ded       = rec.pf_vpf_ded_employee
                esi_employee     = rec.esi_ded_employee
                esi_employer     = rec.esi_ded_employer
                salary_advance   = rec.salary_advance
                tds              = rec.tds
                lwf              = rec.lwf
                pt               = rec.pt
                other_deduction  = rec.other_deduction
                total_deduction  = rec.total_deduction
                net_transfer     = rec.net_transfer
                account_number   = rec.account_number or employee.account_number or ""
                ifsc_code        = rec.ifsc_code or employee.ifsc_code or ""
                branch_code      = rec.branch_code or ""
                pf_wage          = rec.pf_wage
                pf               = rec.pf
                eps_wage         = rec.eps_wage
                pf_8_33          = rec.pf_8_33
                pf_3_67          = rec.pf_3_67
                pf_0_50_pf       = rec.pf_0_50_pf_wage
                pf_0_50_eps      = rec.pf_0_50_eps_wage
                pf_0_01          = rec.pf_0_01
                bonus            = rec.bonus
                actual_ctc       = rec.actual_monthly_ctc
                earned_ctc       = rec.earned_monthly_ctc
                remarks          = rec.payment_status
            else:
                # No saved payroll record for this period — show zeroes / blanks
                total_days_cycle = (end_date - start_date).days + 1
                no_of_days       = total_days_cycle
                days_payable     = total_days_cycle
                salary           = employee.salary or 0
                basic            = round(salary * 0.50, 2)
                hra              = round(salary * 0.25, 2)
                lta              = round(salary * 0.05, 2)
                other_allowance  = round(salary * 0.20, 2)
                gross_salary     = salary
                earned_basic = earned_hra = earned_lta = earned_other = 0
                earned_actual_gross = 0
                attendance_bonus = odw = total = 0
                internet_charges = 0
                gross_earned = earned_pf_wages = 0
                pf_ded_employee = pf_ded_employer = vpf = pf_vpf_ded = 0
                esi_employee = esi_employer = 0
                salary_advance = tds = lwf = pt = other_deduction = total_deduction = 0
                net_transfer     = 0
                account_number   = employee.account_number or ""
                ifsc_code        = employee.ifsc_code or ""
                branch_code      = ""
                pf_wage = pf = eps_wage = 0
                pf_8_33 = pf_3_67 = pf_0_50_pf = pf_0_50_eps = pf_0_01 = 0
                bonus            = 0
                actual_ctc = earned_ctc = 0
                remarks          = "Not Processed"

            data = [
                index,
                employee.employee_id,
                employee.gender,
                employee.pf_number,
                employee.uan_number,
                employee.esi_number,
                f"{employee.first_name} {employee.last_name}",
                employee.department,
                employee.designation,
                employee.email,
                employee.joining_date.strftime("%d-%m-%Y") if employee.joining_date else "",
                no_of_days,
                days_payable,
                basic,
                hra,
                lta,
                other_allowance,
                gross_salary,
                earned_basic,
                earned_hra,
                earned_lta,
                earned_other,
                earned_actual_gross,
                attendance_bonus,
                odw,
                total,
                internet_charges,
                gross_earned,
                earned_pf_wages,
                pf_ded_employee,
                pf_ded_employer,
                vpf,
                pf_vpf_ded,
                esi_employee,
                esi_employer,
                salary_advance,
                tds,
                lwf,
                pt,
                other_deduction,
                total_deduction,
                net_transfer,
                account_number,
                ifsc_code,
                branch_code,
                pf_wage,
                pf,
                eps_wage,
                pf_8_33,
                pf_3_67,
                pf_0_50_pf,
                pf_0_50_eps,
                pf_0_01,
                bonus,
                actual_ctc,
                earned_ctc,
                remarks
            ]

            for col_num, value in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col_num, value=value)
                cell.border = thin_border
                if value is not None and value != "":
                    if isinstance(value, (int, float, date, datetime)) or (
                        isinstance(value, str) and (value.isdigit() or value.startswith("EMP"))
                    ):
                        cell.alignment = Alignment(horizontal="center")

            row += 1

        for column_cells in ws.columns:
            try:
                length = max(
                    len(str(cell.value)) if cell.value else 0
                    for cell in column_cells
                )
                ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 5
            except:
                pass

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Paysheet_{calendar.month_name[month]}_{year}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:

        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


def check_is_non_working_day(check_date):
    try:
        from models.holiday import Holiday, HolidayOverride
        override = HolidayOverride.query.filter_by(date=check_date).first()
        if override:
            if override.override_type == "Working Day":
                return False
            elif override.override_type in ["Holiday", "Weekly Off"]:
                return True

        holiday = Holiday.query.filter_by(date=check_date, is_published=True).first()
        if holiday:
            return True
    except Exception:
        pass

    day_name = check_date.strftime("%A")
    if day_name == "Sunday":
        return True
    elif day_name == "Saturday":
        import calendar
        weeks = calendar.monthcalendar(check_date.year, check_date.month)
        sat_count = 0
        for week in weeks:
            sat = week[5]
            if sat != 0:
                sat_count += 1
                if sat == check_date.day:
                    break
        if sat_count in [2, 4]:
            return True
        return False

    return False


def get_last_working_day(ref_date=None):
    if ref_date is None:
        ref_date = date.today() - timedelta(days=1)

    target_date = ref_date
    max_steps = 14
    steps = 0
    while check_is_non_working_day(target_date) and steps < max_steps:
        target_date -= timedelta(days=1)
        steps += 1
    return target_date


@attendance_bp.route(
    "/approve/<int:employee_id>",
    methods=["PUT"]
)
def approve_attendance(employee_id):
    try:
        target_date_str = request.args.get("date")
        if target_date_str:
            target_date = datetime.strptime(target_date_str, "%Y-%m-%d").date()
        else:
            target_date = date.today() - timedelta(days=1)

        from models.employee import Employee
        emp = Employee.query.get(employee_id)
        if not emp:
            return jsonify({"success": False, "error": "Employee not found"}), 404
        target_user_id = emp.user_id

        attendance = Attendance.query.filter_by(
            user_id=target_user_id,
            attendance_date=target_date
        ).first()

        if not attendance:
            # No attendance row — check if employee has approved or pending leave for this day
            from models.leave import LeaveRequest
            from sqlalchemy import or_ as sql_or
            leave = LeaveRequest.query.filter(
                sql_or(
                    LeaveRequest.employee_id == str(emp.id),
                    LeaveRequest.employee_id == emp.employee_id
                ),
                LeaveRequest.from_date <= target_date,
                LeaveRequest.to_date >= target_date
            ).filter(LeaveRequest.status.in_(["Pending", "Approved"])).first()

            if leave:
                # Confirm leave, create Leave attendance row
                attendance = Attendance(
                    user_id=target_user_id,
                    attendance_date=target_date,
                    status="Leave",
                    leave_type=leave.leave_type or "Leave",
                    manager_status="Approved",
                    check_in=None,
                    check_out=None,
                    total_hours=0.0
                )
                db.session.add(attendance)
                if leave.status == "Pending":
                    leave.status = "Approved"
                    leave.approved_by = emp.reporting_manager or "Manager"
                    leave.approved_at = datetime.now()
            else:
                # ABSENT employee who submitted regularization or leave via employee portal
                # This path should not normally be reached — employee must submit first.
                # Fallback: create a basic Absent/LOP record
                attendance = Attendance(
                    user_id=target_user_id,
                    attendance_date=target_date,
                    status="Absent",
                    manager_status="Approved",
                    is_lop=True,
                    check_in=None,
                    check_out=None,
                    total_hours=0.0
                )
                db.session.add(attendance)

        else:
            # Attendance row exists
            attendance.manager_status = "Approved"

            if attendance.is_regularization:
                # Employee submitted regularization times — apply them now
                attendance.check_in = attendance.regularization_check_in
                attendance.check_out = attendance.regularization_check_out
                attendance.total_hours = attendance.regularization_total_hours or 0.0
                
                calculate_attendance_status(attendance)
                
                # Keep regularization fields populated for history retrieval
            elif attendance.status in ("Leave", "Absent"):
                # Leave confirmation — find the pending LeaveRequest and approve it
                from models.leave import LeaveRequest
                from sqlalchemy import or_ as sql_or
                leave_req = LeaveRequest.query.filter(
                    sql_or(
                        LeaveRequest.employee_id == str(emp.id),
                        LeaveRequest.employee_id == emp.employee_id
                    ),
                    LeaveRequest.from_date <= target_date,
                    LeaveRequest.to_date >= target_date,
                    LeaveRequest.status == "Pending"
                ).first()
                if leave_req:
                    attendance.status = "Leave"
                    attendance.leave_type = leave_req.leave_type
                    leave_req.status = "Approved"
                    leave_req.approved_by = emp.reporting_manager or "Manager"
                    leave_req.approved_at = datetime.now()
            else:
                # Normal present record — just flip manager_status
                # If they forgot portal punches but card punches exist, populate them
                if not attendance.check_in and attendance.card_check_in:
                    attendance.check_in = attendance.card_check_in
                if not attendance.check_out and attendance.card_check_out:
                    attendance.check_out = attendance.card_check_out
                
                # Recalculate hours if we now have check-in and check-out
                if attendance.check_in and attendance.check_out:
                    total_seconds = (attendance.check_out - attendance.check_in).total_seconds()
                    break_minutes = attendance.total_break_minutes or 0
                    if not break_minutes:
                        break_minutes = (attendance.lunch_minutes or 0) + (attendance.tea_minutes or 0)
                    gap_minutes = attendance.total_gap_minutes or 0
                    paused_minutes = attendance.paused_minutes or 0
                    total_seconds -= (break_minutes + gap_minutes + paused_minutes) * 60
                    hours_decimal = max(total_seconds, 0) / 3600
                    attendance.total_hours = int(hours_decimal * 100) / 100

                # Determine correct status
                web_hrs = attendance.total_hours or 0.0
                card_hrs = attendance.card_working_hours or 0.0
                max_hrs = max(web_hrs, card_hrs)

                active_hrs = max_hrs
                if not (attendance.check_out or attendance.card_check_out):
                    effective_in = attendance.check_in or attendance.card_check_in
                    if effective_in:
                        now = get_ist_now()
                        if attendance.attendance_date == now.date():
                            paused_seconds = (attendance.paused_minutes or 0) * 60
                            if attendance.is_paused and attendance.paused_start:
                                elapsed_seconds = (attendance.paused_start - effective_in).total_seconds()
                            else:
                                elapsed_seconds = (now - effective_in).total_seconds()
                            break_seconds = (attendance.total_break_minutes or 0) * 60
                            hours_decimal = max(elapsed_seconds - break_seconds - paused_seconds, 0) / 3600
                            active_hrs = max(hours_decimal, max_hrs)

                if active_hrs < 4.0:
                    attendance.status = "Absent"
                elif active_hrs < 8.0:
                    attendance.status = "Half Day"
                else:
                    attendance.status = "Present"

        db.session.commit()

        # Emit real-time update
        try:
            from extensions import socketio
            if emp:
                payload = {
                    "id": emp.id,
                    "user_id": emp.user_id,
                    "date": attendance.attendance_date.strftime("%Y-%m-%d"),
                    "attendance_status": attendance.status or "Present",
                    "check_in": attendance.check_in.strftime("%I:%M %p") if (attendance and attendance.check_in) else None,
                    "check_out": attendance.check_out.strftime("%I:%M %p") if (attendance and attendance.check_out) else None,
                    "working_hours": attendance.total_hours or 0.0,
                    "lunch_minutes": attendance.lunch_minutes or 0,
                    "tea_minutes": attendance.tea_minutes or 0,
                    "shift": emp.shift_timing or "General Shift",
                    "manager_status": attendance.manager_status or "Approved",
                    "checked_in": (attendance and attendance.check_in is not None and attendance.check_out is None),
                    "lunch_break": attendance.lunch_break or False,
                    "tea_break": attendance.tea_break or False
                }
                socketio.emit("attendance_update", payload)
        except Exception as socket_err:
            print("Failed to emit approve socket:", str(socket_err))

        return jsonify({
            "success": True,
            "message": "Attendance Approved"
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500


# ─────────────────────────────────────────────────────────────────────────────
# Employee Resolution Endpoints (called from Employee Dashboard)
# ─────────────────────────────────────────────────────────────────────────────

@attendance_bp.route(
    "/submit-regularization/<int:employee_id>",
    methods=["PUT"]
)
def submit_regularization(employee_id):
    """Employee submits check-in/check-out times for an absent day (regularization request)."""
    try:
        data = request.get_json(silent=True) or {}
        target_date_str = data.get("date") or request.args.get("date")
        check_in_str = data.get("check_in")   # e.g. "09:15"
        check_out_str = data.get("check_out") # e.g. "18:30"
        reason = (data.get("reason") or "").strip()

        if not target_date_str or not check_in_str or not check_out_str:
            return jsonify({"success": False, "error": "date, check_in, and check_out are required"}), 400

        target_date = datetime.strptime(target_date_str, "%Y-%m-%d").date()

        from models.employee import Employee
        emp = Employee.query.get(employee_id)
        if not emp:
            emp = Employee.query.filter_by(employee_id=str(employee_id)).first()
        if not emp:
            emp = Employee.query.filter_by(user_id=employee_id).first()
        if not emp:
            return jsonify({"success": False, "error": "Employee not found"}), 404

        # Parse check_in / check_out as full datetime on that date
        check_in_dt = datetime.combine(target_date, datetime.strptime(check_in_str, "%H:%M").time())
        check_out_dt = datetime.combine(target_date, datetime.strptime(check_out_str, "%H:%M").time())

        target_user_id = emp.user_id if (emp and emp.user_id) else employee_id

        attendance = Attendance.query.filter_by(
            user_id=target_user_id,
            attendance_date=target_date
        ).first()

        msg_entry = {
            "id": f"msg_{int(datetime.now().timestamp())}",
            "sender_role": "employee",
            "sender_name": f"{emp.first_name} {emp.last_name}",
            "comment": f"Regularization request: {check_in_str} – {check_out_str}. {reason}",
            "timestamp": datetime.now().isoformat()
        }

        # Calculate hours decimal based on check-in and check-out (minus breaks if any)
        total_seconds = (check_out_dt - check_in_dt).total_seconds()
        break_mins = 0
        if attendance:
            break_mins = attendance.total_break_minutes or 0
            if not break_mins:
                break_mins = (attendance.lunch_minutes or 0) + (attendance.tea_minutes or 0)
        total_seconds -= break_mins * 60
        hours_decimal = max(total_seconds, 0) / 3600
        calculated_total_hours = int(hours_decimal * 100) / 100

        if not attendance:
            attendance = Attendance(
                user_id=target_user_id,
                attendance_date=target_date,
                status="Absent",
                manager_status="Clarification Provided",
                is_regularization=True,
                regularization_reason=reason,
                regularization_submitted_at=datetime.now(),
                regularization_check_in=check_in_dt,
                regularization_check_out=check_out_dt,
                regularization_total_hours=calculated_total_hours,
                check_in=None,
                check_out=None,
                total_hours=0.0,
                clarification_history=[msg_entry]
            )
            db.session.add(attendance)
        else:
            attendance.manager_status = "Clarification Provided"
            attendance.is_regularization = True
            attendance.regularization_reason = reason
            attendance.regularization_submitted_at = datetime.now()
            attendance.regularization_check_in = check_in_dt
            attendance.regularization_check_out = check_out_dt
            attendance.regularization_total_hours = calculated_total_hours
            history = list(attendance.clarification_history or [])
            history.append(msg_entry)
            attendance.clarification_history = history

        flag_modified(attendance, "clarification_history")
        flag_modified(attendance, "manager_status")
        flag_modified(attendance, "is_regularization")

        db.session.commit()

        try:
            from extensions import socketio
            socketio.emit("attendance_update", {"user_id": target_user_id, "manager_status": "Clarification Provided", "date": attendance.attendance_date.strftime("%Y-%m-%d")})
        except Exception as socket_err:
            print("Failed to emit regularization socket:", str(socket_err))

        return jsonify({"success": True, "message": "Regularization submitted. Awaiting manager approval."})

    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500


@attendance_bp.route(
    "/pending-regularizations/<int:manager_user_id>",
    methods=["GET"]
)
def get_pending_regularizations(manager_user_id):
    try:
        from models.employee import Employee
        from models.attendance import Attendance
        from models.user import User

        # Find manager employee record
        manager = Employee.query.filter_by(user_id=manager_user_id).first()
        
        # Check if the manager is Admin
        user = User.query.get(manager_user_id)
        is_admin = False
        if user:
            role_name = (user.role.name or "").lower() if user.role else ""
            access_level = (user.access_level or "").lower()
            if "admin" in role_name or "admin" in access_level:
                is_admin = True

        from routes.employees import get_all_employees_cached, is_manager_match

        all_employees = [e for e in get_all_employees_cached() if e.is_active != False]

        if is_admin:
            reporting_employees = all_employees
        else:
            if not manager:
                return jsonify([])
            manager_full_name = f"{manager.first_name} {manager.last_name}".strip()
            reporting_employees = [e for e in all_employees if is_manager_match(e.reporting_manager, manager_full_name)]

        reporting_user_ids = [e.user_id for e in reporting_employees if e.user_id]
        if not reporting_user_ids:
            return jsonify([])

        # Query Attendance table for pending regularizations
        pending_records = Attendance.query.filter(
            Attendance.user_id.in_(reporting_user_ids),
            Attendance.is_regularization == True
        ).order_by(Attendance.attendance_date.desc()).all()

        # Completed/rejected records are kept so they can be viewed in the history tab

        results = []
        # Build lookup for employee details
        emp_lookup = {e.user_id: e for e in reporting_employees if e.user_id}

        for record in pending_records:
            emp = emp_lookup.get(record.user_id)
            emp_name = f"{emp.first_name} {emp.last_name}".strip() if emp else "Employee"
            emp_code = emp.employee_id if emp else "-"
            
            results.append({
                "id": record.id,
                "employee_id": emp.id if emp else record.user_id,
                "employee_code": emp_code,
                "employee_name": emp_name,
                "date": record.attendance_date.strftime("%Y-%m-%d"),
                "attendance_date_formatted": record.attendance_date.strftime("%d %b %Y"),
                "check_in": record.regularization_check_in.strftime("%I:%M %p") if record.regularization_check_in else "-",
                "check_out": record.regularization_check_out.strftime("%I:%M %p") if record.regularization_check_out else "-",
                "reason": record.regularization_reason or "",
                "status": record.status or "Absent",
                "manager_status": record.manager_status
            })

        return jsonify(results)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500



@attendance_bp.route(
    "/apply-leave/<int:employee_id>",
    methods=["PUT"]
)
def apply_leave_for_absent_day(employee_id):
    """Employee applies leave for an absent day flagged by manager as Need Clarification."""
    try:
        data = request.get_json(silent=True) or {}
        target_date_str = data.get("date") or request.args.get("date")
        leave_type = (data.get("leave_type") or "Casual Leave").strip()

        if not target_date_str:
            return jsonify({"success": False, "error": "date is required"}), 400

        target_date = datetime.strptime(target_date_str, "%Y-%m-%d").date()

        from models.employee import Employee
        emp = Employee.query.get(employee_id)
        if not emp:
            emp = Employee.query.filter_by(employee_id=str(employee_id)).first()
        if not emp:
            emp = Employee.query.filter_by(user_id=employee_id).first()
        if not emp:
            return jsonify({"success": False, "error": "Employee not found"}), 404

        target_user_id = emp.user_id if (emp and emp.user_id) else employee_id

        # Deduct leave balance from employee_leave_balances table
        from models.leave import EmployeeLeaveBalance
        from sqlalchemy import func

        balance = EmployeeLeaveBalance.query.filter(
            EmployeeLeaveBalance.employee_id == emp.id,
            func.lower(EmployeeLeaveBalance.leave_type) == leave_type.lower()
        ).first()

        if not balance or balance.available < 1:
            return jsonify({"success": False, "error": f"Insufficient {leave_type} balance. Please choose LOP instead."}), 400

        balance.available -= 1

        reason = (data.get("reason") or "").strip()
        msg_entry = {
            "id": f"msg_{int(datetime.now().timestamp())}",
            "sender_role": "employee",
            "sender_name": f"{emp.first_name} {emp.last_name}",
            "comment": f"Applied {leave_type} for this absent day. Reason: {reason}",
            "timestamp": datetime.now().isoformat()
        }

        # Create corresponding LeaveRequest row so it appears in leave history
        from models.leave import LeaveRequest
        leave_req = LeaveRequest(
            employee_id=emp.employee_id,
            employee_name=f"{emp.first_name} {emp.last_name}",
            request_type="Leave",
            leave_type=leave_type,
            reporting_manager=emp.reporting_manager,
            reason=reason or f"Applied {leave_type} via clarification reply.",
            from_date=target_date,
            to_date=target_date,
            total_days=1,
            status="Pending"
        )
        db.session.add(leave_req)

        attendance = Attendance.query.filter_by(
            user_id=target_user_id,
            attendance_date=target_date
        ).first()

        if not attendance:
            attendance = Attendance(
                user_id=target_user_id,
                attendance_date=target_date,
                status="Leave",
                leave_type=leave_type,
                manager_status="Clarification Provided",
                check_in=None,
                check_out=None,
                total_hours=0.0,
                clarification_history=[msg_entry]
            )
            db.session.add(attendance)
        else:
            attendance.status = "Leave"
            attendance.leave_type = leave_type
            attendance.manager_status = "Clarification Provided"
            attendance.check_in = None
            attendance.check_out = None
            attendance.total_hours = 0.0
            history = list(attendance.clarification_history or [])
            history.append(msg_entry)
            attendance.clarification_history = history

        flag_modified(attendance, "clarification_history")
        flag_modified(attendance, "manager_status")
        flag_modified(attendance, "leave_type")

        db.session.commit()

        try:
            from extensions import socketio
            socketio.emit("attendance_update", {"user_id": target_user_id, "manager_status": "Clarification Provided", "date": attendance.attendance_date.strftime("%Y-%m-%d")})
        except Exception as socket_err:
            print("Failed to emit leave socket:", str(socket_err))

        return jsonify({"success": True, "message": f"{leave_type} applied. Awaiting manager approval."})

    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500


@attendance_bp.route(
    "/accept-lop/<int:employee_id>",
    methods=["PUT"]
)
def accept_lop(employee_id):
    """Employee accepts Loss of Pay for an absent day."""
    try:
        data = request.get_json(silent=True) or {}
        target_date_str = data.get("date") or request.args.get("date")

        if not target_date_str:
            return jsonify({"success": False, "error": "date is required"}), 400

        target_date = datetime.strptime(target_date_str, "%Y-%m-%d").date()

        from models.employee import Employee
        emp = Employee.query.get(employee_id)
        if not emp:
            emp = Employee.query.filter_by(user_id=employee_id).first()
        if not emp:
            return jsonify({"success": False, "error": "Employee not found"}), 404

        target_user_id = emp.user_id if (emp and emp.user_id) else employee_id

        msg_entry = {
            "id": f"msg_{int(datetime.now().timestamp())}",
            "sender_role": "employee",
            "sender_name": f"{emp.first_name} {emp.last_name}",
            "comment": "Accepted Loss of Pay (LOP) for this absent day.",
            "timestamp": datetime.now().isoformat()
        }

        attendance = Attendance.query.filter_by(
            user_id=target_user_id,
            attendance_date=target_date
        ).first()

        if not attendance:
            attendance = Attendance(
                user_id=target_user_id,
                attendance_date=target_date,
                status="Absent",
                manager_status="Clarification Provided",
                is_lop=True,
                check_in=None,
                check_out=None,
                total_hours=0.0,
                clarification_history=[msg_entry]
            )
            db.session.add(attendance)
        else:
            attendance.status = "Absent"
            attendance.manager_status = "Clarification Provided"
            attendance.is_lop = True
            attendance.check_in = None
            attendance.check_out = None
            history = list(attendance.clarification_history or [])
            history.append(msg_entry)
            attendance.clarification_history = history
            flag_modified(attendance, "clarification_history")
            flag_modified(attendance, "manager_status")
            flag_modified(attendance, "is_lop")

        db.session.commit()

        try:
            from extensions import socketio
            socketio.emit("attendance_update", {"user_id": target_user_id, "manager_status": "Clarification Provided", "date": attendance.attendance_date.strftime("%Y-%m-%d")})
        except Exception as socket_err:
            print("Failed to emit LOP socket:", str(socket_err))

        return jsonify({"success": True, "message": "LOP accepted. Awaiting manager approval."})

    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500




@attendance_bp.route(
    "/reject/<int:employee_id>",
    methods=["PUT"]
)
@attendance_bp.route(
    "/need-clarification/<int:employee_id>",
    methods=["PUT"]
)
def reject_attendance(employee_id):

    try:

        is_clarification = "need-clarification" in request.url.path
        target_status = "Need Clarification" if is_clarification else "Rejected"
        
        target_date_str = request.args.get("date")
        data = request.get_json(silent=True) or {}
        reason = data.get("reason") or request.args.get("reason") or ""
        
        if target_date_str:
            target_date = datetime.strptime(target_date_str, "%Y-%m-%d").date()
        else:
            target_date = get_last_working_day()

        from models.employee import Employee
        emp = Employee.query.get(employee_id)
        if not emp:
            emp = Employee.query.filter_by(user_id=employee_id).first()
        if not emp:
            return jsonify({"success": False, "error": "Employee not found"}), 404
        target_user_id = emp.user_id

        attendances = Attendance.query.filter_by(
            user_id=target_user_id,
            attendance_date=target_date
        ).all()

        msg_entry = {
            "id": f"msg_{int(datetime.now().timestamp())}",
            "sender_role": "manager",
            "sender_name": "Manager",
            "comment": reason,
            "timestamp": datetime.now().isoformat()
        }

        if not attendances:
            attendance = Attendance(
                user_id=target_user_id,
                attendance_date=target_date,
                status="Absent",
                manager_status=target_status,
                is_regularization=False,
                clarification_history=[msg_entry],
                check_in=None,
                check_out=None,
                total_hours=0.0
            )
            # Check if this day is a non-working day
            if check_is_non_working_day(target_date):
                is_weekend = target_date.weekday() in (5, 6)
                attendance.status = "Weekly Off" if is_weekend else "Holiday"
            db.session.add(attendance)
        else:
            for att in attendances:
                if not is_clarification:
                    att.check_in = att.card_check_in
                    att.check_out = att.card_check_out
                    
                    # Recalculate hours based on card check-in/out if present
                    if att.check_in and att.check_out:
                        total_seconds = (att.check_out - att.check_in).total_seconds()
                        break_minutes = att.total_break_minutes or 0
                        if not break_minutes:
                            break_minutes = (att.lunch_minutes or 0) + (att.tea_minutes or 0)
                        gap_minutes = att.total_gap_minutes or 0
                        total_seconds -= (break_minutes + gap_minutes) * 60
                        hours_decimal = max(total_seconds, 0) / 3600
                        att.total_hours = int(hours_decimal * 100) / 100
                        att.status = "Present"
                    else:
                        att.total_hours = 0.0
                        if check_is_non_working_day(target_date):
                            is_weekend = target_date.weekday() in (5, 6)
                            att.status = "Weekly Off" if is_weekend else "Holiday"
                        else:
                            att.status = "Absent"

                att.manager_status = target_status
                # Keep regularization fields populated for history retrieval
                history = list(att.clarification_history or [])
                history.append(msg_entry)
                att.clarification_history = history
            attendance = attendances[0]

        db.session.commit()

        # Emit attendance_update socket event for real-time dashboard updates
        try:
            from extensions import socketio
            if emp:
                payload = {
                    "id": emp.id,
                    "user_id": emp.user_id,
                    "date": attendance.attendance_date.strftime("%Y-%m-%d"),
                    "attendance_status": attendance.status or "Absent",
                    "check_in": attendance.check_in.strftime("%I:%M %p") if (attendance and attendance.check_in) else None,
                    "check_out": attendance.check_out.strftime("%I:%M %p") if (attendance and attendance.check_out) else None,
                    "working_hours": attendance.total_hours or 0.0,
                    "lunch_minutes": attendance.lunch_minutes or 0,
                    "tea_minutes": attendance.tea_minutes or 0,
                    "shift": emp.shift_timing or "General Shift",
                    "manager_status": target_status,
                    "reason": reason,
                    "clarification_history": attendance.clarification_history or [],
                    "checked_in": (attendance and attendance.check_in is not None and attendance.check_out is None),
                    "lunch_break": attendance.lunch_break or False,
                    "tea_break": attendance.tea_break or False
                }
                socketio.emit("attendance_update", payload)
        except Exception as socket_err:
            print("Failed to emit socket:", str(socket_err))

        msg = "Need Clarification request sent successfully" if is_clarification else "Regularization request rejected successfully"
        return jsonify({
            "success": True,
            "message": msg,
            "reason": reason
        })

    except Exception as e:

        db.session.rollback()

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@attendance_bp.route(
    "/reply-clarification/<int:employee_id>",
    methods=["PUT"]
)
def reply_clarification(employee_id):
    try:
        data = request.get_json(silent=True) or {}
        reply_text = (data.get("reply") or "").strip()
        target_date_str = data.get("date") or request.args.get("date")

        if target_date_str:
            target_date = datetime.strptime(target_date_str, "%Y-%m-%d").date()
        else:
            target_date = get_last_working_day()

        from models.employee import Employee
        emp = Employee.query.get(employee_id)
        if not emp:
            emp = Employee.query.filter_by(user_id=employee_id).first()
        if not emp:
            return jsonify({"success": False, "error": "Employee not found"}), 404

        target_user_id = emp.user_id if (emp and emp.user_id) else employee_id

        attendances = Attendance.query.filter_by(
            user_id=target_user_id,
            attendance_date=target_date
        ).all()

        if not attendances:
            # If no attendance record exists for an absent day query, create the record
            att = Attendance(
                user_id=target_user_id,
                attendance_date=target_date,
                status="Absent",
                manager_status="Clarification Provided",
                clarification_history=[]
            )
            db.session.add(att)
            attendances = [att]

        msg_entry = {
            "id": f"msg_{int(datetime.now().timestamp())}",
            "sender_role": "employee",
            "sender_name": f"{emp.first_name} {emp.last_name}".strip(),
            "comment": reply_text,
            "timestamp": datetime.now().isoformat()
        }

        from sqlalchemy.orm.attributes import flag_modified

        for att in attendances:
            att.manager_status = "Clarification Provided"
            history = list(att.clarification_history or [])
            history.append(msg_entry)
            att.clarification_history = history
            flag_modified(att, "clarification_history")
            flag_modified(att, "manager_status")

        db.session.commit()

        try:
            from extensions import socketio
            socketio.emit("attendance_update", {"user_id": emp.user_id, "manager_status": "Clarification Provided", "date": att.attendance_date.strftime("%Y-%m-%d")})
        except Exception as socket_err:
            print("Failed to emit reply socket:", str(socket_err))

        return jsonify({
            "success": True,
            "message": "Clarification reply submitted successfully"
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500


@attendance_bp.route(
    "/pending-clarifications/<int:user_id>",
    methods=["GET"]
)
def get_pending_clarifications(user_id):
    try:
        from models.attendance import Attendance
        
        # Clean up any remaining 'Need Clarification' records in the database
        # by converting them to 'Rejected' and reverting the times
        need_clarif_records = Attendance.query.filter_by(
            user_id=user_id,
            manager_status="Need Clarification"
        ).order_by(Attendance.attendance_date.desc()).all()
        
        result = []
        for rec in need_clarif_records:
            check_in_time = (
                rec.check_in.strftime("%I:%M %p")
                if rec.check_in else "-"
            )
            check_out_time = (
                rec.check_out.strftime("%I:%M %p")
                if rec.check_out else "-"
            )
            card_check_in_time = (
                rec.card_check_in.strftime("%I:%M %p")
                if rec.card_check_in else "-"
            )
            card_check_out_time = (
                rec.card_check_out.strftime("%I:%M %p")
                if rec.card_check_out else "-"
            )
            break_str = f"{rec.total_break_minutes} min" if rec.total_break_minutes else "0 min"

            result.append({
                "attendance_date": rec.attendance_date.strftime("%Y-%m-%d"),
                "check_in": check_in_time,
                "check_out": check_out_time,
                "card_check_in": card_check_in_time,
                "card_check_out": card_check_out_time,
                "total_hours": rec.total_hours,
                "card_working_hours": rec.card_working_hours,
                "break_str": break_str,
                "total_break_minutes": rec.total_break_minutes or 0,
                "status": rec.status,
                "clarification_history": rec.clarification_history or []
            })

        return jsonify(result)

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@attendance_bp.route(
    "/approve-all",
    methods=["PUT"]
)
def approve_all_attendance():

    try:
        manager_id = request.args.get("manager_id")
        target_date_str = request.args.get("date")
        if target_date_str:
            yesterday = datetime.strptime(target_date_str, "%Y-%m-%d").date()
        else:
            yesterday = get_last_working_day()

        if manager_id:
            from models.employee import Employee
            manager = Employee.query.filter_by(user_id=int(manager_id)).first()
            if manager:
                manager_name = f"{manager.first_name} {manager.last_name}".strip().lower()
                reporting_employees = [e for e in get_all_employees_cached() if (e.status or "").lower() != "inactive"]
                for employee in reporting_employees:
                    if not employee.reporting_manager:
                        continue
                    e_mgr = employee.reporting_manager.strip().lower()
                    is_match = (e_mgr == manager_name) or (len(e_mgr.split()) == 1 and manager_name.split()[0] == e_mgr) or (len(manager_name.split()) == 1 and e_mgr.split()[0] == manager_name)
                    if is_match:
                        attendances = Attendance.query.filter_by(
                            user_id=employee.user_id,
                            attendance_date=yesterday
                        ).all()
                        if not attendances:
                            attendance = Attendance(
                                user_id=employee.user_id,
                                attendance_date=yesterday,
                                status="Absent",
                                manager_status="Approved",
                                check_in=None,
                                check_out=None,
                                total_hours=0.0
                            )
                            db.session.add(attendance)
                        else:
                            for att in attendances:
                                if att.manager_status != "Need Clarification":
                                    att.manager_status = "Approved"
        else:
            attendances = Attendance.query.filter_by(
                attendance_date=yesterday
            ).all()
            for attendance in attendances:
                if attendance.manager_status != "Need Clarification":
                    attendance.manager_status = "Approved"

        db.session.commit()

        # Emit attendance_approved_all socket event for real-time dashboard updates
        try:
            from extensions import socketio
            socketio.emit("attendance_approved_all", {"status": "Approved"})
        except Exception as socket_err:
            print("Failed to emit approve-all socket:", str(socket_err))

        return jsonify({
            "success": True,
            "message": "All attendance approved"
        })

    except Exception as e:

        db.session.rollback()

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@attendance_bp.route(
    "/reject-all",
    methods=["PUT"]
)
def reject_all_attendance():

    try:
        manager_id = request.args.get("manager_id")
        target_date_str = request.args.get("date")
        data = request.get_json(silent=True) or {}
        reason = data.get("reason") or request.args.get("reason") or ""

        if target_date_str:
            yesterday = datetime.strptime(target_date_str, "%Y-%m-%d").date()
        else:
            yesterday = get_last_working_day()

        if manager_id:
            from models.employee import Employee
            manager = Employee.query.filter_by(user_id=int(manager_id)).first()
            if manager:
                manager_name = f"{manager.first_name} {manager.last_name}".strip().lower()
                reporting_employees = [e for e in get_all_employees_cached() if (e.status or "").lower() != "inactive"]
                for employee in reporting_employees:
                    if not employee.reporting_manager:
                        continue
                    e_mgr = employee.reporting_manager.strip().lower()
                    is_match = (e_mgr == manager_name) or (len(e_mgr.split()) == 1 and manager_name.split()[0] == e_mgr) or (len(manager_name.split()) == 1 and e_mgr.split()[0] == manager_name)
                    if is_match:
                        attendance = Attendance.query.filter_by(
                            user_id=employee.user_id,
                            attendance_date=yesterday
                        ).first()
                        if not attendance:
                            attendance = Attendance(
                                user_id=employee.user_id,
                                attendance_date=yesterday,
                                status="Absent",
                                manager_status="Rejected",
                                rejection_reason=reason,
                                check_in=None,
                                check_out=None,
                                total_hours=0.0
                            )
                            db.session.add(attendance)
                        else:
                            attendance.manager_status = "Rejected"
                            attendance.rejection_reason = reason
        else:
            attendances = Attendance.query.filter_by(
                attendance_date=yesterday
            ).all()
            for attendance in attendances:
                attendance.manager_status = "Rejected"
                attendance.rejection_reason = reason

        db.session.commit()

        try:
            from extensions import socketio
            socketio.emit("attendance_approved_all", {"status": "Rejected", "reason": reason})
        except Exception as socket_err:
            print("Failed to emit reject-all socket:", str(socket_err))

        return jsonify({
            "success": True,
            "message": "All attendance rejected",
            "reason": reason
        })

    except Exception as e:

        db.session.rollback()

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@attendance_bp.route("/upload-excel", methods=["POST"])
def upload_attendance_excel():
    try:
        import xlrd
        from datetime import time
        
        if "file" not in request.files:
            return jsonify({"success": False, "error": "No file uploaded"}), 400
            
        file = request.files["file"]
        if not file or not file.filename.endswith((".xls", ".xlsx")):
            return jsonify({"success": False, "error": "Invalid file format. Please upload an .xls or .xlsx file"}), 400

        # Read Excel content directly in memory
        file_contents = file.read()
        
        # Check if legacy XLS
        if file.filename.endswith(".xls"):
            workbook = xlrd.open_workbook(file_contents=file_contents)
            sheet = workbook.sheet_by_index(0)
            
            # Read Attendance Date from Row index 6, Column index 4
            att_date_raw = sheet.cell_value(6, 4)
            if not att_date_raw:
                return jsonify({"success": False, "error": "Could not locate report date in Excel sheet"}), 400
                
            att_date_str = str(att_date_raw).strip()
            try:
                parsed_date = datetime.strptime(att_date_str, "%d-%b-%Y").date()
            except Exception as de:
                return jsonify({"success": False, "error": f"Failed to parse report date '{att_date_str}': {str(de)}"}), 400

            print("Standardized Excel Upload Date:", parsed_date)
            
            # Loop starting from Row index 10 (11th row)
            r = 10
            processed_count = 0
            batch_attendance = {}
            
            while r < sheet.nrows:
                sno_val = sheet.cell_value(r, 1) # Column 2
                e_code_val = sheet.cell_value(r, 2) # Column 3
                
                # Stop if both columns are empty
                if not sno_val and not e_code_val:
                    break
                    
                # Convert E. Code to string
                if isinstance(e_code_val, float):
                    e_code = str(int(e_code_val))
                else:
                    e_code = str(e_code_val).strip()
                    
                sno_str = str(sno_val).strip()
                
                # Skip header/duplicates/non-records
                if not e_code or e_code == "E. Code" or not sno_str.replace(".0", "").isdigit():
                    r += 1
                    continue
                    
                status = str(sheet.cell_value(r, 13)).strip() # Column 14 (Status)
                if status == "Absent":
                    r += 1
                    continue
                    
                in_time_val = sheet.cell_value(r, 7) # Column 8 (InTime)
                out_time_val = sheet.cell_value(r, 8) # Column 9 (OutTime)
                
                # Convert In/Out values to string time representation
                def get_time_str(cell_val):
                    if not cell_val:
                        return ""
                    val_str = str(cell_val).strip()
                    if val_str == "--:--" or val_str == "0.0" or val_str == "0":
                        return ""
                    if isinstance(cell_val, float) and 0.0 < cell_val < 1.0:
                        total_minutes = int(cell_val * 24 * 60)
                        hours = total_minutes // 60
                        minutes = total_minutes % 60
                        return f"{hours:02d}:{minutes:02d}"
                    return val_str
                    
                in_time_str = get_time_str(in_time_val)
                out_time_str = get_time_str(out_time_val)
                
                # Skip if no punches
                if not in_time_str and not out_time_str:
                    r += 1
                    continue
                    
                # Find employee
                from models.employee import Employee
                employee = Employee.query.filter_by(employee_id=e_code).first()
                if not employee:
                    print(f"Excel skip: Employee '{e_code}' not found")
                    r += 1
                    continue
                    
                cache_key = (employee.user_id, parsed_date)
                if cache_key in batch_attendance:
                    attendance = batch_attendance[cache_key]
                else:
                    attendance = Attendance.query.filter_by(
                        user_id=employee.user_id,
                        attendance_date=parsed_date
                    ).first()
                    
                    if not attendance:
                        attendance = Attendance(
                            user_id=employee.user_id,
                            attendance_date=parsed_date,
                            status="Present",
                            shift_timing=employee.shift_timing or "General Shift"
                        )
                        db.session.add(attendance)
                    batch_attendance[cache_key] = attendance
                    
                # Parse log datetime
                if in_time_str:
                    time_parts = list(map(int, in_time_str.split(":")))
                    card_in_dt = datetime.combine(parsed_date, time(time_parts[0], time_parts[1]))
                    if not attendance.card_check_in:
                        attendance.card_check_in = card_in_dt
                    else:
                        attendance.card_check_in = min(attendance.card_check_in, card_in_dt)
                        
                if out_time_str:
                    time_parts = list(map(int, out_time_str.split(":")))
                    card_out_dt = datetime.combine(parsed_date, time(time_parts[0], time_parts[1]))
                    if not attendance.card_check_out:
                        attendance.card_check_out = card_out_dt
                    else:
                        attendance.card_check_out = max(attendance.card_check_out, card_out_dt)
                        
                # Calculate hours
                if attendance.card_check_in and attendance.card_check_out:
                    total_seconds = (attendance.card_check_out - attendance.card_check_in).total_seconds()
                    hours_decimal = max(total_seconds, 0) / 3600
                    attendance.card_working_hours = int(hours_decimal * 100) / 100
                else:
                    attendance.card_working_hours = 0.0
                    
                # Recalculate Status
                web_hrs = attendance.total_hours or 0.0
                card_hrs = attendance.card_working_hours or 0.0
                max_hrs = max(web_hrs, card_hrs)
                
                if max_hrs < 4.0:
                    attendance.status = "Absent"
                elif max_hrs < 8.0:
                    attendance.status = "Half Day"
                else:
                    attendance.status = "Present"
                    
                processed_count += 1
                
                # Emit socket event
                try:
                    from extensions import socketio
                    web_in = attendance.check_in.strftime("%I:%M %p") if attendance.check_in else None
                    web_out = attendance.check_out.strftime("%I:%M %p") if attendance.check_out else None
                    card_in = attendance.card_check_in.strftime("%I:%M %p") if attendance.card_check_in else None
                    card_out = attendance.card_check_out.strftime("%I:%M %p") if attendance.card_check_out else None
                    
                    payload = {
                        "id": employee.id,
                        "user_id": employee.user_id,
                        "attendance_status": attendance.status,
                        "check_in": web_in,
                        "check_out": web_out,
                        "working_hours": attendance.total_hours or 0.0,
                        "card_check_in": card_in,
                        "card_check_out": card_out,
                        "card_working_hours": attendance.card_working_hours or 0.0,
                        "shift": employee.shift_timing or "General Shift",
                        "manager_status": attendance.manager_status or "Pending",
                        "checked_in": (attendance.check_in is not None and attendance.check_out is None),
                        "card_checked_in": (attendance.card_check_in is not None and attendance.card_check_out is None),
                    }
                    socketio.emit("attendance_update", payload)
                except Exception as se:
                    print("Socket emit error during Excel upload:", se)
                    
                r += 1
                
            db.session.commit()
            return jsonify({"success": True, "message": f"Successfully processed {processed_count} employee records"}), 200
        else:
            return jsonify({"success": False, "error": "Only .xls legacy format is currently supported"}), 400
            
    except Exception as e:
        db.session.rollback()
        print("EXCEL UPLOAD ERROR:", str(e))
        return jsonify({"success": False, "error": str(e)}), 500
@attendance_bp.route("/update-attendance-record", methods=["POST"])
@jwt_required()
def update_attendance_record():
    try:
        from models.attendance import Attendance
        from models.employee import Employee
        from models.user import User
        
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        # Verify if caller is manager/admin/hr
        if not current_user or current_user.access_level.lower() not in ["manager", "admin", "hr"]:
            return jsonify({"success": False, "error": "Access denied"}), 403
            
        data = request.json
        employee_id = data.get("employee_id") # can be int id
        attendance_date_str = data.get("attendance_date") # YYYY-MM-DD
        status = data.get("status")
        
        check_in_str = data.get("check_in") # HH:MM AM/PM or None
        check_out_str = data.get("check_out") # HH:MM AM/PM or None
        card_check_in_str = data.get("card_check_in") # HH:MM AM/PM or None
        card_check_out_str = data.get("card_check_out") # HH:MM AM/PM or None
        
        emp = None
        try:
            emp = Employee.query.get(int(employee_id))
        except Exception:
            pass
        if not emp:
            emp = Employee.query.filter_by(employee_id=str(employee_id)).first()
        if not emp:
            return jsonify({"success": False, "error": f"Employee not found for ID: {employee_id}"}), 404
            
        att_date = datetime.strptime(attendance_date_str, "%Y-%m-%d").date()
        
        # Query attendance record for this date
        attendance = Attendance.query.filter_by(
            user_id=emp.user_id,
            attendance_date=att_date
        ).first()
        
        # Parse helper for check-in/out times (HH:MM AM/PM to datetime)
        def parse_time_str(time_str, base_date):
            if not time_str or time_str == "-":
                return None
            try:
                # Expecting format like "09:00 AM"
                t = datetime.strptime(time_str.strip(), "%I:%M %p").time()
                return datetime.combine(base_date, t)
            except ValueError:
                try:
                    # Fallback to "09:00:00" format
                    t = datetime.strptime(time_str.strip(), "%H:%M:%S").time()
                    return datetime.combine(base_date, t)
                except ValueError:
                    try:
                        # Fallback to "09:00" format
                        t = datetime.strptime(time_str.strip(), "%H:%M").time()
                        return datetime.combine(base_date, t)
                    except ValueError:
                        return None

        parsed_check_in = parse_time_str(check_in_str, att_date)
        parsed_check_out = parse_time_str(check_out_str, att_date)
        parsed_card_check_in = parse_time_str(card_check_in_str, att_date)
        parsed_card_check_out = parse_time_str(card_check_out_str, att_date)
        
        lunch_minutes = int(data.get("lunch_minutes") or 0)
        tea_minutes = int(data.get("tea_minutes") or 0)
        paused_minutes = int(data.get("paused_minutes") or 0)

        if not attendance:
            # Create a new attendance record if it doesn't exist
            attendance = Attendance(
                user_id=emp.user_id,
                attendance_date=att_date,
                status=status,
                check_in=parsed_check_in,
                check_out=parsed_check_out,
                card_check_in=parsed_card_check_in,
                card_check_out=parsed_card_check_out,
                lunch_minutes=lunch_minutes,
                tea_minutes=tea_minutes,
                paused_minutes=paused_minutes,
                total_break_minutes=lunch_minutes + tea_minutes,
                manager_status="Approved"
            )
            db.session.add(attendance)
        else:
            # Update existing record
            attendance.status = status
            attendance.check_in = parsed_check_in
            attendance.check_out = parsed_check_out
            attendance.card_check_in = parsed_card_check_in
            attendance.card_check_out = parsed_card_check_out
            attendance.lunch_minutes = lunch_minutes
            attendance.tea_minutes = tea_minutes
            attendance.paused_minutes = paused_minutes
            attendance.total_break_minutes = lunch_minutes + tea_minutes
            attendance.manager_status = "Approved"

        # Calculate working hours if both check-in and check-out are present
        if attendance.check_in and attendance.check_out:
            diff_seconds = (attendance.check_out - attendance.check_in).total_seconds()
            break_minutes = attendance.total_break_minutes or 0
            gap_minutes = attendance.total_gap_minutes or 0
            paused_minutes = attendance.paused_minutes or 0
            diff_seconds -= (break_minutes + gap_minutes + paused_minutes) * 60
            attendance.total_hours = max(0.0, int((diff_seconds / 3600.0) * 100) / 100)
            
            # Auto-calculate status based on hours
            is_weekend = attendance.attendance_date.weekday() >= 5
            req_hours = 7.0 if is_weekend else 8.0
            
            if attendance.total_hours >= req_hours:
                attendance.status = "Present"
            elif attendance.total_hours >= 4.0:
                attendance.status = "Half Day"
            else:
                attendance.status = "Absent"
        else:
            attendance.total_hours = 0.0

        if attendance.card_check_in and attendance.card_check_out:
            diff_seconds = (attendance.card_check_out - attendance.card_check_in).total_seconds()
            attendance.card_working_hours = max(0.0, round(diff_seconds / 3600.0, 2))
        else:
            attendance.card_working_hours = 0.0
            
        db.session.commit()
        return jsonify({"success": True, "message": "Attendance record updated successfully"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500


@attendance_bp.route("/trigger-db-sync", methods=["POST"])
def trigger_db_sync():
    try:
        mysql_host = os.environ.get("MYSQL_BIOMETRIC_HOST", "10.1.1.18")
        mysql_port = int(os.environ.get("MYSQL_BIOMETRIC_PORT", 3306))
        mysql_user = os.environ.get("MYSQL_BIOMETRIC_USER", "Muralibalu")
        mysql_password = os.environ.get("MYSQL_BIOMETRIC_PASSWORD", "Murali@12")
        mysql_db = os.environ.get("MYSQL_BIOMETRIC_DB", "TimeTrack")

        if not mysql_password:
            return jsonify({
                "success": False,
                "error": "MYSQL_BIOMETRIC_PASSWORD is not set in environment (.env)"
            }), 400

        # 1. Connect to MySQL database
        connection = pymysql.connect(
            host=mysql_host,
            port=mysql_port,
            user=mysql_user,
            password=mysql_password,
            database=mysql_db,
            cursorclass=pymysql.cursors.DictCursor
        )
        
        # 2. Query absolute MIN and MAX punch times per employee per day for the last 30 days
        with connection.cursor() as cursor:
            sql = """
                SELECT 
                    EmployeeCode,
                    MIN(LogDateTime) AS FirstPunch,
                    MAX(LogDateTime) AS LastPunch
                FROM AttendanceLogs
                WHERE LogDate >= DATE_SUB(CURDATE(), INTERVAL 30 DAY)
                  AND LogDate < CURDATE()
                GROUP BY EmployeeCode, LogDate
            """
            cursor.execute(sql)
            rows = cursor.fetchall()
            
        connection.close()
        
        if not rows:
            return jsonify({
                "success": True,
                "message": "No biometric logs found in MySQL for the last 30 days"
            }), 200

        # 3. Process logs directly per employee per day
        processed_count = 0
        
        for row in rows:
            emp_code = str(row["EmployeeCode"]).strip()
            first_punch = row["FirstPunch"]
            last_punch = row["LastPunch"]

            if not first_punch:
                continue

            employee = Employee.query.filter_by(employee_id=emp_code).first()
            if not employee:
                continue

            att_date = first_punch.date()

            # Ignore same day (today) entries
            if att_date == get_ist_today():
                continue

            # Find existing or create new attendance
            attendance = Attendance.query.filter_by(
                user_id=employee.user_id,
                attendance_date=att_date
            ).first()

            # Determine target values
            target_check_in = first_punch
            target_check_out = last_punch if last_punch != first_punch else None

            # Skip if we already have the exact same biometric details in the DB
            # (No need to update same date details if already synced and identical)
            if attendance:
                if attendance.card_check_in == target_check_in and attendance.card_check_out == target_check_out:
                    continue

            if not attendance:
                attendance = Attendance(
                    user_id=employee.user_id,
                    attendance_date=att_date,
                    status="Present",
                    shift_timing=employee.shift_timing or "General Shift"
                )
                db.session.add(attendance)

            # Update biometric check-in and check-out
            attendance.card_check_in = target_check_in
            attendance.card_check_out = target_check_out

            # Recalculate card working hours
            if attendance.card_check_in and attendance.card_check_out:
                total_seconds = (attendance.card_check_out - attendance.card_check_in).total_seconds()
                hours_decimal = max(total_seconds, 0) / 3600
                attendance.card_working_hours = int(hours_decimal * 100) / 100
            else:
                attendance.card_working_hours = 0.0

            # Recalculate status
            calculate_attendance_status(attendance)

            processed_count += 1

            # Emit Socket.IO updates for live dashboard
            try:
                from extensions import socketio
                web_in_str = attendance.check_in.strftime("%I:%M %p") if attendance.check_in else None
                web_out_str = attendance.check_out.strftime("%I:%M %p") if attendance.check_out else None
                card_in_str = attendance.card_check_in.strftime("%I:%M %p") if attendance.card_check_in else None
                card_out_str = attendance.card_check_out.strftime("%I:%M %p") if attendance.card_check_out else None

                # Determine UI status
                ui_status = "Absent"
                if attendance.check_in or attendance.card_check_in:
                    if attendance.check_in and not attendance.check_out:
                        ui_status = "Present"
                    elif attendance.check_out or attendance.card_check_out:
                        from datetime import date
                        is_today = (attendance.attendance_date == date.today())
                        if is_today and attendance.card_check_out and not attendance.check_out:
                            punch_out_hour = attendance.card_check_out.hour
                            working_hrs = attendance.card_working_hours or 0.0
                            if punch_out_hour >= 15 or working_hrs >= 4.0:
                                ui_status = "Checked Out"
                            else:
                                ui_status = "Present"
                        else:
                            ui_status = "Checked Out"
                    else:
                        ui_status = "Present"
                else:
                    ui_status = "Absent"

                payload = {
                    "id": employee.id,
                    "user_id": employee.user_id,
                    "attendance_status": ui_status,
                    "check_in": web_in_str,
                    "check_out": web_out_str,
                    "working_hours": attendance.total_hours or 0.0,
                    "card_check_in": card_in_str,
                    "card_check_out": card_out_str,
                    "card_working_hours": attendance.card_working_hours or 0.0,
                    "shift": employee.shift_timing or "General Shift",
                    "manager_status": attendance.manager_status or "Pending",
                    "checked_in": (attendance.check_in is not None and attendance.check_out is None),
                    "card_checked_in": (attendance.card_check_in is not None and attendance.card_check_out is None),
                }
                socketio.emit("attendance_update", payload)
            except Exception as se:
                print("Socket emit error during db sync:", se)

        db.session.commit()
        return jsonify({
            "success": True,
            "message": f"Successfully synced {processed_count} logs from Biometric DB"
        }), 200

    except Exception as e:
        db.session.rollback()
        print("DB SYNC TRIGGER ERROR:", str(e))
        return jsonify({
            "success": False,
            "error": f"Failed to sync database: {str(e)}"
        }), 500


@attendance_bp.route("/db-check", methods=["GET"])
def db_check_temp():
    try:
        from models.employee import Employee
        from models.attendance import Attendance
        employees = Employee.query.all()
        emp_list = []
        for emp in employees:
            emp_list.append({
                "id": emp.id,
                "user_id": emp.user_id,
                "name": f"{emp.first_name} {emp.last_name}",
                "manager": emp.reporting_manager,
                "employee_id": emp.employee_id
            })
        
        regs = Attendance.query.filter((Attendance.is_regularization == True) | (Attendance.manager_status == "Clarification Provided")).all()
        reg_list = []
        for rec in regs:
            reg_list.append({
                "id": rec.id,
                "user_id": rec.user_id,
                "date": str(rec.attendance_date),
                "check_in": str(rec.check_in),
                "check_out": str(rec.check_out),
                "manager_status": rec.manager_status,
                "is_regularization": rec.is_regularization,
                "reason": rec.regularization_reason
            })
            
        return jsonify({
            "employees": emp_list,
            "regularizations": reg_list
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# Pending Attendance — 25th→24th Cycle
# ---------------------------------------------------------------------------

@attendance_bp.route("/pending-cycle/<int:manager_user_id>", methods=["GET"])
def get_pending_cycle_attendance(manager_user_id):
    """
    Return all attendance records for the manager's reporting team that are
    still Pending within the current 25th-of-last-month → 24th-of-this-month
    payroll cycle (IST timezone).
    """
    try:
        import calendar

        # ── 1. Calculate cycle boundaries in IST (no third-party deps) ──────
        ist_today = datetime.now(ZoneInfo("Asia/Kolkata")).date()
        y, m, d = ist_today.year, ist_today.month, ist_today.day

        if d >= 25:
            cycle_start = ist_today.replace(day=25)
            # 24th of next month
            if m == 12:
                cycle_end = ist_today.replace(year=y + 1, month=1, day=24)
            else:
                cycle_end = ist_today.replace(month=m + 1, day=24)
        else:
            # 25th of previous month
            if m == 1:
                cycle_start = ist_today.replace(year=y - 1, month=12, day=25)
            else:
                cycle_start = ist_today.replace(month=m - 1, day=25)
            cycle_end = ist_today.replace(day=24)

        # ── 2. Resolve manager's full name ──────────────────────────────────
        manager_emp = Employee.query.filter_by(user_id=manager_user_id).first()
        if not manager_emp:
            # Try to look up via User table
            manager_user = User.query.filter_by(id=manager_user_id).first()
            manager_full_name = manager_user.full_name if manager_user else None
        else:
            manager_full_name = f"{manager_emp.first_name} {manager_emp.last_name}".strip()

        if not manager_full_name:
            return jsonify({"success": False, "error": "Manager not found"}), 404

        # ── 3. Find reporting team (Admins and HR see all, Managers see only their team)
        user = User.query.get(manager_user_id)
        is_admin = False
        if user:
            role_name = (user.role.name or "").lower() if user.role else ""
            access_level = (user.access_level or "").lower()
            if "admin" in role_name or "admin" in access_level or "hr" in role_name or "hr" in access_level:
                is_admin = True

        from routes.employees import is_manager_match
        all_employees = Employee.query.all()

        if is_admin:
            reporting_team = [emp for emp in all_employees if emp.user_id != manager_user_id]
        else:
            reporting_team = [
                emp for emp in all_employees
                if emp.user_id != manager_user_id and is_manager_match(emp.reporting_manager, manager_full_name)
            ]

        if not reporting_team:
            return jsonify({
                "success": True,
                "cycle_start": str(cycle_start),
                "cycle_end": str(cycle_end),
                "pending_count": 0,
                "records": []
            })

        team_user_ids = [emp.user_id for emp in reporting_team if emp.user_id]
        team_by_user_id = {emp.user_id: emp for emp in reporting_team}

        # Calculate the last working day (yesterday) to exclude it from cycle view (it is shown in Yesterday Summary card)
        last_working_day = get_last_working_day(ist_today - timedelta(days=1))

        # ── 4. Query pending attendance for the cycle (excluding today's active date and yesterday's summary date)
        pending_records = Attendance.query.filter(
            Attendance.user_id.in_(team_user_ids),
            Attendance.attendance_date <= cycle_end,
            Attendance.attendance_date < ist_today,
            Attendance.attendance_date != last_working_day,
            Attendance.manager_status == "Pending",
            Attendance.status != "Leave",
        ).order_by(Attendance.attendance_date.desc()).all()

        # ── 5. Serialise ────────────────────────────────────────────────────
        def fmt_time(dt):
            if not dt:
                return "—"
            try:
                return dt.strftime("%I:%M %p")
            except Exception:
                return str(dt)

        results = []
        for rec in pending_records:
            emp = team_by_user_id.get(rec.user_id)
            if not emp:
                continue

            from models.shift_request import ShiftRequest
            wages_req = ShiftRequest.query.filter(
                ShiftRequest.employee_id.in_([emp.id, emp.employee_id]),
                ShiftRequest.request_type == "One Day Wages",
                ShiftRequest.from_date <= rec.attendance_date,
                ShiftRequest.to_date >= rec.attendance_date
            ).first()
            wages_status = wages_req.status if wages_req else None
            is_one_day_wages_final = (wages_req is not None)

            results.append({
                "id":                   rec.id,
                "db_employee_id":       emp.id,
                "employee_id":          emp.employee_id or "",
                "employee_name":        f"{emp.first_name} {emp.last_name}".strip(),
                "department":           emp.department or "",
                "designation":          emp.designation or "",
                "attendance_date":      str(rec.attendance_date),
                "web_checkin":          fmt_time(rec.check_in),
                "web_checkout":         fmt_time(rec.check_out),
                "biometric_checkin":    fmt_time(rec.card_check_in),
                "biometric_checkout":   fmt_time(rec.card_check_out),
                "working_hours":        round(rec.total_hours or 0, 2),
                "status":               rec.status or "Absent",
                "manager_status":       rec.manager_status or "Pending",
                "check_in":             fmt_time(rec.check_in),
                "check_out":            fmt_time(rec.check_out),
                "lunch_minutes":        rec.lunch_minutes or 0,
                "tea_minutes":          rec.tea_minutes or 0,
                "total_break_minutes":  rec.total_break_minutes or 0,
                "clarification_history": rec.clarification_history or [],
                "reporting_manager":    manager_full_name,
                "is_one_day_wages":     is_one_day_wages_final,
                "wages_status":         wages_status,
            })

        return jsonify({
            "success":      True,
            "cycle_start":  str(cycle_start),
            "cycle_end":    str(cycle_end),
            "pending_count": len(results),
            "records":      results,
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500