from utils.compat import Blueprint, request, jsonify
from models.database import db
from models.leave import LeaveRequest
from datetime import datetime
from models.employee import Employee
from utils.employee_cache import get_all_employees_cached
from openpyxl import Workbook
from utils.compat import send_file
from io import BytesIO
from utils.jwt_helper import jwt_required, get_jwt_identity
from models.user import User

from services.leave_balance_service import (
    update_all_employee_leave_balances
)

leave_bp = Blueprint(
    "leave",
    __name__
)

def is_date_week_off(d):
    if d.weekday() == 6:
        return True
    if d.weekday() == 5:
        sat_count = 0
        for day_num in range(1, d.day + 1):
            from datetime import date
            if date(d.year, d.month, day_num).weekday() == 5:
                sat_count += 1
        if sat_count in (2, 4):
            return True
    return False

def get_cancellable_dates(from_date, to_date):
    from models.holiday import Holiday, HolidayOverride
    from datetime import timedelta
    
    # Pre-fetch overrides and published holidays
    overrides = {o.date: o.override_type for o in HolidayOverride.query.filter(HolidayOverride.date >= from_date, HolidayOverride.date <= to_date).all()}
    holidays = {h.date for h in Holiday.query.filter(Holiday.date >= from_date, Holiday.date <= to_date, Holiday.is_published == True).all()}
    
    valid_dates = []
    curr = from_date
    while curr <= to_date:
        override_type = overrides.get(curr)
        is_holiday = False
        is_week_off = False
        
        if override_type:
            if override_type == "Holiday":
                is_holiday = True
            elif override_type == "Weekly Off":
                is_week_off = True
        else:
            if curr in holidays:
                is_holiday = True
            elif is_date_week_off(curr):
                is_week_off = True
                
        if not (is_holiday or is_week_off):
            valid_dates.append(curr.strftime("%Y-%m-%d"))
            
        curr += timedelta(days=1)
        
    return valid_dates

def serialize_leave(leave):
    emp_string_id = leave.employee_id
    try:
        if emp_string_id and str(emp_string_id).isdigit():
            emp = Employee.query.get(int(emp_string_id))
            if emp and emp.employee_id:
                emp_string_id = emp.employee_id
    except:
        pass

    leave_duration = "Full Day"
    if leave.total_days and leave.total_days <= 0.5:
        if leave.reason and " (First Half)" in leave.reason:
            leave_duration = "First Half"
        elif leave.reason and " (Second Half)" in leave.reason:
            leave_duration = "Second Half"
        else:
            leave_duration = "First Half"

    return {
        "id": leave.id,
        "employee_id": emp_string_id,
        "employee_name": leave.employee_name,
        "leave_type": leave.leave_type,
        "from_date": str(leave.from_date) if leave.from_date else None,
        "to_date": str(leave.to_date) if leave.to_date else None,
        "total_days": leave.total_days,
        "leave_duration": leave_duration,
        "reporting_manager": leave.reporting_manager,
        "handover_to": leave.handover_to,
        "emergency_contact": leave.emergency_contact,
        "reason": leave.reason,
        "status": leave.status,
        "request_type": leave.request_type,
        "permission_date": str(leave.permission_date) if leave.permission_date else None,
        "from_time": str(leave.from_time) if leave.from_time else None,
        "to_time": str(leave.to_time) if leave.to_time else None,
        "cancelled_at": leave.cancelled_at.isoformat() + "Z" if hasattr(leave, "cancelled_at") and leave.cancelled_at else None,
        "cancelled_by": leave.cancelled_by if hasattr(leave, "cancelled_by") else None,
        "cancellation_reason": leave.cancellation_reason if hasattr(leave, "cancellation_reason") else None,
        "created_at": leave.created_at.isoformat() + "Z" if hasattr(leave, "created_at") and leave.created_at else None,
        "approved_at": leave.approved_at.isoformat() + "Z" if hasattr(leave, "approved_at") and leave.approved_at else None,
        "rejected_at": leave.rejected_at.isoformat() + "Z" if hasattr(leave, "rejected_at") and leave.rejected_at else None,
        "cancelled_dates": leave.cancelled_dates or [],
        "cancellable_dates": get_cancellable_dates(leave.from_date, leave.to_date) if (leave.request_type == "Leave" and leave.from_date and leave.to_date) else [],
    }

@leave_bp.route("/", methods=["POST"])
def apply_leave():

    try:

        data = request.get_json()

        emp_id = data.get("employee_id")
        if not emp_id:
            return jsonify({"success": False, "error": "Employee ID is required."}), 400

        from models.employee import Employee
        if str(emp_id).isdigit():
            employee = Employee.query.filter(
                (Employee.id == int(emp_id)) | (Employee.employee_id == str(emp_id))
            ).first()
        else:
            employee = Employee.query.filter(Employee.employee_id == str(emp_id)).first()

        request_type = data.get("request_type", "Leave")

        leave = LeaveRequest(
            employee_id=data.get("employee_id"),
            employee_name=data.get("employee_name"),
            request_type=request_type,
            leave_type=data.get("leave_type"),
            reporting_manager=data.get("reporting_manager"),
            handover_to=data.get("handover_to"),
            reason=data.get("reason")
        )

        # ===========================
        # LEAVE REQUEST
        # ===========================
        if request_type == "Leave":

            leave.from_date = datetime.strptime(
                data.get("from_date"),
                "%Y-%m-%d"
            ).date()

            leave.to_date = datetime.strptime(
                data.get("to_date"),
                "%Y-%m-%d"
            ).date()

            leave.total_days = float(data.get("total_days", 0))

            if data.get("from_time"):
                leave.from_time = datetime.strptime(data.get("from_time"), "%H:%M").time()
            if data.get("to_time"):
                leave.to_time = datetime.strptime(data.get("to_time"), "%H:%M").time()

            # Query balance at time of application to pre-mark as LOP if balance is 0
            leave_type_lower = (leave.leave_type or "").strip().lower()
            from models.leave import EmployeeLeaveBalance
            from sqlalchemy import func
            balance = EmployeeLeaveBalance.query.filter(
                EmployeeLeaveBalance.employee_id == employee.id if employee else None,
                func.lower(EmployeeLeaveBalance.leave_type) == leave_type_lower
            ).first()

            available_balance = 0.0
            if balance:
                available_balance = float(balance.available or 0.0)

            if available_balance <= 0.0:
                leave.leave_type = f"{leave.leave_type} (Loss of Pay)"

        # ===========================
        # PERMISSION REQUEST
        # ===========================
        elif request_type == "Permission":

            leave.permission_date = datetime.strptime(
                data.get("permission_date"),
                "%Y-%m-%d"
            ).date()

            leave.from_time = datetime.strptime(
                data.get("from_time"),
                "%H:%M"
            ).time()

            leave.to_time = datetime.strptime(
                data.get("to_time"),
                "%H:%M"
            ).time()

            # Ensure minimum duration is 1 hour
            from datetime import timedelta
            dt_from = datetime.combine(leave.permission_date, leave.from_time)
            dt_to = datetime.combine(leave.permission_date, leave.to_time)
            if dt_to < dt_from:
                dt_to += timedelta(days=1)
            duration_hours = (dt_to - dt_from).total_seconds() / 3600.0
            duration_minutes = (dt_to - dt_from).total_seconds() / 60.0
            
            if abs(duration_minutes - 60.0) > 0.1:
                return jsonify({
                    "success": False,
                    "error": "Permission duration must be exactly 1 hour."
                }), 400

            # Validate against database permission balance row
            from models.leave import EmployeeLeaveBalance
            from sqlalchemy import func
            perm_bal = EmployeeLeaveBalance.query.filter(
                EmployeeLeaveBalance.employee_id == employee.id if employee else None,
                func.lower(EmployeeLeaveBalance.leave_type) == "permission"
            ).first()

            available_hours = float(perm_bal.available) if perm_bal else 2.0
            if duration_hours > available_hours:
                return jsonify({
                    "success": False,
                    "error": f"Applying this permission would exceed your remaining monthly permission limit. You have {available_hours:.2f} hours remaining."
                }), 400

            leave.total_days = 0

        db.session.add(leave)
        db.session.commit()

        # Send manager notification email
        try:
            from services.request_email_service import send_manager_request_email
            send_manager_request_email(leave, request_type)
        except Exception as email_err:
            print("Failed to send manager request email:", str(email_err))

        # Emit leave_update socket event for real-time dashboard updates
        try:
            from extensions import socketio
            socketio.emit("leave_update", serialize_leave(leave))
        except Exception as socket_err:
            print("Failed to emit leave socket:", str(socket_err))

        return jsonify({
            "success": True,
            "message": f"{request_type} Applied Successfully"
        }), 200

    except Exception as e:

        import traceback
        traceback.print_exc()
        db.session.rollback()

        import traceback
        traceback.print_exc()

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500
    
@leave_bp.route("/", methods=["GET"])
def get_leaves():

    try:

        leaves = LeaveRequest.query.order_by(
            LeaveRequest.id.desc()
        ).all()


        leaves_data = [serialize_leave(leave) for leave in leaves]
        return jsonify(leaves_data), 200

    except Exception as e:
        import traceback
        traceback.print_exc()

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500
    

@leave_bp.route(
    "/approve/<int:leave_id>",
    methods=["PUT"]
)
@jwt_required()
def approve_leave(leave_id):

    try:

        leave = LeaveRequest.query.get(leave_id)

        if not leave:
            return jsonify({
                "success": False,
                "error": "Leave not found"
            }), 404

        # Prevent double approval/rejection
        if leave.status != "Pending":
            return jsonify({
                "success": False,
                "error": f"Leave already {leave.status.lower()}"
            }), 400

        print("Leave Employee ID:", leave.employee_id)

        employee = Employee.query.filter(
            (Employee.id == int(leave.employee_id)) | (Employee.employee_id == str(leave.employee_id))
        ).first()

        print("Employee Found:", employee)

        if not employee:
            return jsonify({
                "success": False,
                "error": "Employee not found"
            }), 404

        # ===========================
        # PERMISSION REQUEST
        # ===========================
        
        user_id = get_jwt_identity()
        approver = User.query.get(int(user_id)) if user_id else None
        approver_name = approver.full_name if approver else "Manager"
        
        if leave.request_type == "Permission":
            leave.status = "Approved"
            leave.approved_by = approver_name
            leave.approved_at = datetime.utcnow()

            # Deduct from permission balance row in DB
            from models.leave import EmployeeLeaveBalance
            from sqlalchemy import func
            perm_bal = EmployeeLeaveBalance.query.filter(
                EmployeeLeaveBalance.employee_id == employee.id,
                func.lower(EmployeeLeaveBalance.leave_type) == "permission"
            ).first()
            if perm_bal:
                ft = leave.from_time
                tt = leave.to_time
                perm_seconds = (tt.hour * 3600 + tt.minute * 60) - (ft.hour * 3600 + ft.minute * 60)
                perm_hours = max(perm_seconds, 0) / 3600.0
                perm_bal.available = max(0.0, (perm_bal.available or 0.0) - perm_hours)

            # Recalculate attendance status if it exists
            from models.attendance import Attendance
            from routes.attendance import calculate_attendance_status
            att = Attendance.query.filter_by(
                user_id=employee.user_id,
                attendance_date=leave.permission_date
            ).first()
            if att:
                calculate_attendance_status(att)

            db.session.commit()

            # Emit leave_update socket event for real-time dashboard updates
            try:
                from extensions import socketio
                socketio.emit("leave_update", serialize_leave(leave))
            except Exception as socket_err:
                print("Failed to emit leave socket:", str(socket_err))

            # Send employee status email
            try:
                from services.request_email_service import send_employee_status_email
                send_employee_status_email(leave, employee, "Approved", leave.request_type)
            except Exception as email_err:
                print("Failed to send employee status email:", str(email_err))

            return jsonify({
                "success": True,
                "message": "Permission Approved Successfully"
            }), 200

        # ===========================
        # LEAVE REQUEST
        # ===========================

        leave_type = (leave.leave_type or "").strip().lower()
        if " (loss of pay)" in leave_type:
            leave_type = leave_type.replace(" (loss of pay)", "").strip()
        elif " (lop)" in leave_type:
            leave_type = leave_type.replace(" (lop)", "").strip()

        leave_days = leave.total_days or 0

        # Try to find corresponding EmployeeLeaveBalance
        from models.leave import EmployeeLeaveBalance
        from sqlalchemy import func
        balance = EmployeeLeaveBalance.query.filter(
            EmployeeLeaveBalance.employee_id == employee.id,
            func.lower(EmployeeLeaveBalance.leave_type) == leave_type
        ).first()

        # For dynamic custom leave categories, check if we found a balance record
        if leave_type not in ["sick leave", "casual leave", "earned leave", "privilege leave"] and not balance:
            return jsonify({
                "success": False,
                "error": f"Invalid leave type or no balance record found: {leave.leave_type}"
            }), 400

        available_balance = 0.0
        if balance:
            available_balance = float(balance.available or 0.0)

        lop_leave = None

        if available_balance <= 0.0:
            # Entire leave is LOP
            leave.leave_type = f"{leave.leave_type} (Loss of Pay)"
            leave.status = "Approved"
            leave.approved_by = approver_name
            leave.approved_at = datetime.utcnow()
        elif leave_days > available_balance:
            # Split the leave request:
            # 1. First part (Paid): up to available_balance
            import math
            from datetime import timedelta
            paid_days = available_balance
            lop_days = leave_days - available_balance
            
            # Save original to_date
            original_to_date = leave.to_date
            
            # Update current request to represent the paid portion
            leave.status = "Approved"
            leave.approved_by = approver_name
            leave.approved_at = datetime.utcnow()
            leave.total_days = paid_days
            
            paid_days_count = int(math.ceil(paid_days))
            new_to_date = leave.from_date + timedelta(days=max(0, paid_days_count - 1))
            if new_to_date > original_to_date:
                new_to_date = original_to_date
            leave.to_date = new_to_date
            
            # Deduct the balance to 0
            if balance:
                balance.available = 0.0
            if leave_type == "sick leave":
                employee.sick_leave = 0.0
            elif leave_type == "casual leave":
                employee.casual_leave = 0.0
            elif leave_type in ["earned leave", "privilege leave"]:
                employee.privilege_leave = 0.0
                
            # 2. Second part (LOP): new request for remainder
            lop_from_date = new_to_date + timedelta(days=1)
            if lop_from_date > original_to_date:
                lop_from_date = original_to_date
                
            lop_leave = LeaveRequest(
                employee_id=leave.employee_id,
                employee_name=leave.employee_name,
                request_type="Leave",
                leave_type=f"{leave.leave_type} (Loss of Pay)",
                from_date=lop_from_date,
                to_date=original_to_date,
                total_days=lop_days,
                status="Approved",
                reason=(leave.reason or "") + " (Loss of Pay split)",
                reporting_manager=leave.reporting_manager,
                handover_to=leave.handover_to,
                emergency_contact=leave.emergency_contact,
                approved_by=approver_name,
                approved_at=datetime.utcnow()
            )
            db.session.add(lop_leave)
        else:
            # Normal deduction (leave fits within balance)
            leave.status = "Approved"
            leave.approved_by = approver_name
            leave.approved_at = datetime.utcnow()

            if balance:
                balance.available = max(0.0, (balance.available or 0.0) - leave_days)

            if leave_type == "sick leave":
                employee.sick_leave = max(
                    0.0,
                    (employee.sick_leave or 0.0) - leave_days
                )
            elif leave_type == "casual leave":
                employee.casual_leave = max(
                    0.0,
                    (employee.casual_leave or 0.0) - leave_days
                )
            elif leave_type in ["earned leave", "privilege leave"]:
                employee.privilege_leave = max(
                    0.0,
                    (employee.privilege_leave or 0.0) - leave_days
                )

        db.session.commit()

        # Emit leave_update socket event for real-time dashboard updates
        try:
            from extensions import socketio
            socketio.emit("leave_update", serialize_leave(leave))
            if lop_leave:
                socketio.emit("leave_update", serialize_leave(lop_leave))
        except Exception as socket_err:
            print("Failed to emit leave socket:", str(socket_err))

        return jsonify({
            "success": True,
            "message": "Leave Approved Successfully",
            "leave_balance": {
                "sick_leave": employee.sick_leave,
                "casual_leave": employee.casual_leave,
                "privilege_leave": employee.privilege_leave,
                "earned_leave": employee.privilege_leave,
                "total_balance":
                    (employee.sick_leave or 0) +
                    (employee.casual_leave or 0) +
                    (employee.privilege_leave or 0)
            }
        }), 200

    except Exception as e:

        import traceback
        traceback.print_exc()
        db.session.rollback()

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@leave_bp.route(
    "/reject/<int:leave_id>",
    methods=["PUT"]
)
@jwt_required()
def reject_leave(leave_id):

    leave = LeaveRequest.query.get(leave_id)

    if not leave:
        return jsonify({
            "error": "Leave not found"
        }), 404

    # Prevent double approval/rejection
    if leave.status != "Pending":
        return jsonify({
            "success": False,
            "error": f"Leave already {leave.status.lower()}"
        }), 400

    employee = Employee.query.filter(
        (Employee.id == int(leave.employee_id)) | (Employee.employee_id == str(leave.employee_id))
    ).first()

    leave.status = "Rejected"
    
    user_id = get_jwt_identity()
    rejecter = User.query.get(int(user_id)) if user_id else None
    
    leave.rejected_by = rejecter.full_name if rejecter else "Manager"
    leave.rejected_at = datetime.utcnow()

    db.session.commit()

    # Send employee status email
    if employee:
        try:
            from services.request_email_service import send_employee_status_email
            send_employee_status_email(leave, employee, "Rejected", leave.request_type)
        except Exception as email_err:
            print("Failed to send employee status email:", str(email_err))

    # Emit leave_update socket event for real-time dashboard updates
    try:
        from extensions import socketio
        socketio.emit("leave_update", serialize_leave(leave))
    except Exception as socket_err:
        print("Failed to emit leave socket:", str(socket_err))

    return jsonify({
        "success": True,
        "message": "Leave Rejected"
    })


@leave_bp.route(
    "/cancel/<int:leave_id>",
    methods=["PUT"]
)
def cancel_leave(leave_id):

    try:

        leave = LeaveRequest.query.get(
            leave_id
        )

        if not leave:
            return jsonify({
                "success": False,
                "error": "Leave not found"
            }), 404

        if leave.status == "Cancelled":
            return jsonify({
                "success": False,
                "error": "Already cancelled"
            }), 400

        from zoneinfo import ZoneInfo
        from datetime import datetime
        
        today = datetime.now(ZoneInfo("Asia/Kolkata")).date()
        start_date = leave.permission_date if leave.request_type == "Permission" else leave.from_date
        
        # Date constraint bypassed to allow manager cancellations of past/active requests

        previous_status = leave.status
        leave.status = "Cancelled"
        
        if previous_status == "Approved" and leave.request_type == "Leave":
            if leave.employee_id and str(leave.employee_id).isdigit():
                employee = Employee.query.filter(
                    (Employee.id == int(leave.employee_id)) | (Employee.employee_id == str(leave.employee_id))
                ).first()
            else:
                employee = Employee.query.filter(
                    Employee.employee_id == str(leave.employee_id)
                ).first()

            if employee:
                leave_type = (leave.leave_type or "").strip().lower()
                leave_days = leave.total_days or 0

                if "loss of pay" in leave_type or "lop" in leave_type:
                    # Do not restore balance for LOP leaves
                    pass
                else:
                    if leave_type == "sick leave":
                        employee.sick_leave = (employee.sick_leave or 0) + leave_days
                    elif leave_type == "casual leave":
                        employee.casual_leave = (employee.casual_leave or 0) + leave_days
                    elif leave_type in ["earned leave", "privilege leave"]:
                        employee.privilege_leave = (employee.privilege_leave or 0) + leave_days
                    # Restore EmployeeLeaveBalance
                    from models.leave import EmployeeLeaveBalance
                    from sqlalchemy import func
                    balance = EmployeeLeaveBalance.query.filter(
                        EmployeeLeaveBalance.employee_id == employee.id,
                        func.lower(EmployeeLeaveBalance.leave_type) == leave_type
                    ).first()
                    if balance:
                        balance.available = (balance.available or 0) + leave_days

                # Restore attendance records for dates covered by the leave
                if leave.from_date and leave.to_date:
                    from models.attendance import Attendance
                    from routes.attendance import sync_biometric_to_web_entry
                    from datetime import timedelta

                    curr_date = leave.from_date
                    while curr_date <= leave.to_date:
                        att = Attendance.query.filter_by(
                            user_id=employee.user_id,
                            attendance_date=curr_date
                        ).first()
                        if att and att.status == "Leave":
                            has_punches = bool(att.check_in or att.card_check_in or att.check_out or att.card_check_out)
                            if not has_punches:
                                db.session.delete(att)
                            else:
                                att.status = "Absent"
                                att.leave_type = None
                                sync_biometric_to_web_entry(att)
                        curr_date += timedelta(days=1)

                # Check if there is a related split request and cancel it too
                from datetime import timedelta
                related_split = None
                if leave.reason and " (Loss of Pay split)" in leave.reason:
                    paid_reason = leave.reason.replace(" (Loss of Pay split)", "")
                    related_split = LeaveRequest.query.filter(
                        LeaveRequest.employee_id == leave.employee_id,
                        LeaveRequest.to_date == leave.from_date - timedelta(days=1),
                        LeaveRequest.reason == paid_reason,
                        LeaveRequest.status.in_(["Approved", "Pending"])
                    ).first()
                elif leave.reason:
                    lop_reason = leave.reason + " (Loss of Pay split)"
                    related_split = LeaveRequest.query.filter(
                        LeaveRequest.employee_id == leave.employee_id,
                        LeaveRequest.from_date == leave.to_date + timedelta(days=1),
                        LeaveRequest.reason == lop_reason,
                        LeaveRequest.status.in_(["Approved", "Pending"])
                    ).first()

                if related_split:
                    prev_status_related = related_split.status
                    related_split.status = "Cancelled"
                    related_split.cancelled_by = "Manager"
                    related_split.cancelled_at = datetime.now(ZoneInfo("Asia/Kolkata"))

                    # Restore balance for the related split request if it was Approved
                    if prev_status_related == "Approved" and related_split.request_type == "Leave":
                        related_type = (related_split.leave_type or "").strip().lower()
                        related_days = related_split.total_days or 0

                        if "loss of pay" not in related_type and "lop" not in related_type:
                            if related_type == "sick leave":
                                employee.sick_leave = (employee.sick_leave or 0) + related_days
                            elif related_type == "casual leave":
                                employee.casual_leave = (employee.casual_leave or 0) + related_days
                            elif related_type in ["earned leave", "privilege leave"]:
                                employee.privilege_leave = (employee.privilege_leave or 0) + related_days

                            # Restore EmployeeLeaveBalance
                            from models.leave import EmployeeLeaveBalance
                            from sqlalchemy import func
                            bal_related = EmployeeLeaveBalance.query.filter(
                                EmployeeLeaveBalance.employee_id == employee.id,
                                func.lower(EmployeeLeaveBalance.leave_type) == related_type
                            ).first()
                            if bal_related:
                                bal_related.available = (bal_related.available or 0) + related_days

                    # Restore attendance records for the related split request
                    if related_split.from_date and related_split.to_date:
                        curr_date = related_split.from_date
                        while curr_date <= related_split.to_date:
                            att = Attendance.query.filter_by(
                                user_id=employee.user_id,
                                attendance_date=curr_date
                            ).first()
                            if att and att.status == "Leave":
                                has_punches = bool(att.check_in or att.card_check_in or att.check_out or att.card_check_out)
                                if not has_punches:
                                    db.session.delete(att)
                                else:
                                    att.status = "Absent"
                                    att.leave_type = None
                                    sync_biometric_to_web_entry(att)
                            curr_date += timedelta(days=1)

                    # Emit SocketIO update for the related request
                    try:
                        from extensions import socketio
                        socketio.emit("leave_update", serialize_leave(related_split))
                    except Exception as socket_err:
                        print("Failed to emit related leave socket:", str(socket_err))

        db.session.commit()

        # Emit leave_update socket event for real-time dashboard updates
        try:
            from extensions import socketio
            socketio.emit("leave_update", serialize_leave(leave))
            if employee and employee.user_id:
                socketio.emit("attendance_update", {"user_id": employee.user_id})
        except Exception as socket_err:
            print("Failed to emit leave socket:", str(socket_err))

        return jsonify({
            "success": True,
            "message": "Leave Cancelled Successfully"
        })

    except Exception as e:

        import traceback
        traceback.print_exc()
        db.session.rollback()

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@leave_bp.route(
    "/cancel-date/<int:leave_id>",
    methods=["PUT"]
)
@jwt_required()
def cancel_leave_date(leave_id):
    try:
        from zoneinfo import ZoneInfo
        from datetime import datetime, date, timedelta
        from sqlalchemy.orm.attributes import flag_modified
        from models.leave import LeaveRequest, EmployeeLeaveBalance, LeaveAuditLog
        from models.employee import Employee
        from models.user import User
        from models.attendance import Attendance
        from routes.attendance import sync_biometric_to_web_entry

        leave = LeaveRequest.query.get(leave_id)
        if not leave:
            return jsonify({
                "success": False,
                "error": "Leave not found"
            }), 404

        # Validate request is approved
        if leave.status != "Approved":
            return jsonify({
                "success": False,
                "error": "Only approved leave requests can be cancelled date-wise"
            }), 400

        # Validate that it is of request_type "Leave"
        if leave.request_type != "Leave":
            return jsonify({
                "success": False,
                "error": "Only leave requests can be cancelled date-wise"
            }), 400

        data = request.get_json() or {}
        dates_list = data.get("dates")
        
        # Backwards compatibility for single "date" parameter
        if not dates_list:
            single_date = (data.get("date") or "").strip()
            if single_date:
                dates_list = [single_date]
                
        if not dates_list or not isinstance(dates_list, list):
            return jsonify({
                "success": False,
                "error": "At least one date is required in 'dates' list"
            }), 400

        # Parse and validate dates
        parsed_dates = []
        for date_str in dates_list:
            date_str = date_str.strip()
            if not date_str:
                continue
            try:
                d_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
                parsed_dates.append((date_str, d_obj))
            except ValueError:
                return jsonify({
                    "success": False,
                    "error": f"Invalid date format: {date_str}. Expected YYYY-MM-DD"
                }), 400

        if not parsed_dates:
            return jsonify({
                "success": False,
                "error": "No valid dates provided"
            }), 400

        # Get current date in Asia/Kolkata
        tz = ZoneInfo("Asia/Kolkata")
        today = datetime.now(tz).date()

        # Backend constraint validation: selected_date >= current_date, in request range, not already cancelled
        cancelled_dates = leave.cancelled_dates or []
        for date_str, selected_date in parsed_dates:
            if selected_date < today:
                return jsonify({
                    "success": False,
                    "error": f"Cannot cancel leave date {date_str} in the past. Date must be today or in the future."
                }), 400

            if not (leave.from_date <= selected_date <= leave.to_date):
                return jsonify({
                    "success": False,
                    "error": f"Date {date_str} is not within the approved leave request range."
                }), 400

            if date_str in cancelled_dates:
                return jsonify({
                    "success": False,
                    "error": f"Date {date_str} has already been cancelled"
                }), 400

        # Validate manager authorization
        user_id = get_jwt_identity()
        approver = User.query.get(int(user_id)) if user_id else None
        if not approver:
            return jsonify({
                "success": False,
                "error": "Approver user not found"
            }), 404

        employee = Employee.query.filter(
            (Employee.id == int(leave.employee_id)) | (Employee.employee_id == str(leave.employee_id))
        ).first()

        if not employee:
            return jsonify({
                "success": False,
                "error": "Employee associated with leave request not found"
            }), 404

        # Check if the manager is authorized
        approver_name = approver.full_name or "Manager"
        
        def is_manager_match(manager_field, approver_name_str):
            if not manager_field:
                return False
            m_field = manager_field.strip().lower()
            a_name = approver_name_str.strip().lower()
            if m_field == a_name:
                return True
            m_parts = m_field.split()
            a_parts = a_name.split()
            if len(m_parts) == 1 and len(a_parts) > 0 and a_parts[0] == m_parts[0]:
                return True
            if len(a_parts) == 1 and len(m_parts) > 0 and m_parts[0] == a_parts[0]:
                return True
            return False

        is_authorized = (
            approver.access_level.lower() == "admin" or
            is_manager_match(employee.reporting_manager, approver_name) or
            is_manager_match(leave.reporting_manager, approver_name) or
            is_manager_match(leave.handover_to, approver_name)
        )

        if not is_authorized:
            return jsonify({
                "success": False,
                "error": "Unauthorized. You are not authorized to cancel this employee's leave."
            }), 403

        # Process cancellation for each selected date
        if leave.cancelled_dates is None:
            leave.cancelled_dates = []

        total_days_restored = 0.0

        for date_str, selected_date in parsed_dates:
            leave.cancelled_dates.append(date_str)
            
            # Determine days to restore for this date
            if leave.from_date == leave.to_date and leave.total_days == 0.5:
                days_to_restore = 0.5
            else:
                days_to_restore = 1.0
            
            total_days_restored += days_to_restore

            # Delete/update Attendance for that date
            att = Attendance.query.filter_by(
                user_id=employee.user_id,
                attendance_date=selected_date
            ).first()
            if att and att.status == "Leave":
                has_punches = bool(att.check_in or att.card_check_in or att.check_out or att.card_check_out)
                if not has_punches:
                    db.session.delete(att)
                else:
                    att.status = "Absent"
                    att.leave_type = None
                    sync_biometric_to_web_entry(att)

            # Log Leave Audit
            audit = LeaveAuditLog(
                leave_id=leave.id,
                employee_name=leave.employee_name,
                action=f"Manager cancelled leave date: {date_str}",
                previous_status="Approved",
                new_status="Approved",
                cancelled_at=datetime.utcnow(),
                cancelled_by=approver_name
            )
            db.session.add(audit)

        flag_modified(leave, "cancelled_dates")

        # Deduct total days from request
        leave.total_days = max(0.0, (leave.total_days or 0) - total_days_restored)

        # Restore balance on Employee & EmployeeLeaveBalance
        leave_type = (leave.leave_type or "").strip().lower()
        if "loss of pay" in leave_type or "lop" in leave_type:
            # Do not restore balance for LOP leaves
            pass
        else:
            if leave_type == "sick leave":
                employee.sick_leave = (employee.sick_leave or 0) + total_days_restored
            elif leave_type == "casual leave":
                employee.casual_leave = (employee.casual_leave or 0) + total_days_restored
            elif leave_type in ["earned leave", "privilege leave"]:
                employee.privilege_leave = (employee.privilege_leave or 0) + total_days_restored

            from sqlalchemy import func
            balance = EmployeeLeaveBalance.query.filter(
                EmployeeLeaveBalance.employee_id == employee.id,
                func.lower(EmployeeLeaveBalance.leave_type) == leave_type
            ).first()
            if balance:
                balance.available = (balance.available or 0) + total_days_restored

        # If all dates of range are cancelled, mark the entire request as Cancelled
        cancellable_dates_in_range = get_cancellable_dates(leave.from_date, leave.to_date)
        all_cancellable_cancelled = all(d in leave.cancelled_dates for d in cancellable_dates_in_range)
        
        if all_cancellable_cancelled or leave.total_days <= 0:
            leave.status = "Cancelled"
            leave.cancelled_by = approver_name
            leave.cancelled_at = datetime.utcnow()
            leave.cancellation_reason = f"All dates cancelled by manager. Last cancelled: {', '.join(dates_list)}"

        db.session.commit()

        # Emit update
        try:
            from extensions import socketio
            socketio.emit("leave_update", serialize_leave(leave))
            if employee and employee.user_id:
                socketio.emit("attendance_update", {"user_id": employee.user_id})
        except Exception as socket_err:
            print("Failed to emit leave socket:", str(socket_err))

        return jsonify({
            "success": True,
            "message": f"Successfully cancelled leaves for {', '.join(dates_list)}",
            "leave": serialize_leave(leave)
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        db.session.rollback()
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@leave_bp.route(
    "/update/<int:leave_id>",
    methods=["PUT"]
)
def update_leave(leave_id):

    try:

        leave = LeaveRequest.query.get(leave_id)

        if not leave:
            return jsonify({
                "success": False,
                "error": "Request not found"
            }), 404

        data = request.get_json()

        # Common fields
        leave.reason = data.get("reason", leave.reason)
        leave.handover_to = data.get(
            "handover_to",
            leave.handover_to
        )

        # ===========================
        # UPDATE LEAVE
        # ===========================
        if leave.request_type == "Leave":

            leave.leave_type = data.get("leave_type")

            leave.from_date = datetime.strptime(
                data.get("from_date"),
                "%Y-%m-%d"
            ).date()

            leave.to_date = datetime.strptime(
                data.get("to_date"),
                "%Y-%m-%d"
            ).date()

            leave.total_days = float(
                data.get("total_days", 0)
            )

            # Clear permission fields but retain/update timing if present
            leave.permission_date = None
            if data.get("from_time"):
                leave.from_time = datetime.strptime(data.get("from_time"), "%H:%M").time()
            else:
                leave.from_time = None
            if data.get("to_time"):
                leave.to_time = datetime.strptime(data.get("to_time"), "%H:%M").time()
            else:
                leave.to_time = None

        # ===========================
        # UPDATE PERMISSION
        # ===========================
        elif leave.request_type == "Permission":

            leave.permission_date = datetime.strptime(
                data.get("permission_date"),
                "%Y-%m-%d"
            ).date()

            leave.from_time = datetime.strptime(
                data.get("from_time"),
                "%H:%M"
            ).time()

            leave.to_time = datetime.strptime(
                data.get("to_time"),
                "%H:%M"
            ).time()

            # Clear leave fields
            leave.leave_type = None
            leave.from_date = None
            leave.to_date = None
            leave.total_days = 0

        db.session.commit()

        return jsonify({
            "success": True,
            "message": f"{leave.request_type} Updated Successfully"
        }), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        db.session.rollback()
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500
@leave_bp.route(
    "/<int:leave_id>/cancel",
    methods=["POST"]
)
def cancel_approved_leave(leave_id):
    try:
        from zoneinfo import ZoneInfo
        from models.notification import Notification
        from models.leave import LeaveAuditLog

        leave = LeaveRequest.query.get(leave_id)
        if not leave:
            return jsonify({
                "success": False,
                "message": "Leave request not found"
            }), 404

        data = request.get_json() or {}
        employee_id = data.get("employee_id")

        # 1. Validate employee ownership
        if not employee_id:
            return jsonify({
                "success": False,
                "message": "employee_id is required to validate ownership"
            }), 400

        # Resolve employee details to match either primary key ID or employee_id string
        from models.employee import Employee
        from sqlalchemy import or_

        # Safe integer conversion to avoid db comparison issues with strings
        emp_id_int = None
        if isinstance(employee_id, int):
            emp_id_int = employee_id
        elif isinstance(employee_id, str) and employee_id.isdigit():
            emp_id_int = int(employee_id)

        filters = [Employee.employee_id == str(employee_id)]
        if emp_id_int is not None:
            filters.append(Employee.id == emp_id_int)

        employee = Employee.query.filter(or_(*filters)).first()

        # Check if ownership is valid (matches database ID, employee code, or raw passed ID)
        valid_ids = {str(employee_id).strip().lower()}
        if employee:
            if employee.id:
                valid_ids.add(str(employee.id))
            if employee.employee_id:
                valid_ids.add(str(employee.employee_id).strip().lower())
            if employee.user_id:
                valid_ids.add(str(employee.user_id))

        leave_emp_id = str(leave.employee_id).strip().lower() if leave.employee_id else ""

        if leave_emp_id not in valid_ids:
            return jsonify({
                "success": False,
                "message": "You do not own this leave request."
            }), 403

        # 2. Validate leave status is Approved or Pending
        if leave.status not in ["Approved", "Pending"]:
            return jsonify({
                "success": False,
                "message": "Only approved or pending leave requests can be cancelled."
            }), 400

        # 3. Validate current date is before leave start date
        today = datetime.now(ZoneInfo("Asia/Kolkata")).date()
        start_date = leave.permission_date if leave.request_type == "Permission" else leave.from_date

        if not start_date:
            return jsonify({
                "success": False,
                "message": "Leave start date is missing."
            }), 400

        if leave.status == "Approved" and today >= start_date:
            return jsonify({
                "success": False,
                "message": "Leave cannot be cancelled once the leave start date has begun."
            }), 400

        # 4. Start database transaction (SQLAlchemy transaction auto-managed, committed below)
        previous_status = leave.status
        leave.status = "Cancelled"
        leave.cancelled_by = "Employee"
        leave.cancelled_at = datetime.now(ZoneInfo("Asia/Kolkata"))
        leave.cancellation_reason = data.get("cancellation_reason", "")

        # 5. Restore balance only if it was Approved
        if previous_status == "Approved":
            if leave.employee_id and str(leave.employee_id).isdigit():
                employee = Employee.query.filter(
                    (Employee.id == int(leave.employee_id)) | (Employee.employee_id == str(leave.employee_id))
                ).first()
            else:
                employee = Employee.query.filter(
                    Employee.employee_id == str(leave.employee_id)
                ).first()

            if not employee:
                return jsonify({
                    "success": False,
                    "message": "Employee record not found to restore balance."
                }), 404

            if leave.request_type == "Permission":
                # Restore database permission balance row
                from models.leave import EmployeeLeaveBalance
                from sqlalchemy import func
                perm_bal = EmployeeLeaveBalance.query.filter(
                    EmployeeLeaveBalance.employee_id == employee.id,
                    func.lower(EmployeeLeaveBalance.leave_type) == "permission"
                ).first()
                if perm_bal:
                    ft = leave.from_time
                    tt = leave.to_time
                    perm_seconds = (tt.hour * 3600 + tt.minute * 60) - (ft.hour * 3600 + ft.minute * 60)
                    perm_hours = max(perm_seconds, 0) / 3600.0
                    perm_bal.available = (perm_bal.available or 0.0) + perm_hours

                # Recalculate attendance status if it exists
                from models.attendance import Attendance
                from routes.attendance import calculate_attendance_status
                att = Attendance.query.filter_by(
                    user_id=employee.user_id,
                    attendance_date=leave.permission_date
                ).first()
                if att:
                    calculate_attendance_status(att)

            elif leave.request_type == "Leave":
                leave_type = (leave.leave_type or "").strip().lower()
                leave_days = leave.total_days or 0

                if "loss of pay" in leave_type or "lop" in leave_type:
                    # Do not restore balance for LOP leaves
                    pass
                else:
                    if leave_type == "sick leave":
                        employee.sick_leave = (employee.sick_leave or 0) + leave_days
                    elif leave_type == "casual leave":
                        employee.casual_leave = (employee.casual_leave or 0) + leave_days
                    elif leave_type in ["earned leave", "privilege leave"]:
                        employee.privilege_leave = (employee.privilege_leave or 0) + leave_days
                    else:
                        return jsonify({
                            "success": False,
                            "message": f"Invalid leave type for balance restoration: {leave.leave_type}"
                        }), 400

                    # Restore EmployeeLeaveBalance
                    from models.leave import EmployeeLeaveBalance
                    from sqlalchemy import func
                    balance = EmployeeLeaveBalance.query.filter(
                        EmployeeLeaveBalance.employee_id == employee.id,
                        func.lower(EmployeeLeaveBalance.leave_type) == leave_type
                    ).first()
                    if balance:
                        balance.available = (balance.available or 0) + leave_days

                # Restore attendance records for dates covered by the leave
                if leave.from_date and leave.to_date:
                    from models.attendance import Attendance
                    from routes.attendance import sync_biometric_to_web_entry
                    from datetime import timedelta

                    curr_date = leave.from_date
                    while curr_date <= leave.to_date:
                        att = Attendance.query.filter_by(
                            user_id=employee.user_id,
                            attendance_date=curr_date
                        ).first()
                        if att and att.status == "Leave":
                            has_punches = bool(att.check_in or att.card_check_in or att.check_out or att.card_check_out)
                            if not has_punches:
                                db.session.delete(att)
                            else:
                                att.status = "Absent"
                                att.leave_type = None
                                sync_biometric_to_web_entry(att)
                        curr_date += timedelta(days=1)

                # Check if there is a related split request and cancel it too
                from datetime import timedelta
                related_split = None
                if leave.reason and " (Loss of Pay split)" in leave.reason:
                    paid_reason = leave.reason.replace(" (Loss of Pay split)", "")
                    related_split = LeaveRequest.query.filter(
                        LeaveRequest.employee_id == leave.employee_id,
                        LeaveRequest.to_date == leave.from_date - timedelta(days=1),
                        LeaveRequest.reason == paid_reason,
                        LeaveRequest.status.in_(["Approved", "Pending"])
                    ).first()
                elif leave.reason:
                    lop_reason = leave.reason + " (Loss of Pay split)"
                    related_split = LeaveRequest.query.filter(
                        LeaveRequest.employee_id == leave.employee_id,
                        LeaveRequest.from_date == leave.to_date + timedelta(days=1),
                        LeaveRequest.reason == lop_reason,
                        LeaveRequest.status.in_(["Approved", "Pending"])
                    ).first()

                if related_split:
                    prev_status_related = related_split.status
                    related_split.status = "Cancelled"
                    related_split.cancelled_by = "Employee"
                    related_split.cancelled_at = datetime.now(ZoneInfo("Asia/Kolkata"))
                    related_split.cancellation_reason = leave.cancellation_reason

                    # Restore balance for the related split request if it was Approved
                    if prev_status_related == "Approved" and related_split.request_type == "Leave":
                        related_type = (related_split.leave_type or "").strip().lower()
                        related_days = related_split.total_days or 0

                        if "loss of pay" not in related_type and "lop" not in related_type:
                            if related_type == "sick leave":
                                employee.sick_leave = (employee.sick_leave or 0) + related_days
                            elif related_type == "casual leave":
                                employee.casual_leave = (employee.casual_leave or 0) + related_days
                            elif related_type in ["earned leave", "privilege leave"]:
                                employee.privilege_leave = (employee.privilege_leave or 0) + related_days

                            # Restore EmployeeLeaveBalance
                            from models.leave import EmployeeLeaveBalance
                            from sqlalchemy import func
                            bal_related = EmployeeLeaveBalance.query.filter(
                                EmployeeLeaveBalance.employee_id == employee.id,
                                func.lower(EmployeeLeaveBalance.leave_type) == related_type
                            ).first()
                            if bal_related:
                                bal_related.available = (bal_related.available or 0) + related_days

                    # Restore attendance records for the related split request
                    if related_split.from_date and related_split.to_date:
                        curr_date = related_split.from_date
                        while curr_date <= related_split.to_date:
                            att = Attendance.query.filter_by(
                                user_id=employee.user_id,
                                attendance_date=curr_date
                            ).first()
                            if att and att.status == "Leave":
                                has_punches = bool(att.check_in or att.card_check_in or att.check_out or att.card_check_out)
                                if not has_punches:
                                    db.session.delete(att)
                                else:
                                    att.status = "Absent"
                                    att.leave_type = None
                                    sync_biometric_to_web_entry(att)
                            curr_date += timedelta(days=1)

                    # Emit SocketIO update for the related request
                    try:
                        from extensions import socketio
                        socketio.emit("leave_update", serialize_leave(related_split))
                    except Exception as socket_err:
                        print("Failed to emit related leave socket:", str(socket_err))

        # 6. Create manager notification
        if leave.request_type == "Permission":
            date_str = str(leave.permission_date)
            msg = f"{leave.employee_name} has cancelled the {previous_status.lower()} permission scheduled for {date_str}."
        else:
            from_date_str = str(leave.from_date)
            to_date_str = str(leave.to_date)
            msg = f"{leave.employee_name} has cancelled the {previous_status.lower()} leave scheduled from {from_date_str} to {to_date_str}."

        receiver_name = leave.reporting_manager or (employee.reporting_manager if employee else None) or "Admin"

        notification = Notification(
            receiver_name=receiver_name,
            title="Leave Cancelled",
            message=msg,
            related_id=leave.id,
            related_type="Leave",
            notification_type="Leave Cancellation",
            status="Pending",
            action_required=False
        )
        db.session.add(notification)

        # 7. Create audit log
        audit_msg = f"Employee {leave.employee_name} cancelled Leave Request #{leave.id}."
        audit_log = LeaveAuditLog(
            leave_id=leave.id,
            employee_name=leave.employee_name,
            action=audit_msg,
            previous_status=previous_status,
            new_status="Cancelled",
            cancelled_at=datetime.now(ZoneInfo("Asia/Kolkata")),
            cancelled_by="Employee"
        )
        db.session.add(audit_log)

        db.session.commit()

        # Emit leave_update socket event for real-time dashboard updates
        try:
            from extensions import socketio
            socketio.emit("leave_update", serialize_leave(leave))
        except Exception as socket_err:
            print("Failed to emit leave socket:", str(socket_err))

        return jsonify({
            "success": True,
            "message": "Your leave has been cancelled successfully."
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@leave_bp.route(
    "/export-leave-report",
    methods=["GET"]
)
def export_leave_report():

    try:

        from io import BytesIO

        from openpyxl import Workbook
        from datetime import date

        from openpyxl.styles import (
            PatternFill,
            Font,
            Border,
            Side,
            Alignment
        )

        from openpyxl.utils import (
            get_column_letter
        )

        wb = Workbook()

        ws = wb.active

        ws.title = "Leave Working Update"

        # =========================
        # STYLES
        # =========================

        blue_fill = PatternFill(
            "solid",
            fgColor="4F81BD"
        )

        green_fill = PatternFill(
            "solid",
            fgColor="00E5C3"
        )

        yellow_fill = PatternFill(
            "solid",
            fgColor="FFD966"
        )

        black_fill = PatternFill(
            "solid",
            fgColor="000000"
        )

        orange_fill = PatternFill(
            "solid",
            fgColor="F4B183"
        )

        white_font = Font(
            bold=True,
            color="FFFFFF"
        )

        bold_font = Font(
            bold=True
        )

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # =========================
        # TITLE
        # =========================

        ws.merge_cells("A1:T1")

        ws["A1"] = "LEAVE WORKING UPDATE"

        ws["A1"].font = Font(
            bold=True,
            size=16
        )

        ws["A1"].alignment = Alignment(
            horizontal="center"
        )

        # =========================
        # GROUP HEADERS
        # =========================

        ws.merge_cells("F3:H3")
        ws["F3"] = "Opening Balance"

        ws.merge_cells("I3:K3")
        ws["I3"] = "Monthly Credit"

        ws.merge_cells("L3:N3")
        ws["L3"] = "Leaves Deducted"

        ws.merge_cells("P3:R3")
        ws["P3"] = "Closing Balance"

        for cell in [
            "F3",
            "I3",
            "L3",
            "P3"
        ]:

            ws[cell].fill = black_fill
            ws[cell].font = white_font
            ws[cell].alignment = Alignment(
                horizontal="center"
            )

        # =========================
        # COLUMN HEADERS
        # =========================

        headers = [

            "Employee ID",
            "Employee Name",
            "Department",
            "Designation",
            "DOJ",

            "Opening CL",
            "Opening SL",
            "Opening PL",

            "CL Credit",
            "SL Credit",
            "PL Credit",

            "CL Taken",
            "SL Taken",
            "PL Taken",

            "Total Deducted",

            "Closing CL",
            "Closing SL",
            "Closing PL",

            "Total Balance",

            "Remarks"
        ]

        for col_num, header in enumerate(
            headers,
            start=1
        ):

            cell = ws.cell(
                row=4,
                column=col_num
            )

            cell.value = header
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="center"
            )

            if col_num <= 5:
                cell.fill = blue_fill

            elif 6 <= col_num <= 8:
                cell.fill = yellow_fill

            elif 9 <= col_num <= 11:
                cell.fill = green_fill

            elif 12 <= col_num <= 15:
                cell.fill = orange_fill

            elif 16 <= col_num <= 19:
                cell.fill = yellow_fill

        # =========================
        # EMPLOYEE DATA
        # =========================

        employees = [e for e in Employee.query.order_by(
            Employee.first_name
        ).all() if e.is_active != False]

        row = 5

        for employee in employees:

            cl_taken = 0
            sl_taken = 0
            pl_taken = 0

            approved_leaves = LeaveRequest.query.filter(
                LeaveRequest.employee_id == employee.employee_id,
                LeaveRequest.status == "Approved",
                LeaveRequest.request_type == "Leave"
            ).all()

            for leave in approved_leaves:

                leave_type = (
                    leave.leave_type or ""
                ).strip().lower()

                leave_days = (
                    leave.total_days or 0
                )

                if leave_type == "casual leave":
                    cl_taken += leave_days

                elif leave_type == "sick leave":
                    sl_taken += leave_days

                elif leave_type in ["earned leave", "privilege leave"]:
                    pl_taken += leave_days

            current_cl = employee.casual_leave or 0
            current_sl = employee.sick_leave or 0
            current_pl = employee.privilege_leave or 0

            credit_cl = 6.0
            credit_sl = 6.0
            credit_pl = 15.0

            opening_cl = current_cl + cl_taken
            opening_sl = current_sl + sl_taken
            opening_pl = current_pl + pl_taken

            closing_cl = current_cl
            closing_sl = current_sl
            closing_pl = current_pl

            total_deducted = (
                cl_taken +
                sl_taken +
                pl_taken
            )

            total_balance = (
                closing_cl +
                closing_sl +
                closing_pl
            )

            ws.cell(row=row, column=1, value=employee.employee_id)
            ws.cell(row=row, column=2, value=f"{employee.first_name} {employee.last_name}")
            ws.cell(row=row, column=3, value=employee.department)
            ws.cell(row=row, column=4, value=employee.designation)
            ws.cell(row=row, column=5, value=employee.joining_date.strftime("%d-%m-%Y") if employee.joining_date else "")

            ws.cell(row=row, column=6, value=opening_cl)
            ws.cell(row=row, column=7, value=opening_sl)
            ws.cell(row=row, column=8, value=opening_pl)

            ws.cell(row=row, column=9, value=credit_cl)
            ws.cell(row=row, column=10, value=credit_sl)
            ws.cell(row=row, column=11, value=credit_pl)

            ws.cell(row=row, column=12, value=cl_taken)
            ws.cell(row=row, column=13, value=sl_taken)
            ws.cell(row=row, column=14, value=pl_taken)

            ws.cell(row=row, column=15, value=total_deducted)

            ws.cell(row=row, column=16, value=closing_cl)
            ws.cell(row=row, column=17, value=closing_sl)
            ws.cell(row=row, column=18, value=closing_pl)

            ws.cell(row=row, column=19, value=total_balance)

            ws.cell(row=row, column=20, value="")

            for col in range(1, 21):

                c = ws.cell(
                    row=row,
                    column=col
                )
                c.border = thin_border
                val = c.value
                if val is not None and val != "":
                    if isinstance(val, (int, float, date, datetime)) or (isinstance(val, str) and (val.isdigit() or val.startswith("EMP"))):
                        c.alignment = Alignment(horizontal="center")

            row += 1

        # =========================
        # AUTO WIDTH
        # =========================

        for column_cells in ws.columns:

            length = max(
                len(str(cell.value))
                if cell.value
                else 0
                for cell in column_cells
            )

            ws.column_dimensions[
                get_column_letter(
                    column_cells[0].column
                )
            ].width = length + 5

        # =========================
        # FILTER
        # =========================

        ws.auto_filter.ref = f"A4:T{row}"

        # =========================
        # DOWNLOAD
        # =========================

        excel_file = BytesIO()

        wb.save(excel_file)

        excel_file.seek(0)

        return send_file(
            excel_file,
            as_attachment=True,
            download_name="Leave_Working_Update.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@leave_bp.route(
    "/credit-monthly-leaves",
    methods=["POST"]
)
def credit_monthly_leaves():

    try:

        result = (
            update_all_employee_leave_balances()
        )

        return jsonify(result)

    except Exception as e:

        import traceback
        traceback.print_exc()
        db.session.rollback()

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@leave_bp.route("/policies", methods=["GET"])
def get_leave_policies():
    try:
        from models.leave import LeavePolicy
        policies = LeavePolicy.query.all()
        return jsonify([
            {
                "id": p.id,
                "leave_type": p.leave_type,
                "yearly_limit": p.yearly_limit,
                "applicable_gender": p.applicable_gender
            }
            for p in policies
        ]), 200
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@leave_bp.route("/policies", methods=["POST"])
def create_leave_policy():
    try:
        data = request.get_json() or {}
        leave_type = data.get("leave_type")
        yearly_limit = float(data.get("yearly_limit", 0.0))
        applicable_gender = data.get("applicable_gender", "All")

        if not leave_type:
            return jsonify({"success": False, "error": "Leave type is required"}), 400

        from models.leave import LeavePolicy, EmployeeLeaveBalance
        from models.employee import Employee

        exists = LeavePolicy.query.filter_by(leave_type=leave_type).first()
        if exists:
            return jsonify({"success": False, "error": "Leave policy already exists"}), 400

        policy = LeavePolicy(leave_type=leave_type, yearly_limit=yearly_limit, applicable_gender=applicable_gender)
        db.session.add(policy)
        db.session.commit()

        employees = [e for e in get_all_employees_cached() if e.is_active != False]
        for emp in employees:
            emp_gender = (emp.gender or "").strip().lower()
            pol_gender = applicable_gender.strip().lower()
            is_applicable = (pol_gender == "all") or (emp_gender == pol_gender)

            if is_applicable:
                balance = EmployeeLeaveBalance(
                    employee_id=emp.id,
                    leave_type=leave_type,
                    available=yearly_limit
                )
                db.session.add(balance)
        db.session.commit()

        return jsonify({
            "success": True,
            "message": "Leave category created successfully",
            "policy": {
                "id": policy.id,
                "leave_type": policy.leave_type,
                "yearly_limit": policy.yearly_limit,
                "applicable_gender": policy.applicable_gender
            }
        }), 201
    except Exception as e:
        import traceback
        traceback.print_exc()
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500

@leave_bp.route("/policies", methods=["PUT"])
def update_leave_policies():
    try:
        data = request.get_json() or {}
        if not isinstance(data, list):
            return jsonify({"success": False, "error": "Invalid format, list of policies expected"}), 400

        from models.leave import LeavePolicy, EmployeeLeaveBalance
        for item in data:
            policy_id = item.get("id")
            new_limit = float(item.get("yearly_limit", 0.0))
            policy = LeavePolicy.query.get(policy_id)
            if policy:
                old_limit = policy.yearly_limit
                diff = new_limit - old_limit
                policy.yearly_limit = new_limit

                # Adjust each employee's available balance by the difference
                if diff != 0:
                    balances = EmployeeLeaveBalance.query.filter_by(leave_type=policy.leave_type).all()
                    for bal in balances:
                        bal.available = max(0.0, (bal.available or 0.0) + diff)

        db.session.commit()
        return jsonify({"success": True, "message": "Leave policies updated successfully"}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500

@leave_bp.route("/policies/<int:policy_id>", methods=["DELETE"])
def delete_leave_policy(policy_id):
    try:
        from models.leave import LeavePolicy, EmployeeLeaveBalance
        policy = LeavePolicy.query.get(policy_id)
        if not policy:
            return jsonify({"success": False, "error": "Policy not found"}), 404

        db.session.delete(policy)
        EmployeeLeaveBalance.query.filter_by(leave_type=policy.leave_type).delete()
        db.session.commit()

        return jsonify({"success": True, "message": "Leave category deleted successfully"}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500

@leave_bp.route("/balances/<int:employee_id>", methods=["GET"])
def get_employee_balances(employee_id):
    try:
        from models.leave import EmployeeLeaveBalance
        balances = EmployeeLeaveBalance.query.filter_by(employee_id=employee_id).all()
        
        # Clean up duplicates in database if they exist
        seen_types = {}
        duplicates_to_delete = []
        for bal in balances:
            lt_lower = (bal.leave_type or "").strip().lower()
            if not lt_lower:
                continue
            if lt_lower in seen_types:
                existing = seen_types[lt_lower]
                # Keep the one with the higher available balance, delete the other duplicate
                if bal.available > existing.available:
                    duplicates_to_delete.append(existing)
                    seen_types[lt_lower] = bal
                else:
                    duplicates_to_delete.append(bal)
            else:
                seen_types[lt_lower] = bal
        
        if duplicates_to_delete:
            for dup in duplicates_to_delete:
                db.session.delete(dup)
            db.session.commit()
            # Reload balances after deduplication
            balances = EmployeeLeaveBalance.query.filter_by(employee_id=employee_id).all()

        if not balances:
            from models.leave import LeavePolicy
            from models.employee import Employee
            emp = Employee.query.get(employee_id)
            if emp:
                policies = LeavePolicy.query.all()
                for pol in policies:
                    avail = pol.yearly_limit
                    if pol.leave_type == "Sick Leave" and emp.sick_leave is not None:
                        avail = emp.sick_leave
                    elif pol.leave_type == "Casual Leave" and emp.casual_leave is not None:
                        avail = emp.casual_leave
                    elif pol.leave_type == "Privilege Leave" and emp.privilege_leave is not None:
                        avail = emp.privilege_leave
                    
                    bal = EmployeeLeaveBalance(
                        employee_id=emp.id,
                        leave_type=pol.leave_type,
                        available=avail
                    )
                    db.session.add(bal)
                db.session.commit()
                balances = EmployeeLeaveBalance.query.filter_by(employee_id=employee_id).all()

        return jsonify([
            {
                "id": b.id,
                "leave_type": b.leave_type,
                "available": b.available
            }
            for b in balances
        ]), 200
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@leave_bp.route("/balances/<int:employee_id>", methods=["PUT"])
def update_employee_balances(employee_id):
    try:
        from models.leave import EmployeeLeaveBalance
        data = request.get_json()
        
        for leave_type, available in data.items():
            balance = EmployeeLeaveBalance.query.filter_by(
                employee_id=employee_id,
                leave_type=leave_type
            ).first()
            if balance:
                balance.available = float(available)
            else:
                balance = EmployeeLeaveBalance(
                    employee_id=employee_id,
                    leave_type=leave_type,
                    available=float(available)
                )
                db.session.add(balance)
        
        db.session.commit()
        return jsonify({"success": True, "message": "Leave balances updated successfully"}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500

@leave_bp.route("/resolve-absent", methods=["POST"])
def resolve_absent():
    try:
        from models.employee import Employee
        from models.attendance import Attendance
        from models.leave import EmployeeLeaveBalance, LeaveRequest
        
        data = request.get_json()
        employee_id = data.get("employee_id")
        date_str = data.get("date")
        action = data.get("action")
        reason = data.get("reason", "").strip() or "Applied from Attendance Calendar"
        
        employee = Employee.query.get(employee_id)
        if not employee:
            return jsonify({"success": False, "error": "Employee not found"}), 404
            
        target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        
        att = Attendance.query.filter_by(
            user_id=employee.user_id,
            attendance_date=target_date
        ).first()
        
        now = datetime.utcnow()
        
        if action == "LOP":
            if not att:
                att = Attendance(
                    user_id=employee.user_id,
                    attendance_date=target_date,
                    status="Absent",
                    is_lop=True
                )
                db.session.add(att)
            else:
                att.status = "Absent"
                att.is_lop = True
                
            leave = LeaveRequest(
                employee_id=employee.id,
                employee_name=f"{employee.first_name} {employee.last_name}".strip(),
                request_type="Leave",
                leave_type="Loss of Pay",
                from_date=target_date,
                to_date=target_date,
                total_days=1.0,
                status="Approved",
                reason=reason,
                reporting_manager=employee.reporting_manager,
                approved_by="Auto Approved",
                created_at=now,
                approved_at=now
            )
            db.session.add(leave)
            db.session.commit()
            
            try:
                from extensions import socketio
                socketio.emit("leave_update", {
                    "id": leave.id,
                    "employee_id": employee.employee_id,
                    "employee_name": leave.employee_name,
                    "leave_type": leave.leave_type,
                    "from_date": str(leave.from_date),
                    "to_date": str(leave.to_date),
                    "total_days": leave.total_days,
                    "status": leave.status
                })
            except Exception as se:
                print("Resolve absent socket emit failed for LOP:", se)
                
            return jsonify({"success": True, "message": "Marked as Loss of Pay (LOP) and leave request recorded"}), 200
            
        else:
            leave_type = action.strip()
            
            from sqlalchemy import func
            balance = EmployeeLeaveBalance.query.filter(
                EmployeeLeaveBalance.employee_id == employee.id,
                func.lower(EmployeeLeaveBalance.leave_type) == leave_type.lower()
            ).first()
            
            # Fallback for combined CL/SL policy
            if not balance and leave_type.lower() in ["sick leave", "casual leave"]:
                balance = EmployeeLeaveBalance.query.filter(
                    EmployeeLeaveBalance.employee_id == employee.id,
                    func.lower(EmployeeLeaveBalance.leave_type).in_(["cl/sl", "cl / sl", "sl/cl", "sl / cl"])
                ).first()
                if balance:
                    leave_type = balance.leave_type
            
            if not balance or balance.available < 1.0:
                return jsonify({"success": False, "error": "No leave balance available"}), 400
            
            # Deduct 1 day from the leave balance immediately (auto-approve)
            balance.available = max(balance.available - 1.0, 0.0)

            # Create the leave request as already Approved
            leave = LeaveRequest(
                employee_id=employee.id,
                employee_name=f"{employee.first_name} {employee.last_name}".strip(),
                request_type="Leave",
                leave_type=leave_type,
                from_date=target_date,
                to_date=target_date,
                total_days=1.0,
                status="Approved",
                reason=reason,
                reporting_manager=employee.reporting_manager,
                approved_by="Auto Approved",
                created_at=now,
                approved_at=now
            )
            db.session.add(leave)

            # Update the attendance record to reflect the leave
            if not att:
                att = Attendance(
                    user_id=employee.user_id,
                    attendance_date=target_date,
                    status="Leave",
                    leave_type=leave_type,
                    check_in=None,
                    check_out=None,
                    total_hours=0.0
                )
                db.session.add(att)
            else:
                att.status = "Leave"
                att.leave_type = leave_type
                att.is_lop = False
                att.check_in = None
                att.check_out = None
                att.total_hours = 0.0

            db.session.commit()
            
            try:
                from extensions import socketio
                socketio.emit("leave_update", {
                    "id": leave.id,
                    "employee_id": employee.employee_id,
                    "employee_name": leave.employee_name,
                    "leave_type": leave.leave_type,
                    "from_date": str(leave.from_date),
                    "to_date": str(leave.to_date),
                    "total_days": leave.total_days,
                    "status": leave.status
                })
            except Exception as se:
                print("Resolve absent socket emit failed:", se)
                
            return jsonify({"success": True, "message": f"{leave_type} applied and approved automatically."}), 200
            
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500